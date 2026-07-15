"""Shared helpers for the M6 benchmark tests: build extraction workbooks,
usage CSVs, manifests, and ground-truth files inside pytest tmp_path. All
offline - no PDFs, no provider calls, no real data."""

import json
from pathlib import Path

import openpyxl

from invoice_extractor.benchmark.report import REPORT_SHEETS  # noqa: F401 (re-export)
from invoice_extractor.excel_export import (
    INVOICE_COLUMNS,
    LINE_ITEM_COLUMNS,
    NEEDS_REVIEW_COLUMNS,
)


def write_workbook(path: Path, invoices: list[dict], line_items: list[dict],
                   needs_review: list[dict] | None = None) -> Path:
    """Write a 3-sheet extraction-style workbook from raw row dicts (bypassing
    export_workbook so tests can craft duplicates / extras / partial rows)."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("Invoices")
    ws.append(INVOICE_COLUMNS)
    for row in invoices:
        ws.append([row.get(c) for c in INVOICE_COLUMNS])
    ws = wb.create_sheet("LineItems")
    ws.append(LINE_ITEM_COLUMNS)
    for row in line_items:
        ws.append([row.get(c) for c in LINE_ITEM_COLUMNS])
    ws = wb.create_sheet("NeedsReview")
    ws.append(NEEDS_REVIEW_COLUMNS)
    for row in (needs_review or []):
        ws.append([row.get(c) for c in NEEDS_REVIEW_COLUMNS])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)
    return path


def invoice_row(invoice_id, source_file, *, needs_review=False, review_reason=None,
                extraction_method="text", provider="openrouter", model="m1", **fields) -> dict:
    row = {"invoice_id": invoice_id, "source_file": source_file,
           "needs_review": needs_review, "review_reason": review_reason,
           "extraction_method": extraction_method, "provider": provider, "model": model}
    row.update(fields)
    return row


def line_row(invoice_id, source_file, line_number, **fields) -> dict:
    row = {"invoice_id": invoice_id, "source_file": source_file, "line_number": line_number}
    row.update(fields)
    return row


def usage_csv(path: Path, records: list[dict]) -> Path:
    """Write a usage CSV with the production column order."""
    from invoice_extractor.usage import USAGE_CSV_COLUMNS
    import csv
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(USAGE_CSV_COLUMNS)
        for r in records:
            w.writerow([r.get(c, "") for c in USAGE_CSV_COLUMNS])
    return path


def usage_record(source_file, *, route="text", page_range="1", attempt_type="primary",
                 requested_model="req-m", actual_model="act-m", cost_usd="0.0002",
                 accepted="True", **fields) -> dict:
    rec = {"run_id": "run-1", "source_file": source_file, "route": route,
           "page_range": page_range, "attempt_type": attempt_type, "ladder_index": "0",
           "requested_model": requested_model, "actual_model": actual_model,
           "structured_mode": "json_schema", "input_tokens": "100", "output_tokens": "50",
           "reasoning_tokens": "5", "total_tokens": "150", "cost_usd": cost_usd,
           "finish_reason": "stop", "native_finish_reason": "STOP", "generation_id": "gen-1",
           "latency_ms": "1000", "accepted": accepted, "rejection_category": "",
           "http_status": ""}
    rec.update(fields)
    return rec


def write_manifest(dir_path: Path, cases: list[dict], ground_truths: dict[str, dict],
                   thresholds: dict | None = None) -> Path:
    """cases: manifest entries (without ground_truth path). ground_truths:
    case_id -> GT dict. Writes GT files + manifest, returns the manifest path."""
    gt_dir = dir_path / "ground_truth"
    gt_dir.mkdir(parents=True, exist_ok=True)
    entries = []
    for c in cases:
        cid = c["case_id"]
        gt_rel = f"ground_truth/{cid}.json"
        (dir_path / gt_rel).write_text(json.dumps(ground_truths[cid]), encoding="utf-8")
        entries.append({**c, "ground_truth": gt_rel})
    manifest = {"cases": entries}
    if thresholds is not None:
        manifest["thresholds"] = thresholds
    path = dir_path / "manifest.json"
    path.write_text(json.dumps(manifest), encoding="utf-8")
    return path


def gt(case_id, *, invoice=None, line_items=None, expected_needs_review=False,
       ignored_fields=None, accepted_review_categories=None, field_tolerances=None,
       notes="") -> dict:
    return {
        "case_id": case_id, "invoice": invoice or {}, "line_items": line_items or [],
        "expected_needs_review": expected_needs_review,
        "ignored_fields": ignored_fields or [],
        "accepted_review_categories": accepted_review_categories or [],
        "field_tolerances": field_tolerances or {}, "notes": notes,
    }


def manifest_entry(case_id, source_file, document_type="text_single_page",
                   expected_outcome="extracted") -> dict:
    return {"case_id": case_id, "source_file": source_file,
            "document_type": document_type, "expected_outcome": expected_outcome}

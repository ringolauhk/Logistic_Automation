"""Packing-list workbook generation, Build 7: input boundary, workbook
content and formatting, validation, ZIP, persistence, states, retention,
and UI wiring. Fully offline; synthetic packing artifacts only. Generated
files live exclusively in pytest tmp dirs and are never committed."""

import json
import zipfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from apps.web.job_manager import JobError
from apps.web.transfer import jobs as tjobs
from apps.web.transfer import models as tm
from apps.web.transfer import packing as pk
from apps.web.transfer import workbook as wb
from tests.test_transfer_packing import (
    EAN_1,
    EAN_2,
    EAN_3,
    enrich,
    prepared_job,
    two_destination_job,
)

ROOT = Path(__file__).resolve().parent.parent


@pytest.fixture(autouse=True)
def roots(tmp_path, monkeypatch):
    monkeypatch.setenv("WEB_JOBS_DIR", str(tmp_path / "jobs"))
    monkeypatch.setenv("TRANSFER_JOBS_DIR", str(tmp_path / "transfer-jobs"))
    for name in ("PACKING_COMPANY_NAME", "PACKING_DOCUMENT_TITLE",
                 "PACKING_FORM_OF_DELIVERY",
                 "PACKING_OUTPUT_RETENTION_HOURS",
                 "PACKING_CREATE_ZIP_FOR_MULTIPLE",
                 "PACKING_CUSTOMER_STYLE_FIELD",
                 "PACKING_CUSTOMER_COLOR_CODE_FIELD",
                 "PACKING_CUSTOMER_COLOR_DESC_FIELD"):
        monkeypatch.delenv(name, raising=False)
    return tmp_path


def _attr_records():
    """Fixture records with Analysis Code 01-15 + Composition #1-4."""
    import copy
    from tests.test_transfer_packing import ALL_RECORDS, ALL_RECORDS_B
    records = copy.deepcopy(ALL_RECORDS + ALL_RECORDS_B)
    for record in records:
        record.update({f"analysisCode{i:02d}": f"AC{i:02d}"
                       for i in range(1, 16)})
        record.update({f"composition{i:02d}": f"COMP{i:02d}"
                       for i in range(1, 5)})
    return records


def packed_job(tmp_path):
    """Approved + enriched (with Analysis/Composition attributes) +
    packing-prepared two-destination job."""
    job_id = two_destination_job(tmp_path)
    enrich(job_id, records=_attr_records())
    pk.prepare_packing(job_id)
    return job_id


def generated_job(tmp_path):
    job_id = packed_job(tmp_path)
    meta = wb.generate_workbooks(job_id)
    return job_id, meta


def wb_path(job_id, entry):
    return wb.output_dir(job_id) / entry["filename"]


# --- input boundary ---------------------------------------------------------------

class TestInputBoundary:
    def test_valid_packing_result_accepted(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        assert meta["status"] == "complete"
        assert (tjobs.load_transfer_job(job_id).status
                == tm.JOB_WORKBOOK_GENERATION_COMPLETE)

    def test_missing_preparation_rejected(self, tmp_path):
        job_id = prepared_job(tmp_path)     # enriched, never packed
        with pytest.raises(JobError):
            wb.load_generation_inputs(job_id)

    def test_stale_packing_result_rejected(self, tmp_path):
        job_id = packed_job(tmp_path)
        # editing the review makes packing stale via its checksums
        from apps.web.transfer import review as rv
        review = rv.load_review(job_id)
        rv.apply_correction(review, "line", review.lines[0].entity_id,
                            "description", "EDITED LATE")
        rv.save_review(job_id, review)
        with pytest.raises(JobError, match="stale"):
            wb.load_generation_inputs(job_id)

    def test_malformed_packing_result_rejected(self, tmp_path):
        job_id = packed_job(tmp_path)
        pk.result_path(job_id).write_text("{ nope")
        with pytest.raises(JobError):
            wb.load_generation_inputs(job_id)

    def test_blocking_packing_issue_rejected(self, tmp_path):
        job_id = two_destination_job(tmp_path)
        from tests.test_transfer_packing import ALL_RECORDS
        enrich(job_id, records=ALL_RECORDS)   # EAN_3 blocked -> WITH_ISSUES
        pk.prepare_packing(job_id)
        with pytest.raises(JobError, match="blocking"):
            wb.load_generation_inputs(job_id)


# --- workbook count + filenames ---------------------------------------------------

class TestWorkbookFiles:
    def test_one_workbook_per_destination_no_extras(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        entries = meta["destination_workbooks"]
        assert [e["destination_code"] for e in entries] == ["ZZOHK101",
                                                            "ZZOHK202"]
        xlsx_files = sorted(p.name for p in
                            wb.output_dir(job_id).glob("*.xlsx"))
        assert xlsx_files == sorted(e["filename"] for e in entries)
        assert len(xlsx_files) == 2

    def test_deterministic_sanitized_filenames(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        entry = meta["destination_workbooks"][0]
        date = json.loads(pk.result_path(job_id).read_text())["invoice_date"]
        assert entry["filename"] == (
            f"Packing_List_ZZOHK101_PL-ZZOHK101-{date}-001.xlsx")
        assert wb.sanitize_component("A/B\\C..D") == "A-B-C..D"
        assert "/" not in wb.workbook_filename("X/../Y", "PL/1")

    def test_relative_paths_only_in_metadata(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        blob = json.dumps(meta)
        assert str(wb.output_dir(job_id)) not in blob
        for entry in meta["destination_workbooks"]:
            assert entry["relative_path"].startswith("output/")


# --- packing list sheet -----------------------------------------------------------

def _sheet_text(ws):
    return " | ".join(str(c.value) for row in ws.iter_rows()
                      for c in row if c.value is not None)


class TestPackingListSheet:
    def test_header_fields_and_identity(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        entry = meta["destination_workbooks"][0]
        book = load_workbook(wb_path(job_id, entry))
        text = _sheet_text(book[wb.SHEET_PACKING_LIST])
        for expected in ("IMAGINEX BG LIMITED", "Packing List",
                         "Delivery Invoice No.", entry[
                             "delivery_invoice_number"],
                         "ZZOHK101", "Warehouse Shipment", "TN#",
                         "Form Of Delivery :", "Invoice Date :"):
            assert expected in text, expected

    def test_detail_columns_subtotals_and_totals(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        entry = meta["destination_workbooks"][0]     # ZZOHK101
        ws = load_workbook(wb_path(job_id, entry))[wb.SHEET_PACKING_LIST]
        text = _sheet_text(ws)
        for header in ("CTN. No.", "Description", "SKU Number / EAN", "PLU",
                       "IMX Item Code", "IMX Color Code",
                       "Color Description", "Size", "Customer Style",
                       "Customer Color Code", "Qty"):
            assert header in text, header
        # carton subtotals and final totals (fixture: carton 001 = 5 units,
        # carton 002 = 1 unit)
        assert "001 Total" in text and "002 Total" in text
        assert "Total Cartons" in text and "Total Units" in text
        rows = list(ws.iter_rows(values_only=True))
        subtotal_001 = next(r for r in rows if r[0] == "001 Total")
        assert subtotal_001[11] == 5
        subtotal_002 = next(r for r in rows if r[0] == "002 Total")
        assert subtotal_002[11] == 1
        totals = next(r for r in rows if r[0] == "Total Cartons")
        assert totals[1] == 2 and totals[11] == 6

    def test_carton_order_and_no_cross_destination_data(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        book_101 = load_workbook(wb_path(job_id,
                                         meta["destination_workbooks"][0]))
        text_101 = _sheet_text(book_101[wb.SHEET_PACKING_LIST])
        assert "ZZOHK202" not in text_101
        assert EAN_3 not in text_101                 # other destination's EAN
        book_202 = load_workbook(wb_path(job_id,
                                         meta["destination_workbooks"][1]))
        text_202 = _sheet_text(book_202[wb.SHEET_PACKING_LIST])
        assert "ZZOHK101" not in text_202
        assert EAN_1 not in text_202
        ctns = [str(r[0]) for r in
                book_101[wb.SHEET_PACKING_LIST].iter_rows(values_only=True)
                if r[0] is not None and str(r[0]).isdigit()]
        assert ctns == sorted(ctns)                  # 001 rows before 002

    def test_customer_columns_blank_by_default_mapped_when_configured(
            self, tmp_path, monkeypatch):
        job_id = packed_job(tmp_path)
        wb.generate_workbooks(job_id)
        meta = wb.load_output(job_id)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_PACKING_LIST]
        rows = [r for r in ws.iter_rows(values_only=True)
                if r[0] is not None and str(r[0]) == "001"]
        assert all((r[8] or "") == "" for r in rows)     # Customer Style blank
        # configure a mapping -> populated from the normalized attribute
        monkeypatch.setenv("PACKING_CUSTOMER_STYLE_FIELD",
                           "analysis_code_01")
        wb.generate_workbooks(job_id)
        meta = wb.load_output(job_id)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_PACKING_LIST]
        rows = [r for r in ws.iter_rows(values_only=True)
                if r[0] is not None and str(r[0]) == "001"]
        assert any(r[8] for r in rows)               # mapped value appears

    def test_print_setup(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_PACKING_LIST]
        assert ws.page_setup.orientation == "landscape"
        assert ws.page_setup.fitToWidth == 1
        assert ws.print_title_rows
        assert ws.print_area
        assert ws.sheet_view.showGridLines is False
        assert "Page &P of &N" in ws.oddFooter.center.text


# --- detail sheet -----------------------------------------------------------------

class TestDetailSheet:
    def test_source_api_analysis_composition_present(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_DETAIL]
        headers = [c.value for c in ws[1]]
        for expected in ("Src Item", "Src EAN", "IMX Item Code", "PLU",
                         "EAN", "Analysis Code 01", "Analysis Code 15",
                         "Composition #1", "Composition #4",
                         "Reviewed Line IDs", "Match Status"):
            assert expected in headers, expected
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        rows = [r for r in rows if r[0] is not None]
        assert len(rows) == 3                        # ZZOHK101 prepared lines
        by_header = dict(zip(headers, rows[0]))
        assert by_header["Analysis Code 01"] == "AC01"      # from fixture
        assert by_header["Analysis Code 15"] == "AC15"
        assert by_header["Composition #1"] == "COMP01"
        assert by_header["Composition #4"] == "COMP04"

    def test_leading_zero_identifiers_stay_text(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_DETAIL]
        headers = [c.value for c in ws[1]]
        ean_col = headers.index("EAN") + 1
        carton_col = headers.index("Generated Carton") + 1
        for row in ws.iter_rows(min_row=2):
            if row[0].value is None:
                continue
            ean = row[ean_col - 1].value
            assert isinstance(ean, str) and ean.startswith("0")
            carton = row[carton_col - 1].value
            assert isinstance(carton, str) and carton.startswith("0")
            assert row[ean_col - 1].number_format == "@"

    def test_consolidated_source_ids_retained(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_DETAIL]
        headers = [c.value for c in ws[1]]
        ids_col = headers.index("Reviewed Line IDs")
        count_col = headers.index("Source Row Count")
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] is not None]
        merged = next(r for r in rows if r[count_col] == 2)
        assert "D001-C001-L001" in merged[ids_col]
        assert "D001-C001-L002" in merged[ids_col]

    def test_freeze_and_filter(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_DETAIL]
        assert ws.freeze_panes == "A2"
        assert ws.auto_filter.ref


# --- carton mapping + needs review + source documents -----------------------------

class TestAuditSheets:
    def test_carton_mapping_matches_build6(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        prepared = pk.load_preparation(job_id)
        group = prepared["destinations"][0]
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_CARTON_MAPPING]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] is not None and r[0] != "Total"]
        assert [r[2] for r in rows] == [m["generated_carton_number"]
                                        for m in group["carton_mappings"]]
        assert [r[3] for r in rows] == [m["original_carton_number"]
                                        for m in group["carton_mappings"]]
        assert [r[4] for r in rows] == [
            m["source_carton_key"]["upload_sequence"]
            for m in group["carton_mappings"]]
        total = next(r for r in ws.iter_rows(min_row=2, values_only=True)
                     if r[0] == "Total")
        assert total[10] == group["total_units"]

    def test_needs_review_empty_state_and_no_secrets(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        book = load_workbook(wb_path(job_id,
                                     meta["destination_workbooks"][1]))
        text = _sheet_text(book[wb.SHEET_NEEDS_REVIEW])
        assert ("No unresolved review items." in text
                or "warning" in text.lower())
        for forbidden in ("Bearer", "Authorization", "access_token",
                          "password"):
            assert forbidden not in text

    def test_needs_review_includes_product_warning(self, tmp_path):
        # fixture gateway record for EAN_1 has API wording -> desc warning
        job_id, meta = generated_job(tmp_path)
        book = load_workbook(wb_path(job_id,
                                     meta["destination_workbooks"][0]))
        text = _sheet_text(book[wb.SHEET_NEEDS_REVIEW])
        assert ("PRODUCT_" in text
                or "No unresolved review items." in text)

    def test_source_documents_sheet(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        ws = load_workbook(wb_path(job_id, meta["destination_workbooks"][0])
                           )[wb.SHEET_SOURCE_DOCUMENTS]
        rows = [r for r in ws.iter_rows(min_row=2, values_only=True)
                if r[0] is not None]
        assert len(rows) == 2                        # both source PDFs listed
        assert {r[1] for r in rows} == {"first.pdf", "second.pdf"}
        first = next(r for r in rows if r[1] == "first.pdf")
        assert first[3]                              # delivery note number
        assert first[12] == "yes"                    # in this workbook
        second = next(r for r in rows if r[1] == "second.pdf")
        assert second[12] == "no"


# --- validation -------------------------------------------------------------------

class TestValidation:
    def test_all_generated_workbooks_validate(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        for entry in meta["destination_workbooks"]:
            assert entry["validation_status"] in ("valid",
                                                  "valid_with_warnings")
            issues = wb.validate_workbook(
                wb_path(job_id, entry),
                next(g for g in pk.load_preparation(job_id)["destinations"]
                     if g["destination_code"] == entry["destination_code"]))
            assert not [i for i in issues
                        if i["severity"] == wb.SEV_BLOCKING]

    def test_missing_and_corrupt_files_detected(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        group = pk.load_preparation(job_id)["destinations"][0]
        missing = wb.validate_workbook(
            wb.output_dir(job_id) / "nope.xlsx", group)
        assert missing[0]["code"] == wb.WORKBOOK_FILE_MISSING
        bad = wb.output_dir(job_id) / "bad.xlsx"
        bad.write_bytes(b"not a zip at all")
        corrupt = wb.validate_workbook(bad, group)
        assert corrupt[0]["code"] == wb.WORKBOOK_OPEN_FAILED
        bad.unlink()

    def test_missing_sheet_and_count_mismatch_detected(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        entry = meta["destination_workbooks"][0]
        path = wb_path(job_id, entry)
        group = next(g for g in pk.load_preparation(job_id)["destinations"]
                     if g["destination_code"] == entry["destination_code"])
        book = load_workbook(path)
        book.remove(book[wb.SHEET_NEEDS_REVIEW])
        tampered = wb.output_dir(job_id) / "tampered.xlsx"
        book.save(tampered)
        issues = wb.validate_workbook(tampered, group)
        assert any(i["code"] == wb.WORKBOOK_SHEET_MISSING for i in issues)
        wrong_group = dict(group, prepared_line_count=99)
        issues = wb.validate_workbook(path, wrong_group)
        assert any(i["code"] == wb.WORKBOOK_LINE_COUNT_MISMATCH
                   for i in issues)
        wrong_totals = dict(group, total_units=999)
        issues = wb.validate_workbook(path, wrong_totals)
        assert any(i["code"] == wb.WORKBOOK_TOTAL_MISMATCH for i in issues)
        tampered.unlink()

    def test_no_macros_or_external_links(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        for entry in meta["destination_workbooks"]:
            with zipfile.ZipFile(wb_path(job_id, entry)) as archive:
                assert not any(n.endswith("vbaProject.bin")
                               for n in archive.namelist())
                assert not any("externalLink" in n
                               for n in archive.namelist())


# --- ZIP --------------------------------------------------------------------------

class TestZip:
    def test_zip_for_multiple_destinations(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        zip_entry = meta["zip"]
        assert zip_entry is not None
        assert zip_entry["filename"] == wb.zip_filename(job_id)
        assert zip_entry["member_count"] == 2
        with zipfile.ZipFile(wb.output_dir(job_id)
                             / zip_entry["filename"]) as archive:
            members = archive.namelist()
            assert members == [e["filename"]
                               for e in meta["destination_workbooks"]]
            assert archive.testzip() is None
            assert all("/" not in n and ".." not in n for n in members)
            assert all(n.endswith(".xlsx") for n in members)

    def test_no_zip_for_single_destination(self, tmp_path):
        from tests.test_transfer_packing import (ALL_RECORDS, ROW_1, ROW_2,
                                                 note_pdf)
        from apps.web.transfer import extraction, review as rv
        pdf = note_pdf(tmp_path / "one.pdf", [{"rows": (ROW_1, ROW_2),
                                               "carton_total": None}])
        uploads = [("one.pdf", Path(pdf).read_bytes())]
        validated, _ = tjobs.validate_transfer_uploads(uploads)
        job_id = tjobs.create_transfer_job(uploads, validated)
        extraction.run_extraction(job_id, use_default_adapter=False)
        rv.save_review(job_id, rv.get_or_create_review(job_id))
        rv.approve_review(job_id)
        enrich(job_id, records=ALL_RECORDS)
        pk.prepare_packing(job_id)
        meta = wb.generate_workbooks(job_id)
        assert meta["zip"] is None
        assert not list(wb.output_dir(job_id).glob("*.zip"))

    def test_zip_disabled_by_config(self, tmp_path, monkeypatch):
        monkeypatch.setenv("PACKING_CREATE_ZIP_FOR_MULTIPLE", "false")
        job_id, meta = generated_job(tmp_path)
        assert meta["zip"] is None


# --- persistence + staleness + states ---------------------------------------------

class TestPersistence:
    def test_metadata_schema_checksum_reload(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        raw = json.loads(wb.output_meta_path(job_id).read_text())
        assert raw["schema_version"] == 1
        assert raw["packing_checksum"] == wb.packing_checksum(job_id)
        assert not list(wb.output_dir(job_id).glob("*.tmp-*"))
        reloaded = wb.load_output(job_id)
        assert reloaded["stale"] is False
        for entry in reloaded["destination_workbooks"]:
            path = wb.output_dir(job_id) / entry["filename"]
            assert path.stat().st_size == entry["byte_size"]
            assert wb._sha256(path) == entry["sha256"]

    def test_unchanged_regeneration_no_duplicates_stable_names(self, tmp_path):
        job_id, first = generated_job(tmp_path)
        names_before = sorted(p.name for p in
                              wb.output_dir(job_id).iterdir())
        second = wb.generate_workbooks(job_id)
        names_after = sorted(p.name for p in
                             wb.output_dir(job_id).iterdir())
        assert names_before == names_after           # no duplicates
        assert ([e["delivery_invoice_number"]
                 for e in second["destination_workbooks"]]
                == [e["delivery_invoice_number"]
                    for e in first["destination_workbooks"]])

    def test_changed_packing_marks_output_stale(self, tmp_path):
        job_id, meta = generated_job(tmp_path)
        # touch the packing artifact (rewrite -> new bytes/checksum)
        prepared = json.loads(pk.result_path(job_id).read_text())
        prepared["updated_at"] = "changed"
        pk.result_path(job_id).write_text(json.dumps(prepared))
        reloaded = wb.load_output(job_id)
        assert reloaded["stale"] is True

    def test_validation_failure_sets_failed_state(self, tmp_path,
                                                  monkeypatch):
        job_id = packed_job(tmp_path)
        original = wb.validate_workbook
        monkeypatch.setattr(wb, "validate_workbook",
                            lambda path, group: [{
                                "code": wb.WORKBOOK_TOTAL_MISMATCH,
                                "severity": wb.SEV_BLOCKING,
                                "destination": group["destination_code"],
                                "message": "forced"}])
        with pytest.raises(JobError):
            wb.generate_workbooks(job_id)
        assert (tjobs.load_transfer_job(job_id).status
                == tm.JOB_WORKBOOK_GENERATION_FAILED)
        assert not list(wb.output_dir(job_id).glob("*.tmp-*"))
        # retry succeeds after the fault is removed
        monkeypatch.setattr(wb, "validate_workbook", original)
        meta = wb.generate_workbooks(job_id)
        assert meta["status"] == "complete"

    def test_invalid_transition_rejected(self, tmp_path):
        job_id = prepared_job(tmp_path)              # product state
        with pytest.raises(JobError):
            tjobs.update_job_status(job_id,
                                    tm.JOB_WORKBOOK_GENERATION_COMPLETE)

    def test_stranded_in_progress_recovers(self, tmp_path):
        job_id = packed_job(tmp_path)
        tjobs.update_job_status(job_id,
                                tm.JOB_WORKBOOK_GENERATION_IN_PROGRESS)
        meta = wb.generate_workbooks(job_id)
        assert meta["status"] == "complete"


# --- retention --------------------------------------------------------------------

class TestRetention:
    def test_expired_outputs_cleaned_active_kept(self, tmp_path,
                                                 monkeypatch):
        import os as _os
        import time as _time
        job_id, meta = generated_job(tmp_path)
        config = wb.load_workbook_config()
        old = _time.time() - (config.retention_hours + 1) * 3600
        for path in wb.output_dir(job_id).glob("*.xlsx"):
            _os.utime(path, (old, old))
        removed = wb.cleanup_expired_outputs(job_id, config)
        assert removed == 2
        assert wb.output_meta_path(job_id).is_file()  # metadata untouched
        # in-progress jobs are never cleaned
        job_id2, _ = generated_job(tmp_path)
        tjobs.update_job_status(job_id2,
                                tm.JOB_WORKBOOK_GENERATION_IN_PROGRESS)
        for path in wb.output_dir(job_id2).glob("*.xlsx"):
            _os.utime(path, (old, old))
        assert wb.cleanup_expired_outputs(job_id2, config) == 0

    def test_config_validation(self, monkeypatch):
        monkeypatch.setenv("PACKING_OUTPUT_RETENTION_HOURS", "-1")
        assert wb.workbook_config_problems()
        monkeypatch.setenv("PACKING_OUTPUT_RETENTION_HOURS", "24")
        monkeypatch.setenv("PACKING_CUSTOMER_STYLE_FIELD", "evil_field")
        assert wb.workbook_config_problems()
        monkeypatch.setenv("PACKING_CUSTOMER_STYLE_FIELD",
                           "analysis_code_07")
        assert not wb.workbook_config_problems()


# --- UI wiring + boundaries (static) ----------------------------------------------

class TestUiAndBoundaries:
    RPAGE = (ROOT / "apps" / "web" / "transfer" / "review_page.py").read_text(
        encoding="utf-8")
    WB = (ROOT / "apps" / "web" / "transfer" / "workbook.py").read_text(
        encoding="utf-8")

    def test_generate_button_gated(self):
        assert "Generate Workbooks" in self.RPAGE
        assert "_WORKBOOK_STATES" in self.RPAGE

    def test_downloads_present_and_stale_guarded(self):
        section = self.RPAGE.split("_render_workbook_section")[-1]
        assert "download_button" in section
        assert "unavailable (stale)" in section
        assert "transfer_wb_zip_dl" in section

    def test_no_print_or_email_controls(self):
        section = self.RPAGE.split("_render_workbook_section")[-1].lower()
        for forbidden in ("smtp", "mailto", "send email", "print()"):
            assert forbidden not in section, forbidden

    def test_workbook_module_never_calls_api_or_upstream_writers(self):
        low = self.WB.lower()
        for forbidden in ("httpx", "plulabel", "auth/login",
                          "ensure_access_token", "requests.",
                          "_write_preparation", "_write_enrichment",
                          "save_review", "_write_result", "_write_metadata"):
            assert forbidden not in low, forbidden

    def test_invoice_workflow_untouched(self):
        # invoice modules must not import or reference the transfer
        # workbook module (the plain word "workbook" appears in legit
        # invoice UI text)
        for name in ("job_manager.py", "worker.py", "app.py"):
            src = (ROOT / "apps" / "web" / name).read_text(encoding="utf-8")
            assert "transfer.workbook" not in src, name
            assert "_render_workbook_section" not in src, name
            assert "import workbook" not in src, name

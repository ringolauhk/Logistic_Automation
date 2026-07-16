"""M7: operational hardening - progress logs, safe interruption, overwrite
protection, atomic writes, preflight, no-budget warning, doctor readiness,
log hygiene, and the artifact-set safeguard (tests A-AW).

All offline - the autouse network-blocking fixture guards; provider and
export boundaries are mocked only where a test needs to inject behavior.
"""

import json
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest
from click.testing import CliRunner

from invoice_extractor import cli as cli_module
from invoice_extractor import gemini_client, openrouter_client
from invoice_extractor.atomic import StagedArtifacts
from invoice_extractor.cli import cli

from .conftest import TEXT_BODY, build_pdf, invoice_json


_OR_ENV = [
    "LLM_GATEWAY", "OPENROUTER_API_KEY", "OPENROUTER_TEXT_MODELS",
    "OPENROUTER_VISION_MODELS", "MAX_TEXT_PAGES", "MAX_VISION_PAGES", "MAX_RETRIES",
    "MAX_MODEL_ATTEMPTS_PER_FILE", "MAX_COST_USD_PER_FILE", "MAX_COST_USD_PER_RUN",
    "GEMINI_API_KEY", "ANTHROPIC_API_KEY", "ENABLE_CLAUDE_TEXT_FALLBACK",
    "SAVE_DEBUG_ARTIFACTS", "DEBUG_ARTIFACT_DIR",
]


def _or_env(monkeypatch, **extra):
    for var in _OR_ENV:
        monkeypatch.delenv(var, raising=False)
    monkeypatch.setenv("LLM_GATEWAY", "openrouter")
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-or-key")
    monkeypatch.setenv("OPENROUTER_TEXT_MODELS", "tv/text-1,tv/text-2")
    monkeypatch.setenv("OPENROUTER_VISION_MODELS", "tv/vis-1,tv/vis-2")
    monkeypatch.setenv("MAX_TEXT_PAGES", "2")
    monkeypatch.setenv("MAX_VISION_PAGES", "2")
    monkeypatch.setenv("MAX_RETRIES", "1")
    for k, v in extra.items():
        monkeypatch.setenv(k, v)


def _envelope(content, *, model="served-m", finish_reason="stop", **usage):
    u = {"prompt_tokens": 500, "completion_tokens": 100, "total_tokens": 600,
         "cost": 0.0002, "completion_tokens_details": {"reasoning_tokens": 5}}
    u.update(usage)
    return {"id": "gen-1", "model": model,
            "choices": [{"finish_reason": finish_reason, "native_finish_reason": "STOP",
                         "message": {"content": content}}], "usage": u}


class Recorder:
    def __init__(self, responses):
        self.responses = list(responses)
        self.calls = 0

    def __call__(self, cfg, *, model, messages, response_format=None, max_tokens, timeout=None):
        self.calls += 1
        if not self.responses:
            raise AssertionError("provider called more than expected")
        item = self.responses.pop(0)
        if isinstance(item, BaseException):
            raise item
        return item


def _run(args):
    return CliRunner().invoke(cli, args)


def _text_samples(tmp_path, n=1):
    s = tmp_path / "samples"
    s.mkdir(exist_ok=True)
    for i in range(n):
        build_pdf(s / f"inv{i}.pdf", [("text", TEXT_BODY)])
    return s


# --- A-F: progress logs before provider calls ---------------------------------

class TestProgressLogs:
    def test_a_b_e_f_progress_before_text_and_vision(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = tmp_path / "samples"
        s.mkdir()
        build_pdf(s / "mixed.pdf", [("text", TEXT_BODY), ("image",)])
        rec = Recorder([_envelope(invoice_json()), _envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        out = tmp_path / "out" / "r.xlsx"
        result = _run(["run", "--input", str(s), "--output", str(out),
                       "--log-file", str(tmp_path / "run.log")])
        assert result.exit_code == 0, result.output
        log = (tmp_path / "run.log").read_text()
        # A: before first text call; B: before first vision call.
        assert "text chunk 1/1 pages 1 - starting primary model 1/2" in log
        assert "vision chunk 1/1 pages 2 - starting primary model 1/2" in log
        # E: route/page/model/timeout fields present.
        assert "requested=tv/text-1 timeout=120s" in log
        # F: no prompt/invoice/key/base64/response content.
        for forbidden in ("INVOICE INV-1001", "Ocean freight", "test-or-key",
                          "data:image", "base64", "extraction engine"):
            assert forbidden not in log

    def test_c_progress_before_repair(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope("not valid json"), _envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx"),
                       "--log-file", str(tmp_path / "run.log")])
        assert result.exit_code == 0, result.output
        log = (tmp_path / "run.log").read_text()
        assert "starting repair model 1/2" in log

    def test_d_progress_before_escalation(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        err = openrouter_client.ProviderError("boom", category="rate_limited", http_status=429)
        rec = Recorder([err, _envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx"),
                       "--log-file", str(tmp_path / "run.log")])
        assert result.exit_code == 0, result.output
        log = (tmp_path / "run.log").read_text()
        assert "starting escalation model 2/2" in log


# --- G/H: completion + run summary counts -------------------------------------

class TestSummaryCounts:
    def test_g_h_completion_and_run_summary_counts(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope("not valid json"),          # primary (rejected)
                        _envelope(invoice_json(), cost=0.001)])  # repair (accepted)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0, result.output
        # H: run summary reports provider requests + cost + unknown-cost handling.
        assert "Provider requests:    2 (repair=1 escalation=0)" in result.output
        assert "Reported cost (USD):" in result.output


# --- I/J: interruption --------------------------------------------------------

class TestInterruption:
    def test_i_ctrl_c_first_call_exit_130_no_traceback_no_output(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)

        def interrupt(*a, **k):
            raise KeyboardInterrupt()
        monkeypatch.setattr(openrouter_client, "_chat_completion", interrupt)
        out = tmp_path / "out" / "r.xlsx"
        result = _run(["run", "--input", str(s), "--output", str(out)])
        assert result.exit_code == 130
        assert "Traceback" not in result.output
        assert result.exception is None or isinstance(result.exception, SystemExit)
        # One file, interrupted mid-call -> a controlled interrupted row exists,
        # so partial output IS written (>=1 recorded file).
        assert out.exists()
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]

    def test_i_zero_completed_writes_no_output(self, tmp_path, monkeypatch):
        # Interrupt during process_directory BEFORE the first file is recorded
        # (no in-flight identity) -> zero results -> no output written.
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        from invoice_extractor import pipeline

        def interrupt_batch(*a, **k):
            raise pipeline.BatchInterrupted([])
        from invoice_extractor import service as service_module
        monkeypatch.setattr(service_module, "process_directory", interrupt_batch)
        out = tmp_path / "out" / "r.xlsx"
        result = _run(["run", "--input", str(s), "--output", str(out)])
        assert result.exit_code == 130
        assert not out.exists()
        assert "no output written" in result.output

    def test_j_ctrl_c_after_one_file_preserves_completed(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = tmp_path / "samples"
        s.mkdir()
        build_pdf(s / "a.pdf", [("text", TEXT_BODY)])
        build_pdf(s / "b.pdf", [("text", TEXT_BODY)])
        # a.pdf succeeds; b.pdf's first call raises KeyboardInterrupt.
        rec = Recorder([_envelope(invoice_json()), KeyboardInterrupt()])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        out = tmp_path / "out" / "r.xlsx"
        usage = out.parent / "r.usage.csv"
        result = _run(["run", "--input", str(s), "--output", str(out)])
        assert result.exit_code == 130
        assert out.exists() and usage.exists()  # valid partial artifacts
        wb = openpyxl.load_workbook(out)
        # Both files recorded (a.pdf completed; b.pdf interrupted mid-call).
        invoice_sources = {r[list(wb["Invoices"][1]).index(
            next(c for c in wb["Invoices"][1] if c.value == "source_file"))]
            for r in wb["Invoices"].iter_rows(min_row=2, values_only=True)}
        assert {"a.pdf", "b.pdf"} <= invoice_sources
        # NeedsReview: source_file at col 1, review_reason at col 4.
        review = {r[1]: r[4] for r in wb["NeedsReview"].iter_rows(min_row=2, values_only=True)}
        assert "b.pdf" in review
        assert "interrupted by operator" in review["b.pdf"]


# --- K/L/M/N + artifact-set safeguard: atomic writes --------------------------

class TestAtomicWrites:
    def test_k_export_failure_keeps_old_workbook_and_cleans_temp(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        out.parent.mkdir(parents=True)
        out.write_text("OLD-WORKBOOK-SENTINEL")  # pre-existing final
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        def broken_export(results, path):
            raise OSError("disk full (synthetic)")
        from invoice_extractor import service as service_module
        monkeypatch.setattr(service_module, "export_workbook", broken_export)

        result = _run(["run", "--input", str(s), "--output", str(out), "--overwrite"])
        assert result.exit_code != 0
        assert out.read_text() == "OLD-WORKBOOK-SENTINEL"  # untouched
        assert list(out.parent.glob("*.tmp-*")) == []       # temp cleaned

    def test_artifact_set_later_failure_keeps_earlier_final(self, tmp_path, monkeypatch):
        # Workbook temp writes fine; usage CSV write fails -> NEITHER final is
        # replaced (both existing finals untouched, all temps cleaned).
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        usage = out.parent / "r.usage.csv"
        out.parent.mkdir(parents=True)
        out.write_text("OLD-WORKBOOK")
        usage.write_text("OLD-USAGE")
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        def broken_usage(records, path):
            raise OSError("usage write failed (synthetic)")
        from invoice_extractor import service as service_module
        monkeypatch.setattr(service_module, "write_usage_csv", broken_usage)

        result = _run(["run", "--input", str(s), "--output", str(out), "--overwrite"])
        assert result.exit_code != 0
        assert out.read_text() == "OLD-WORKBOOK"   # earlier final NOT replaced
        assert usage.read_text() == "OLD-USAGE"
        assert list(out.parent.glob("*.tmp-*")) == []

    def test_l_m_usage_and_metadata_written_atomically_on_success(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        meta = out.parent / "r.run.json"
        rec = Recorder([_envelope(invoice_json(), cost=0.001)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(out),
                       "--run-metadata", str(meta)])
        assert result.exit_code == 0, result.output
        assert out.exists() and (out.parent / "r.usage.csv").exists() and meta.exists()
        assert list(out.parent.glob("*.tmp-*")) == []  # no temps remain

    def test_staged_artifacts_helper_unit(self, tmp_path):
        a, b = tmp_path / "a.txt", tmp_path / "b.txt"
        a.write_text("OLD-A")
        with pytest.raises(OSError):
            with StagedArtifacts() as stage:
                stage.stage(a, lambda p: p.write_text("NEW-A"))
                stage.stage(b, lambda p: (_ for _ in ()).throw(OSError("boom")))
                stage.commit()
        assert a.read_text() == "OLD-A"          # never replaced
        assert not b.exists()
        assert list(tmp_path.glob("*.tmp-*")) == []


# --- O/P/Q/R/S: overwrite protection ------------------------------------------

class TestOverwriteProtection:
    def test_o_existing_workbook_blocks_before_calls(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        out.parent.mkdir(parents=True)
        out.write_text("EXISTING")
        rec = Recorder([])  # any call is a failure
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(out)])
        assert result.exit_code == 1
        assert "already exists" in result.output
        assert "no provider calls were made" in result.output
        assert rec.calls == 0
        assert out.read_text() == "EXISTING"

    def test_p_existing_usage_sidecar_blocks(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        (out.parent).mkdir(parents=True)
        (out.parent / "r.usage.csv").write_text("EXISTING-USAGE")
        rec = Recorder([])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(out)])
        assert result.exit_code == 1
        assert rec.calls == 0

    def test_q_existing_run_metadata_blocks(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        meta = out.parent / "r.run.json"
        out.parent.mkdir(parents=True)
        meta.write_text("EXISTING-META")
        rec = Recorder([])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(out),
                       "--run-metadata", str(meta)])
        assert result.exit_code == 1
        assert rec.calls == 0

    def test_r_overwrite_replaces_only_after_success(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        out.parent.mkdir(parents=True)
        out.write_text("OLD")
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(out), "--overwrite"])
        assert result.exit_code == 0, result.output
        openpyxl.load_workbook(out)  # real workbook now (replaced)

    def test_s_failed_overwrite_keeps_old(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        out = tmp_path / "out" / "r.xlsx"
        out.parent.mkdir(parents=True)
        out.write_text("OLD-KEEP")
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        from invoice_extractor import service as service_module
        monkeypatch.setattr(service_module, "export_workbook",
                            lambda r, p: (_ for _ in ()).throw(OSError("fail")))
        result = _run(["run", "--input", str(s), "--output", str(out), "--overwrite"])
        assert result.exit_code != 0
        assert out.read_text() == "OLD-KEEP"

    def test_benchmark_output_collision_blocks(self, tmp_path):
        from .benchmark_helpers import gt, invoice_row, manifest_entry, write_manifest, write_workbook
        manifest = write_manifest(tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")})
        wb = write_workbook(tmp_path / "results.xlsx", [invoice_row("INV-1", "a.pdf")], [])
        report_out = tmp_path / "bench.xlsx"
        report_out.write_text("EXISTING-REPORT")
        result = _run(["benchmark", "score", "--manifest", str(manifest),
                       "--workbook", str(wb), "--output", str(report_out)])
        assert result.exit_code == 1
        assert "already exists" in result.output
        assert report_out.read_text() == "EXISTING-REPORT"


# --- T/U/V/W: input/output preflight ------------------------------------------

class TestPreflight:
    def test_t_invalid_input_path(self, tmp_path):
        result = _run(["run", "--input", str(tmp_path / "nope")])
        assert result.exit_code != 0  # click rejects a nonexistent --input

    def test_u_input_is_a_file(self, tmp_path):
        f = tmp_path / "a.pdf"
        f.write_text("x")
        result = _run(["run", "--input", str(f)])
        assert result.exit_code != 0  # click file_okay=False

    def test_v_output_is_a_directory(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        outdir = tmp_path / "out"
        outdir.mkdir()
        result = _run(["run", "--input", str(s), "--output", str(outdir)])
        assert result.exit_code != 0

    def test_w_unwritable_output_parent(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        # Make the intended output parent creation fail.
        monkeypatch.setattr(cli_module.Path, "mkdir",
                            lambda *a, **k: (_ for _ in ()).throw(PermissionError("ro")))
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "ro" / "r.xlsx")])
        assert result.exit_code != 0
        assert "not writable" in result.output
        assert rec.calls == 0


# --- X/Y/Z: empty dir + no-budget warning -------------------------------------

class TestEmptyAndBudgetWarning:
    def test_x_empty_directory_exit_zero(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = tmp_path / "samples"
        s.mkdir()
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0
        assert "No PDFs found" in result.output

    def test_y_no_safety_limits_one_warning(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)  # no budget vars set
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0, result.output
        assert result.output.count("no OpenRouter safety limits configured") == 1

    def test_z_limits_configured_no_warning(self, tmp_path, monkeypatch):
        _or_env(monkeypatch, MAX_MODEL_ATTEMPTS_PER_FILE="3")
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0, result.output
        assert "no OpenRouter safety limits configured" not in result.output


# --- AA/AB/AC: doctor ---------------------------------------------------------

class TestDoctor:
    def test_aa_ab_doctor_openrouter_readiness(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)  # openrouter, no budgets
        result = _run(["doctor", "--input", str(tmp_path), "--output", str(tmp_path / "o")])
        assert "Gateway: openrouter" in result.output
        assert "OPENROUTER_API_KEY" in result.output
        assert "tv/text-1, tv/text-2" in result.output           # model list shown
        assert "no safety limits configured" in result.output    # AB warning
        assert "test-or-key" not in result.output                # key value never printed

    def test_ac_doctor_makes_zero_network_calls(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        called = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: called.append(1))
        monkeypatch.setattr(gemini_client, "_generate", lambda *a, **k: called.append(1))
        _run(["doctor", "--input", str(tmp_path), "--output", str(tmp_path / "o")])
        assert called == []


# --- AD-AH: error messages ----------------------------------------------------

class TestErrorMessages:
    def test_ad_missing_text_models_actionable(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        monkeypatch.delenv("OPENROUTER_TEXT_MODELS", raising=False)
        s = _text_samples(tmp_path)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0  # controlled review, not a crash
        wb = openpyxl.load_workbook(tmp_path / "out" / "r.xlsx")
        reasons = " ".join(str(c.value) for row in wb["NeedsReview"].iter_rows()
                           for c in row if c.value)
        assert "OPENROUTER_TEXT_MODELS" in reasons

    def test_ae_missing_vision_models_actionable(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        monkeypatch.delenv("OPENROUTER_VISION_MODELS", raising=False)
        s = tmp_path / "samples"
        s.mkdir()
        build_pdf(s / "scan.pdf", [("image",)])
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0
        wb = openpyxl.load_workbook(tmp_path / "out" / "r.xlsx")
        reasons = " ".join(str(c.value) for row in wb["NeedsReview"].iter_rows()
                           for c in row if c.value)
        assert "OPENROUTER_VISION_MODELS" in reasons

    def test_ag_all_configured_models_wording_not_all_providers(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        err = openrouter_client.ProviderError("x", category="rate_limited", http_status=429)
        rec = Recorder([err, err])  # both text models fail
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0
        wb = openpyxl.load_workbook(tmp_path / "out" / "r.xlsx")
        reasons = " ".join(str(c.value) for row in wb["NeedsReview"].iter_rows()
                           for c in row if c.value)
        assert "all configured models" in reasons
        assert "all providers" not in reasons

    def test_ah_corrupt_pdf_safe_message(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = tmp_path / "samples"
        s.mkdir()
        (s / "corrupt.pdf").write_bytes(b"NOT-A-PDF-M7")
        rec = Recorder([])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0
        wb = openpyxl.load_workbook(tmp_path / "out" / "r.xlsx")
        reasons = " ".join(str(c.value) for row in wb["NeedsReview"].iter_rows()
                           for c in row if c.value)
        assert "unreadable PDF" in reasons
        assert "NOT-A-PDF-M7" not in reasons


# --- AJ/AK/AL: hermeticity ----------------------------------------------------

class TestHermeticity:
    def test_aj_no_repo_run_log_written(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        repo_log = Path("output/run.log")
        before = repo_log.stat().st_mtime if repo_log.exists() else None
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0, result.output
        after = repo_log.stat().st_mtime if repo_log.exists() else None
        assert before == after  # repo output/run.log untouched (or still absent)

    def test_al_no_temp_files_remain_after_success(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        out = tmp_path / "out" / "r.xlsx"
        _run(["run", "--input", str(s), "--output", str(out),
              "--run-metadata", str(out.parent / "r.run.json")])
        assert list(out.parent.rglob("*.tmp-*")) == []


# --- AM/AN/AO: schema/metadata unchanged --------------------------------------

class TestSchemaUnchanged:
    def test_am_workbook_three_sheets(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        out = tmp_path / "out" / "r.xlsx"
        _run(["run", "--input", str(s), "--output", str(out)])
        assert openpyxl.load_workbook(out).sheetnames == ["Invoices", "LineItems", "NeedsReview"]

    def test_an_usage_csv_schema_unchanged(self, tmp_path, monkeypatch):
        from invoice_extractor.usage import USAGE_CSV_COLUMNS
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        out = tmp_path / "out" / "r.xlsx"
        _run(["run", "--input", str(s), "--output", str(out)])
        import csv
        with open(out.parent / "r.usage.csv", newline="", encoding="utf-8") as f:
            header = next(csv.reader(f))
        assert header == USAGE_CSV_COLUMNS

    def test_ao_run_metadata_only_safe_fields(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = _text_samples(tmp_path)
        rec = Recorder([_envelope(invoice_json(), cost=0.001)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        out = tmp_path / "out" / "r.xlsx"
        meta = out.parent / "r.run.json"
        _run(["run", "--input", str(s), "--output", str(out), "--run-metadata", str(meta)])
        payload = json.loads(meta.read_text())
        assert set(payload) == {"run_id", "started_at", "finished_at", "interrupted",
                                "exit_code", "input_dir", "output_artifacts", "files"}
        row = payload["files"][0]
        assert set(row) == {"source_file", "elapsed_seconds", "extraction_method",
                            "provider", "model", "needs_review", "error", "completed",
                            "interrupted", "request_count", "reported_cost",
                            "unknown_cost_count"}
        blob = meta.read_text()
        for forbidden in ("review_reason", "Ocean freight", "INV-1001", "prompt"):
            assert forbidden not in blob


# --- AQ: CLI help -------------------------------------------------------------

class TestHelp:
    def test_aq_run_help_documents_overwrite_and_interruption(self):
        result = _run(["run", "--help"])
        assert "--overwrite" in result.output
        assert "--run-metadata" in result.output
        low = result.output.lower()
        assert "130" in result.output and "interrupt" in low

    def test_benchmark_help_documents_overwrite(self):
        result = _run(["benchmark", "score", "--help"])
        assert "--overwrite" in result.output


# --- AV: privacy across all M7 surfaces ---------------------------------------

class TestPrivacy:
    def test_av_no_sensitive_content_in_logs_outputs_or_metadata(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        monkeypatch.setenv("OPENROUTER_API_KEY", "SECRET-OR-KEY-M7")
        s = _text_samples(tmp_path)
        body = "UNIQUE-FAKE-BODY-M7"
        b64 = "RkFLRUJBU0U2NC1NNw=="
        rec = Recorder([_envelope(f"not valid json {body} {b64}"),
                        _envelope(f"still not valid json {body} {b64}"),
                        _envelope(f"more junk {body}"),          # tier-2 primary
                        _envelope(f"more junk repair {body}")])  # tier-2 repair
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        out = tmp_path / "out" / "r.xlsx"
        meta = out.parent / "r.run.json"
        log = tmp_path / "run.log"
        result = _run(["run", "--input", str(s), "--output", str(out),
                       "--run-metadata", str(meta), "--log-file", str(log)])
        assert result.exit_code == 0
        blobs = [result.output, log.read_text(), meta.read_text(),
                 (out.parent / "r.usage.csv").read_text()]
        wb = openpyxl.load_workbook(out)
        blobs.append(" ".join(str(c.value) for sh in wb.worksheets
                              for row in sh.iter_rows() for c in row if c.value is not None))
        forbidden = ["SECRET-OR-KEY-M7", body, b64, "data:image/png;base64",
                     "INVOICE INV-1001", "Ocean freight", "Traceback"]
        for blob in blobs:
            for secret in forbidden:
                assert secret not in blob, f"leaked: {secret}"


# --- AW: one file failure never crashes the batch -----------------------------

class TestBatchResilience:
    def test_aw_one_failure_batch_continues(self, tmp_path, monkeypatch):
        _or_env(monkeypatch)
        s = tmp_path / "samples"
        s.mkdir()
        build_pdf(s / "a_good.pdf", [("text", TEXT_BODY)])
        (s / "b_corrupt.pdf").write_bytes(b"NOT-A-PDF")
        build_pdf(s / "c_good.pdf", [("text", TEXT_BODY)])
        rec = Recorder([_envelope(invoice_json()), _envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = _run(["run", "--input", str(s), "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0
        wb = openpyxl.load_workbook(tmp_path / "out" / "r.xlsx")
        assert len(list(wb["Invoices"].iter_rows(min_row=2))) == 3

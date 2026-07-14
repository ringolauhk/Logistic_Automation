import json
from decimal import Decimal

import openpyxl
import pandas as pd

from invoice_extractor.excel_export import export_workbook
from invoice_extractor.pipeline import InvoiceResult
from invoice_extractor.schema import empty_invoice, normalize_invoice, validate_invoice

from .conftest import invoice_dict
from .synthetic_fixtures import provider_responses as pr


def sample_results():
    ok = InvoiceResult(
        source_file="good.pdf",
        invoice=normalize_invoice(invoice_dict()),
        page_count=1,
        document_classification="text-native",
        extraction_method="text",
        provider="gemini",
        model="gemini-test-text",
        text_pages=[1],
    )
    flagged = InvoiceResult(
        source_file="flagged.pdf",
        invoice=normalize_invoice(invoice_dict(total_amount=999.0)),
        page_count=2,
        document_classification="mixed",
        extraction_method="mixed",
        provider="mixed",
        model="gemini-test-text+claude-test-vision",
        text_pages=[1, 2, 5, 6, 7],
        image_pages=[3],
        failed_pages=[4],
        vision_chunk_count=1,
        needs_review=True,
        review_reason="totals inconclusive: synthetic",
    )
    failed = InvoiceResult(  # all-null row: both providers failed
        source_file="failed.pdf",
        invoice=empty_invoice(),
        page_count=1,
        document_classification="image-only",
        extraction_method="failed",
        provider="none",
        model=None,
        image_pages=[1],
        needs_review=True,
        error=True,
        review_reason="vision route failed on all providers: synthetic",
    )
    return [ok, flagged, failed]


class TestWorkbookStructure:
    def test_exactly_three_sheets(self, tmp_path):
        path = export_workbook(sample_results(), tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]

    def test_reopens_with_openpyxl_and_pandas(self, tmp_path):
        path = export_workbook(sample_results(), tmp_path / "out.xlsx")
        openpyxl.load_workbook(path).close()
        for sheet in ("Invoices", "LineItems", "NeedsReview"):
            pd.read_excel(path, sheet_name=sheet)

    def test_provenance_columns_present(self, tmp_path):
        path = export_workbook(sample_results(), tmp_path / "out.xlsx")
        df = pd.read_excel(path, sheet_name="Invoices")
        for col in ("document_classification", "extraction_method", "provider",
                    "model", "text_pages", "image_pages", "blank_pages",
                    "failed_pages", "vision_chunk_count"):
            assert col in df.columns
        assert set(df["extraction_method"]) == {"text", "mixed", "failed"}
        assert set(df["provider"]) == {"gemini", "mixed", "none"}

    def test_page_columns_use_human_readable_ranges(self, tmp_path):
        path = export_workbook(sample_results(), tmp_path / "out.xlsx")
        # assert on raw openpyxl cells (pandas read_excel would type-infer "4" -> 4.0)
        ws = openpyxl.load_workbook(path)["Invoices"]
        headers = [c.value for c in ws[1]]
        row = next(r for r in ws.iter_rows(min_row=2)
                   if r[headers.index("source_file")].value == "flagged.pdf")
        assert row[headers.index("text_pages")].value == "1-2,5-7"  # documented format
        assert row[headers.index("failed_pages")].value == "4"
        assert row[headers.index("vision_chunk_count")].value == 1
        assert "[" not in str(row[headers.index("text_pages")].value)  # no list syntax


class TestReferentialIntegrity:
    def test_line_items_reference_existing_invoices(self, tmp_path):
        path = export_workbook(sample_results(), tmp_path / "out.xlsx")
        invoices = pd.read_excel(path, sheet_name="Invoices")
        items = pd.read_excel(path, sheet_name="LineItems")
        assert not items.empty
        assert set(items["invoice_id"]).issubset(set(invoices["invoice_id"]))

    def test_needs_review_is_exact_subset(self, tmp_path):
        path = export_workbook(sample_results(), tmp_path / "out.xlsx")
        invoices = pd.read_excel(path, sheet_name="Invoices")
        review = pd.read_excel(path, sheet_name="NeedsReview")
        assert bool(review["needs_review"].all())
        expected = set(invoices[invoices["needs_review"]]["invoice_id"])
        assert set(review["invoice_id"]) == expected


class TestEdgeCases:
    def test_no_line_items_still_valid_workbook(self, tmp_path):
        failed_only = [sample_results()[2]]
        path = export_workbook(failed_only, tmp_path / "empty_items.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]
        items = pd.read_excel(path, sheet_name="LineItems")
        assert items.empty

    def test_null_values_do_not_break_column_sizing(self, tmp_path):
        # failed row is all-None across header fields
        path = export_workbook(sample_results(), tmp_path / "nulls.xlsx")
        assert path.exists()

    def test_decimals_exported_as_numeric_cells(self, tmp_path):
        path = export_workbook(sample_results(), tmp_path / "decimals.xlsx")
        wb = openpyxl.load_workbook(path)
        ws = wb["Invoices"]
        headers = [c.value for c in ws[1]]
        total_col = headers.index("total_amount") + 1
        cell = ws.cell(row=2, column=total_col)  # the "good.pdf" row
        assert isinstance(cell.value, (int, float))
        assert not isinstance(cell.value, str)
        assert abs(cell.value - float(Decimal("119.0"))) < 1e-9


def _fixture_08_hallucination_result() -> InvoiceResult:
    """Real fixture-8 hallucination data run through normalize_invoice/
    validate_invoice (schema-level, no PDF/provider mocking needed - the full
    pipeline path for this exact scenario is already covered by
    test_synthetic_pipeline.py::TestFixture08HappyPath)."""
    raw = json.loads(pr.fixture_08_hallucination_response_json())
    inv = normalize_invoice(raw)
    reason = validate_invoice(inv)
    return InvoiceResult(
        source_file="repeated_table_headers.pdf",
        invoice=inv,
        document_classification="text-native",
        extraction_method="text",
        provider="gemini",
        model="gemini-test-text",
        needs_review=reason is not None,
        review_reason=reason,
    )


class TestNeedsReviewSheetContent:
    def test_hallucination_case_flagged_with_line_detail(self, tmp_path):
        result = _fixture_08_hallucination_result()
        assert result.needs_review is True  # sanity: guardrail from the prior milestone fired

        path = export_workbook([result], tmp_path / "out.xlsx")
        review = pd.read_excel(path, sheet_name="NeedsReview")

        assert len(review) == 1
        row = review.iloc[0]
        assert row["source_file"] == "repeated_table_headers.pdf"
        assert row["invoice_number"] == "REP-4400"
        assert "missing an amount" in row["review_reason"]
        assert "possible hallucinated header/label row" in row["review_reason"]
        assert row["line_numbers"] == "2, 5, 8"
        assert row["line_descriptions"] == "Description; Description; Description"

    def test_happy_path_invoice_produces_no_review_row(self, tmp_path):
        ok = InvoiceResult(
            source_file="good.pdf",
            invoice=normalize_invoice(invoice_dict()),
            document_classification="text-native",
            extraction_method="text",
            provider="gemini",
            model="gemini-test-text",
            needs_review=False,
        )
        path = export_workbook([ok], tmp_path / "out.xlsx")
        review = pd.read_excel(path, sheet_name="NeedsReview")
        assert review.empty

    def test_failed_invoice_appears_in_needs_review_without_line_detail(self, tmp_path):
        failed = InvoiceResult(
            source_file="failed.pdf",
            invoice=empty_invoice(),
            document_classification="image-only",
            extraction_method="failed",
            provider="none",
            model=None,
            needs_review=True,
            error=True,
            review_reason="vision route failed on all providers: synthetic",
        )
        path = export_workbook([failed], tmp_path / "out.xlsx")

        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]  # unbroken by a failure

        review = pd.read_excel(path, sheet_name="NeedsReview")
        assert len(review) == 1
        row = review.iloc[0]
        assert row["source_file"] == "failed.pdf"
        assert "failed on all providers" in row["review_reason"]
        assert pd.isna(row["line_numbers"])  # invoice-level reason, no specific rows
        assert pd.isna(row["line_descriptions"])

    def test_multi_file_export_keeps_rows_associated_with_source_file(self, tmp_path):
        clean = _fixture_08_hallucination_result()
        clean.source_file = "second_hallucination.pdf"  # a second, distinct file
        results = [_fixture_08_hallucination_result(), clean, sample_results()[0]]  # ok invoice too

        path = export_workbook(results, tmp_path / "out.xlsx")
        invoices = pd.read_excel(path, sheet_name="Invoices")
        items = pd.read_excel(path, sheet_name="LineItems")
        review = pd.read_excel(path, sheet_name="NeedsReview")

        # every line item and every review row traces back to the correct source_file
        by_id = dict(zip(invoices["invoice_id"], invoices["source_file"]))
        for _, li in items.iterrows():
            assert li["source_file"] == by_id[li["invoice_id"]]
        for _, r in review.iterrows():
            assert r["source_file"] == by_id[r["invoice_id"]]

        assert set(review["source_file"]) == {
            "repeated_table_headers.pdf", "second_hallucination.pdf",
        }
        assert "good.pdf" not in set(review["source_file"])


class TestBlankInvoiceNumber:
    """Real commercial/customs invoices sometimes have no true invoice
    number - only a PO number or other reference (see schema.py's
    missing_identifier). The workbook must handle a blank invoice_number
    like any other optional field, not as an error."""

    def _result(self, **invoice_overrides) -> InvoiceResult:
        inv = normalize_invoice(invoice_dict(invoice_number=None, **invoice_overrides))
        reason = validate_invoice(inv)
        return InvoiceResult(
            source_file="po_only_invoice.pdf",
            invoice=inv,
            document_classification="text-native",
            extraction_method="text",
            provider="gemini",
            model="gemini-test-text",
            needs_review=reason is not None,
            review_reason=reason,
        )

    def test_blank_invoice_number_with_po_number_exports_and_is_not_flagged(
        self, tmp_path
    ):
        result = self._result(po_number="PO-4471")
        assert result.needs_review is False  # sanity: alternative identifier accepted

        path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]

        invoices = pd.read_excel(path, sheet_name="Invoices")
        assert len(invoices) == 1
        assert pd.isna(invoices.iloc[0]["invoice_number"])  # blank, not an error
        assert invoices.iloc[0]["po_number"] == "PO-4471"

        review = pd.read_excel(path, sheet_name="NeedsReview")
        assert review.empty

    def test_blank_invoice_number_and_no_alternative_is_flagged_but_still_exports(
        self, tmp_path,
    ):
        result = self._result()  # no po_number, no reference either
        assert result.needs_review is True
        assert "no alternative PO/reference identifier" in result.review_reason

        path = export_workbook([result], tmp_path / "out.xlsx")
        invoices = pd.read_excel(path, sheet_name="Invoices")
        assert pd.isna(invoices.iloc[0]["invoice_number"])
        # the rest of the invoice is still usable/exported, not discarded
        assert invoices.iloc[0]["total_amount"] == 119.0

        review = pd.read_excel(path, sheet_name="NeedsReview")
        assert len(review) == 1
        assert "no alternative PO/reference identifier" in review.iloc[0]["review_reason"]

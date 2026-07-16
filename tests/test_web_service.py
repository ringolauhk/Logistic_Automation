"""M9 web UI: shared service (T), preflight (S), worker end-to-end +
cancellation (AL-AN), summary (AA), NeedsReview view (AB), downloads
(AC-AG), estimate helper, controlled errors (AO). Offline: the only network
activity is a worker SUBPROCESS talking to a loopback mock provider."""

import json
import os
import time
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from apps.web import cleanup, job_manager, ui_models
from apps.web.estimate import FilePlan, estimate_max_attempts
from apps.web.job_manager import ValidatedUpload
from apps.web.progress import read_events, read_status
from invoice_extractor import openrouter_client
from invoice_extractor.service import run_extraction

from .conftest import TEXT_BODY, build_pdf, invoice_json, make_config
from .webmock_provider import MockProvider

PDF_MIN = b"%PDF-1.4 minimal"


@pytest.fixture(autouse=True)
def jobs_root(tmp_path, monkeypatch):
    root = tmp_path / "jobs"
    monkeypatch.setenv("WEB_JOBS_DIR", str(root))
    return root


def _or_cfg(**over):
    base = dict(llm_gateway="openrouter", openrouter_api_key="test-or-key",
                openrouter_text_models=("tv/text-1",), max_retries=1)
    base.update(over)
    return make_config(**base)


def _envelope(content):
    return {"id": "g", "model": "served",
            "choices": [{"finish_reason": "stop", "native_finish_reason": "S",
                         "message": {"content": content}}],
            "usage": {"prompt_tokens": 1, "completion_tokens": 1, "total_tokens": 2,
                      "cost": 0.0001, "completion_tokens_details": {}}}


# --- T: one shared service behind CLI and UI --------------------------------------

class TestSharedService:
    def test_t_cli_run_delegates_to_run_extraction(self, tmp_path, monkeypatch):
        from click.testing import CliRunner

        from invoice_extractor import service as service_module
        from invoice_extractor.cli import cli
        for var in ("LLM_GATEWAY", "OPENROUTER_API_KEY", "GEMINI_API_KEY",
                    "ANTHROPIC_API_KEY"):
            monkeypatch.delenv(var, raising=False)
        monkeypatch.setenv("GEMINI_API_KEY", "k")
        samples = tmp_path / "samples"
        samples.mkdir()
        build_pdf(samples / "inv.pdf", [("text", TEXT_BODY)])
        calls = []
        real = service_module.run_extraction

        def spy(*args, **kwargs):
            calls.append(1)
            return real(*args, **kwargs)
        monkeypatch.setattr(service_module, "run_extraction", spy)
        from invoice_extractor import gemini_client
        monkeypatch.setattr(gemini_client, "_generate",
                            lambda c, m, ct: invoice_json())
        result = CliRunner().invoke(cli, [
            "run", "--input", str(samples),
            "--output", str(tmp_path / "out" / "r.xlsx")])
        assert result.exit_code == 0, result.output
        assert calls == [1]  # the CLI went through the shared service

    def test_service_summary_counts(self, tmp_path, monkeypatch, logger):
        cfg = _or_cfg()
        build_pdf(tmp_path / "a.pdf", [("text", TEXT_BODY)])
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: _envelope(invoice_json()))
        outcome = run_extraction(tmp_path, tmp_path / "out" / "r.xlsx", cfg, logger)
        counts = outcome.summary_counts()
        assert counts["files_processed"] == 1
        assert counts["extracted"] == 1
        assert counts["requests"] == 1
        assert counts["reported_cost"] == "0.0001"
        assert counts["interrupted"] is False
        assert (tmp_path / "out" / "r.xlsx").exists()
        assert (tmp_path / "out" / "r.usage.csv").exists()


# --- S: preflight failure makes zero provider calls --------------------------------

class TestPreflight:
    def test_s_lock_held_prevents_start_and_calls(self, monkeypatch):
        jid = job_manager.create_job([ValidatedUpload("a.pdf", PDF_MIN)])
        other = job_manager.create_job([ValidatedUpload("b.pdf", PDF_MIN)])
        job_manager.acquire_lock(jid)
        calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: calls.append(1))
        with pytest.raises(job_manager.JobError, match="Another extraction"):
            job_manager.acquire_lock(other)
        assert calls == []  # refused BEFORE any provider work


# --- Worker end-to-end + cancellation (AL/AM/AN) ------------------------------------

def _make_job_with_pdfs(n):
    from tests.conftest import build_pdf as _build
    import tempfile
    uploads = []
    for i in range(n):
        tmp = Path(tempfile.mkdtemp()) / f"inv{i}.pdf"
        _build(tmp, [("text", TEXT_BODY)])
        uploads.append(ValidatedUpload(f"inv{i}.pdf", tmp.read_bytes()))
    return job_manager.create_job(uploads)


def _worker_env(monkeypatch, base_url):
    for var in ("GEMINI_API_KEY", "ANTHROPIC_API_KEY", "MAX_TEXT_PAGES",
                "MAX_VISION_PAGES", "MAX_MODEL_ATTEMPTS_PER_FILE",
                "MAX_COST_USD_PER_FILE", "MAX_COST_USD_PER_RUN"):
        monkeypatch.delenv(var, raising=False)
    monkeypatch.setenv("LLM_GATEWAY", "openrouter")
    monkeypatch.setenv("OPENROUTER_API_KEY", "test-or-key")
    monkeypatch.setenv("OPENROUTER_TEXT_MODELS", "tv/text-1")
    monkeypatch.setenv("OPENROUTER_BASE_URL", base_url)
    monkeypatch.setenv("MAX_RETRIES", "1")


def _wait(predicate, timeout=30.0):
    deadline = time.time() + timeout
    while time.time() < deadline:
        if predicate():
            return True
        time.sleep(0.25)
    return False


class TestWorkerEndToEnd:
    def test_full_worker_run_completes_and_releases_lock(self, monkeypatch):
        with MockProvider() as provider:
            _worker_env(monkeypatch, provider.base_url)
            jid = _make_job_with_pdfs(1)
            token = job_manager.acquire_lock(jid)
            job_manager.spawn_worker(jid, token, settings_env={},
                                     enable_log=True, enable_metadata=True)
            job_dir = job_manager.job_dir_for(jid)
            assert _wait(lambda: (read_status(job_dir) or {}).get("state")
                         in ("completed", "needs_review", "failed"))
            status = read_status(job_dir)
            assert status["state"] == "completed", status
            assert status["exit_code"] == 0
            assert status["summary"]["files_processed"] == 1
            # AA: summary counts match reality (1 file, 1 request).
            assert status["summary"]["requests"] == 1
            assert "results.xlsx" in status["artifacts"]
            assert (job_dir / "output" / "results.xlsx").exists()
            assert (job_dir / "output" / "results.usage.csv").exists()
            assert (job_dir / "output" / "results.run.json").exists()
            assert (job_dir / "logs" / "run.log").exists()
            # P (lock released after success).
            assert _wait(lambda: not job_manager.lock_path().exists(), 10)
            # Events were streamed.
            events, _ = read_events(job_dir / "events.jsonl")
            kinds = {e["event"] for e in events}
            assert {"job_started", "file_started", "provider_request_started",
                    "job_completed"} <= kinds

    def test_al_am_an_cancel_stops_calls_preserves_results_releases_lock(
            self, monkeypatch):
        # File 1 extracts instantly; the request for file 2 BLOCKS on the mock
        # server - we cancel while it is in flight.
        with MockProvider(delay_after=1, delay_seconds=25) as provider:
            _worker_env(monkeypatch, provider.base_url)
            jid = _make_job_with_pdfs(3)
            token = job_manager.acquire_lock(jid)
            job_manager.spawn_worker(jid, token, settings_env={},
                                     enable_log=False, enable_metadata=False)
            job_dir = job_manager.job_dir_for(jid)

            def second_request_inflight():
                return provider.request_count >= 2
            assert _wait(second_request_inflight), "second request never started"
            time.sleep(0.3)
            assert job_manager.cancel_job(jid, token) is True

            assert _wait(lambda: (read_status(job_dir) or {}).get("state")
                         == "cancelled", timeout=30), read_status(job_dir)
            status = read_status(job_dir)
            assert status["exit_code"] == 130
            # AL: no further provider calls after the in-flight one.
            assert provider.request_count == 2
            # AM: completed file preserved in a valid partial workbook.
            wb_path = job_dir / "output" / "results.xlsx"
            assert wb_path.exists()
            wb = openpyxl.load_workbook(wb_path)
            assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]
            rows = list(wb["Invoices"].iter_rows(min_row=2, values_only=True))
            assert len(rows) >= 1
            # AN: lock released after cancellation.
            assert _wait(lambda: not job_manager.lock_path().exists(), 10)
            # AO: no traceback in the safe status/events surface.
            blob = json.dumps(status) + \
                (job_dir / "events.jsonl").read_text()
            assert "Traceback" not in blob


# --- AB: NeedsReview summary uses safe categories -----------------------------------

class TestNeedsReviewView:
    def test_ab_rows_from_status_files(self):
        status = {"files": [
            {"source_file": "a.pdf", "needs_review": True, "error": False,
             "review_categories": ["totals_inconclusive"],
             "extraction_method": "text", "provider": "openrouter",
             "model": "m1"},
            {"source_file": "b.pdf", "needs_review": False, "error": False,
             "review_categories": [], "extraction_method": "text",
             "provider": "openrouter", "model": "m1"},
            {"source_file": "c.pdf", "needs_review": True, "error": True,
             "review_categories": ["malformed_pdf"],
             "extraction_method": "failed", "provider": "none", "model": None},
        ]}
        rows = ui_models.needs_review_rows(status)
        assert [r["File"] for r in rows] == ["a.pdf", "c.pdf"]
        assert rows[0]["Categories"] == "totals_inconclusive"
        assert rows[1]["Outcome"] == "failed"


# --- AC-AG: downloads ----------------------------------------------------------------

class TestDownloads:
    def _job_with_outputs(self):
        jid = job_manager.create_job([ValidatedUpload("secret-invoice.pdf",
                                                      PDF_MIN)])
        out = job_manager.job_dir_for(jid) / "output"
        (out / "results.xlsx").write_bytes(b"WORKBOOK")
        (out / "results.usage.csv").write_bytes(b"USAGE")
        return jid

    def test_ac_ad_workbook_and_usage_downloadable(self):
        jid = self._job_with_outputs()
        names = {n for n, _, _ in ui_models.downloadable_artifacts(jid)}
        assert names == {"results.xlsx", "results.usage.csv"}

    def test_ae_metadata_only_when_present(self):
        jid = self._job_with_outputs()
        out = job_manager.job_dir_for(jid) / "output"
        (out / "results.run.json").write_text("{}")
        names = {n for n, _, _ in ui_models.downloadable_artifacts(jid)}
        assert "results.run.json" in names

    def test_af_arbitrary_paths_not_downloadable(self):
        jid = self._job_with_outputs()
        out = job_manager.job_dir_for(jid) / "output"
        (out / "evil.sh").write_text("#!/bin/sh")
        (out / "notes.txt").write_text("x")
        names = {n for n, _, _ in ui_models.downloadable_artifacts(jid)}
        assert names == {"results.xlsx", "results.usage.csv"}
        # Foreign/invalid job ids yield nothing at all.
        assert ui_models.downloadable_artifacts("../etc") == []
        assert ui_models.downloadable_artifacts("job-x") == []

    def test_ag_uploaded_pdfs_never_downloadable(self):
        jid = self._job_with_outputs()
        names = {n for n, _, _ in ui_models.downloadable_artifacts(jid)}
        assert "secret-invoice.pdf" not in names
        assert not any(n.endswith(".pdf") for n in names)
        # Even a PDF dropped into output/ is not in the fixed allowlist.
        out = job_manager.job_dir_for(jid) / "output"
        (out / "copy.pdf").write_bytes(PDF_MIN)
        names = {n for n, _, _ in ui_models.downloadable_artifacts(jid)}
        assert not any(n.endswith(".pdf") for n in names)


# --- estimate helper ------------------------------------------------------------------

class TestEstimate:
    def test_openrouter_upper_bound_and_cap(self):
        cfg = make_config(llm_gateway="openrouter", openrouter_api_key="k",
                          openrouter_text_models=("a/1", "a/2", "a/3"),
                          openrouter_vision_models=("v/1", "v/2"),
                          max_text_pages=2, max_vision_pages=5, max_retries=3)
        plans = [FilePlan("a.pdf", text_pages=4, image_pages=6,
                          classification="mixed")]
        est = estimate_max_attempts(plans, cfg)
        # (2 text chunks x 3 + 2 vision chunks x 2) = 10 attempts x2 x3 = 60.
        assert est.max_attempts == 60
        assert any("Upper bound" in a for a in est.assumptions)

        capped = make_config(llm_gateway="openrouter", openrouter_api_key="k",
                             openrouter_text_models=("a/1", "a/2", "a/3"),
                             openrouter_vision_models=("v/1", "v/2"),
                             max_text_pages=2, max_vision_pages=5, max_retries=3,
                             max_model_attempts_per_file=3)
        est2 = estimate_max_attempts(plans, capped)
        assert est2.max_attempts == 18  # 3 x2 x3

    def test_direct_gateway_estimate(self):
        cfg = make_config(max_retries=3)  # direct; text gemini-only
        plans = [FilePlan("a.pdf", text_pages=1, image_pages=0,
                          classification="text-native")]
        est = estimate_max_attempts(plans, cfg)
        assert est.max_attempts == 6  # 1 chunk x 1 provider x2 x3

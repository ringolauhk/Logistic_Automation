"""M5: mixed-document routing and full end-to-end batch validation.

Mixed PDFs under LLM_GATEWAY=openrouter: text pages -> text ladder, image
pages -> vision ladder, ONE shared FileBudget, one aggregation, one final
validation. Plus heterogeneous-batch, malformed-input, page-accounting,
chronological-usage, and privacy hardening.

All offline - the autouse network-blocking fixture guards; mocked seams are
openrouter_client._chat_completion (provider boundary) and, where a test
must prove rendering did NOT happen, pdf_utils.render_pages_png.
"""

import csv
import json
from decimal import Decimal
from pathlib import Path

import openpyxl

from invoice_extractor import gemini_client, openrouter_client, pdf_utils
from invoice_extractor.excel_export import export_workbook
from invoice_extractor.pipeline import process_directory, process_file
from invoice_extractor.provider import ProviderError
from invoice_extractor.usage import format_usage_summary, summarize_usage, write_usage_csv

from .conftest import TEXT_BODY, build_pdf, invoice_json, make_config


def mixed_cfg(n_text=2, n_vision=2, **overrides):
    base = dict(
        llm_gateway="openrouter",
        openrouter_api_key="test-or-key",
        openrouter_text_models=tuple(f"tv/text-{i + 1}" for i in range(n_text)),
        openrouter_vision_models=tuple(f"tv/vis-{i + 1}" for i in range(n_vision)),
        max_retries=1,
        max_text_pages=2,
        max_vision_pages=2,
    )
    base.update(overrides)
    return make_config(**base)


def envelope(content, *, model="tv/served", finish_reason="stop",
            native_finish_reason="STOP", generation_id="gen-1", **usage_overrides):
    usage = {
        "prompt_tokens": 500, "completion_tokens": 100, "total_tokens": 600,
        "cost": 0.0002, "completion_tokens_details": {"reasoning_tokens": 5},
    }
    usage.update(usage_overrides)
    return {
        "id": generation_id,
        "model": model,
        "choices": [{
            "finish_reason": finish_reason,
            "native_finish_reason": native_finish_reason,
            "message": {"content": content},
        }],
        "usage": usage,
    }


class Recorder:
    def __init__(self, responses=None):
        self.calls = []
        self.responses = list(responses or [])

    def __call__(self, cfg, *, model, messages, response_format=None, max_tokens, timeout=None):
        self.calls.append({"model": model, "messages": messages})
        if not self.responses:
            raise AssertionError("provider called more times than expected")
        item = self.responses.pop(0)
        if isinstance(item, BaseException):
            raise item
        return item


def rate_limit_error():
    return ProviderError("OpenRouter request failed (HTTP 429)",
                         category="rate_limited", http_status=429)


def header_only(**overrides):
    data = {"line_items": []}
    data.update(overrides)
    return json.dumps(data)


def line_items_only(items):
    return json.dumps({"line_items": items})


def full_headers(**overrides):
    data = dict(invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=100)
    data.update(overrides)
    return header_only(**data)


def item(description="Freight", amount=100, **overrides):
    data = {"description": description, "quantity": 1,
            "unit_price": amount, "amount": amount}
    data.update(overrides)
    return data


def mixed_pdf(pdf_factory, name="mixed.pdf", specs=(("text", TEXT_BODY), ("image",))):
    return Path(pdf_factory(list(specs), name=name))


def render_spy(monkeypatch):
    rendered = []
    real = pdf_utils.render_pages_png

    def spy(path, page_numbers, dpi=200):
        rendered.append(list(page_numbers))
        return real(path, page_numbers, dpi)

    monkeypatch.setattr(pdf_utils, "render_pages_png", spy)
    return rendered


def no_direct_calls(monkeypatch):
    """Guard: any direct Gemini call under openrouter is a routing bug."""
    def forbid(*a, **k):
        raise AssertionError("direct Gemini called under LLM_GATEWAY=openrouter")
    monkeypatch.setattr(gemini_client, "_generate", forbid)


def usage_tuples(result):
    return [(r.page_range, r.attempt_type, r.accepted, r.route)
            for r in result.usage_records]


# --- A: mixed classification - exact, disjoint, complete page sets -------------

class TestClassification:
    def test_a_exact_disjoint_page_sets(self, logger, pdf_factory, monkeypatch):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory, specs=(
            ("text", TEXT_BODY), ("image",), ("blank",), ("text", TEXT_BODY),
        ))
        rec = Recorder([envelope(invoice_json())] * 2)  # 1 text chunk + 1 vision chunk
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        text, image, blank = (set(result.text_pages), set(result.image_pages),
                              set(result.blank_pages))
        assert text == {1, 4}
        assert image == {2}
        assert blank == {3}
        assert text | image | blank == {1, 2, 3, 4}   # no silent page loss
        assert text & image == set()
        assert text & blank == set()
        assert image & blank == set()
        assert result.failed_pages == []


# --- B/C: mixed success + usage CSV -------------------------------------------

class TestMixedSuccess:
    def test_b_text_and_vision_merge_into_one_invoice(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        no_direct_calls(monkeypatch)
        rec = Recorder([
            envelope(line_items_only([item("Text item", 100)])),  # text page 1
            envelope(full_headers()),                              # vision page 2
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.extraction_method == "mixed"
        assert result.provider == "openrouter"
        assert result.needs_review is False
        assert result.invoice.seller_name == "Acme"
        assert [it.description for it in result.invoice.line_items] == ["Text item"]

    def test_c_usage_csv_has_both_routes_chronological(
        self, logger, pdf_factory, monkeypatch, tmp_path
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(line_items_only([item()])),
            envelope(full_headers()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = process_file(pdf, cfg, logger)

        path = write_usage_csv(result.usage_records, tmp_path / "u.usage.csv")
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))

        assert [(r["route"], r["page_range"]) for r in rows] == [
            ("text", "1"), ("vision", "2"),
        ]


# --- Guardrail 1: chronological usage records ----------------------------------

class TestChronologicalUsage:
    def test_failed_middle_chunk_usage_stays_in_execution_order(
        self, logger, pdf_factory, monkeypatch, tmp_path
    ):
        # 6 text pages (chunks 1-2, 3-4, 5-6) + 1 image page (chunk 7).
        # Text chunk 2 fails across BOTH text models (primary + repair each);
        # its 4 rejected attempts must sit BETWEEN chunk 1's and chunk 3's
        # rows, not be appended after all successes.
        cfg = mixed_cfg(n_text=2, n_vision=1)
        pdf = mixed_pdf(pdf_factory, specs=([("text", TEXT_BODY)] * 6 + [("image",)]))
        rec = Recorder([
            envelope(line_items_only([item("A", 10)])),   # c1 tier1 primary: ok
            envelope("not valid json"),                    # c2 tier1 primary: malformed
            envelope("still not valid json"),              # c2 tier1 repair: malformed
            envelope("worse json"),                        # c2 tier2 escalation: malformed
            envelope("bad again"),                         # c2 tier2 repair: malformed
            envelope(line_items_only([item("B", 20)])),   # c3 tier1 primary: ok
            envelope(full_headers(total_amount=30)),       # vision tier1 primary: ok
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        expected = [
            ("1-2", "primary", True, "text"),
            ("3-4", "primary", False, "text"),
            ("3-4", "repair", False, "text"),
            ("3-4", "escalation", False, "text"),
            ("3-4", "repair", False, "text"),
            ("5-6", "primary", True, "text"),
            ("7", "primary", True, "vision"),
        ]
        assert usage_tuples(result) == expected  # exact order, no dupes, no gaps
        assert len(result.usage_records) == len(rec.calls) == 7

        # The CSV preserves the same order row-for-row.
        path = write_usage_csv(result.usage_records, tmp_path / "u.usage.csv")
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert [(r["page_range"], r["attempt_type"], r["route"]) for r in rows] == [
            (pr, at, route) for pr, at, _, route in expected
        ]

    def test_skips_create_no_fabricated_usage_rows(
        self, logger, pdf_factory, monkeypatch
    ):
        # Vision config invalid + text fine: only the ONE real text call may
        # produce a usage row - the skipped vision pages contribute nothing.
        cfg = mixed_cfg(n_vision=0)
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(result.usage_records) == len(rec.calls) == 1
        assert result.usage_records[0].route == "text"


# --- D/E/F: shared budgets across routes (mixed) --------------------------------

class TestSharedBudgets:
    def test_d_text_consumes_attempt_cap_vision_never_renders(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg(max_model_attempts_per_file=1)
        pdf = mixed_pdf(pdf_factory)
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(line_items_only([item("Kept")]))])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1
        assert rendered == []
        assert [it.description for it in result.invoice.line_items] == ["Kept"]
        assert result.needs_review is True
        assert result.review_reason.count("attempt cap") == 1
        assert result.review_reason.count("vision chunks pages 2") == 1

    def test_e_text_response_crosses_file_cost_vision_never_renders(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg(max_cost_usd_per_file=Decimal("0.0004"))
        pdf = mixed_pdf(pdf_factory)
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(line_items_only([item("Kept")]), cost=0.0005)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1
        assert rendered == []
        assert [it.description for it in result.invoice.line_items] == ["Kept"]
        # Guardrail 3: the crossing response's usage record IS retained...
        assert len(result.usage_records) == 1
        assert result.usage_records[0].cost_usd == Decimal("0.0005")
        # ...and the skipped vision chunk fabricated no row.
        assert "cost budget" in result.review_reason

    def test_f_run_budget_crossing_in_mixed_file_stops_everything(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = mixed_cfg(max_cost_usd_per_run=Decimal("0.0004"))
        build_pdf(tmp_path / "a_mixed.pdf", [("text", TEXT_BODY), ("image",)])
        build_pdf(tmp_path / "b_text.pdf", [("text", TEXT_BODY)])
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(invoice_json(), cost=0.0005)])  # a's text call crosses
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        assert len(rec.calls) == 1
        assert rendered == []
        a = next(r for r in results if r.source_file == "a_mixed.pdf")
        b = next(r for r in results if r.source_file == "b_text.pdf")
        assert a.review_reason.count("vision chunks pages 2") == 1
        assert b.error is True and "run-wide" in b.review_reason
        # Guardrail 3: exactly one usage row exists across the whole run.
        assert sum(len(r.usage_records) for r in results) == 1


# --- G/H: one route succeeds, the other fails ordinarily ------------------------

class TestPartialRouteFailure:
    def test_g_text_succeeds_vision_fails(self, logger, pdf_factory, monkeypatch):
        cfg = mixed_cfg(n_vision=1)
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(invoice_json()),  # text ok
            rate_limit_error(),         # vision tier1 (only model) fails
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.invoice.invoice_number == "INV-1001"  # text data retained
        assert result.needs_review is True
        assert result.error is False
        assert result.failed_pages == [2]
        assert "pages 2" in result.review_reason

    def test_h_vision_succeeds_text_fails(self, logger, pdf_factory, monkeypatch):
        cfg = mixed_cfg(n_text=1)
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            rate_limit_error(),         # text tier1 (only model) fails
            envelope(invoice_json()),  # vision ok
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.invoice.invoice_number == "INV-1001"  # vision data retained
        assert result.extraction_method == "vision"
        assert result.needs_review is True
        assert result.error is False
        assert result.failed_pages == [1]
        assert "pages 1" in result.review_reason


# --- I/J/K + Guardrail 2: missing model lists per route -------------------------

class TestMissingModelLists:
    def test_i_missing_vision_models_text_still_runs_no_fallback(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg(n_vision=0)
        pdf = mixed_pdf(pdf_factory)
        no_direct_calls(monkeypatch)
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1  # the text call only
        assert rendered == []
        assert result.invoice.invoice_number == "INV-1001"
        assert result.needs_review is True
        assert "OPENROUTER_VISION_MODELS" in result.review_reason

    def test_j_missing_text_models_vision_still_runs_one_compact_reason(
        self, logger, pdf_factory, monkeypatch
    ):
        # FOUR text pages (would be 2 chunks) + one image page: the text
        # config failure must appear exactly ONCE, covering all text pages -
        # not once per chunk - and the vision route still renders and runs.
        cfg = mixed_cfg(n_text=0)
        pdf = mixed_pdf(pdf_factory, specs=([("text", TEXT_BODY)] * 4 + [("image",)]))
        no_direct_calls(monkeypatch)
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1  # the vision call only
        assert rendered == [[5]]
        assert result.invoice.invoice_number == "INV-1001"  # vision data retained
        assert result.needs_review is True
        assert result.review_reason.count("OPENROUTER_TEXT_MODELS") == 1
        assert result.failed_pages == [1, 2, 3, 4]
        assert len(result.usage_records) == 1  # no fabricated rows for skipped text

    def test_k_both_lists_missing_controlled_row_batch_continues(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = mixed_cfg(n_text=0, n_vision=0)
        build_pdf(tmp_path / "a_mixed.pdf", [("text", TEXT_BODY), ("image",)])
        build_pdf(tmp_path / "b_mixed.pdf", [("text", TEXT_BODY), ("image",)])
        no_direct_calls(monkeypatch)
        rendered = render_spy(monkeypatch)
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))

        results = process_directory(tmp_path, cfg, logger)

        assert or_calls == []
        assert rendered == []
        assert len(results) == 2  # batch continued
        for r in results:
            assert r.error is True
            assert "OPENROUTER_TEXT_MODELS" in r.review_reason
            assert "OPENROUTER_VISION_MODELS" in r.review_reason
            assert r.usage_records == []


# --- L/M: provenance -------------------------------------------------------------

class TestProvenance:
    def test_l_one_actual_model_across_routes(self, logger, pdf_factory, monkeypatch):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(line_items_only([item()]), model="one-served"),
            envelope(full_headers(), model="one-served"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)
        assert result.model == "one-served"
        assert result.extraction_method == "mixed"

    def test_m_multiple_actual_models_compact_workbook_full_csv(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(line_items_only([item()]), model="text-served"),
            envelope(full_headers(), model="vis-served"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.model == "multiple"
        accepted = {r.actual_model for r in result.usage_records if r.accepted}
        assert accepted == {"text-served", "vis-served"}


# --- N/O: header conflicts across routes ----------------------------------------

class TestCrossRouteConflicts:
    def test_n_conflicting_seller_flagged_deterministic_keep(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(full_headers(seller_name="Text Seller")),
            envelope(full_headers(seller_name="Vision Seller")),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "conflict in seller_name" in result.review_reason
        # Deterministic: non-monetary conflicts keep the FIRST route by page
        # order - the text route (page 1).
        assert result.invoice.seller_name == "Text Seller"

    def test_o_conflicting_invoice_numbers_multi_invoice_signal(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(full_headers(invoice_number="INV-A")),
            envelope(full_headers(invoice_number="INV-B")),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "multiple invoices" in result.review_reason


# --- P/Q/R: line items across routes ---------------------------------------------

class TestCrossRouteLineItems:
    def test_p_exact_duplicate_across_routes_deduped(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        dup = item("Freight", 50)
        rec = Recorder([
            envelope(json.dumps({**json.loads(full_headers(total_amount=55)),
                                 "line_items": [dup]})),
            envelope(line_items_only([dup, item("Handling", 5)])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [it.description for it in result.invoice.line_items] == [
            "Freight", "Handling",
        ]

    def test_q_similar_but_not_identical_lines_both_retained(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(line_items_only([item("Freight", 50)])),
            envelope(line_items_only([item("Freight", 60)])),  # same desc, diff amount
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        amounts = [float(it.amount) for it in result.invoice.line_items]
        assert amounts == [50.0, 60.0]  # both kept, source order

    def test_r_line_no_and_item_code_distinct_across_routes(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(line_items_only([
                {"line_no": "1", "item_code": "A1", "description": "From text",
                 "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
            envelope(line_items_only([
                {"line_no": "2", "item_code": "B2", "description": "From vision",
                 "quantity": 1, "unit_price": 20, "amount": 20},
            ])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        items = result.invoice.line_items
        assert [(it.line_no, it.item_code) for it in items] == [("1", "A1"), ("2", "B2")]


# --- S/T: totals reconciliation post-aggregation --------------------------------

class TestTotalsAcrossRoutes:
    def test_s_totals_reconcile_only_after_both_routes(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(line_items_only([item("Only line", 100)])),  # items on text page
            envelope(full_headers(total_amount=100)),               # total on vision page
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)
        assert result.needs_review is False  # 100 == 100 after aggregation

    def test_t_totals_inconclusive_after_both_routes(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        rec = Recorder([
            envelope(line_items_only([item("Only line", 100)])),
            envelope(full_headers(total_amount=999)),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)
        assert result.needs_review is True
        assert "totals inconclusive" in result.review_reason


# --- U/V + Guardrail 4: page accounting -----------------------------------------

class TestPageAccounting:
    def test_u_exact_failed_and_covered_page_sets(
        self, logger, pdf_factory, monkeypatch
    ):
        # 5 pages: text 1-2, image 3-4, blank 5. Vision fails; text succeeds.
        cfg = mixed_cfg(n_vision=1)
        pdf = mixed_pdf(pdf_factory, specs=(
            ("text", TEXT_BODY), ("text", TEXT_BODY), ("image",), ("image",), ("blank",),
        ))
        rec = Recorder([envelope(invoice_json()), rate_limit_error()])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        text, image, blank = (set(result.text_pages), set(result.image_pages),
                              set(result.blank_pages))
        nonblank = text | image
        failed = set(result.failed_pages)
        assert text == {1, 2} and image == {3, 4} and blank == {5}
        assert text | image | blank == {1, 2, 3, 4, 5}
        assert failed == {3, 4}                    # exactly the failed vision pages
        assert failed <= nonblank                   # blank pages never "fail"
        assert failed & text == set()               # successful pages not failed
        covered = nonblank - failed
        assert covered == {1, 2}
        assert covered | failed == nonblank and covered & failed == set()
        assert "result covers pages 1-2 only" in result.review_reason

    def test_v_blank_pages_no_call_no_false_failure(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory, specs=(("text", TEXT_BODY), ("blank",)))
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1  # blank page triggered no provider call
        assert result.blank_pages == [2]
        assert result.failed_pages == []
        assert result.needs_review is False


# --- W/X/Y/AF: failure containment ------------------------------------------------

class TestFailureContainment:
    def test_w_all_routes_fail_null_row_batch_continues(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = mixed_cfg(n_text=1, n_vision=1)
        build_pdf(tmp_path / "a_mixed.pdf", [("text", TEXT_BODY), ("image",)])
        build_pdf(tmp_path / "b_text.pdf", [("text", TEXT_BODY)])
        rec = Recorder([
            rate_limit_error(), rate_limit_error(),  # a: both routes fail
            envelope(invoice_json()),                 # b: ok
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        a = next(r for r in results if r.source_file == "a_mixed.pdf")
        b = next(r for r in results if r.source_file == "b_text.pdf")
        assert a.error is True and a.extraction_method == "failed"
        assert b.error is False

    def test_x_af_corrupt_pdf_safe_row_no_batch_crash(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = mixed_cfg()
        (tmp_path / "a_corrupt.pdf").write_bytes(b"JUNK-NOT-A-PDF-M5")
        build_pdf(tmp_path / "b_good.pdf", [("text", TEXT_BODY)])
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        corrupt = next(r for r in results if r.source_file == "a_corrupt.pdf")
        good = next(r for r in results if r.source_file == "b_good.pdf")
        assert corrupt.error is True
        assert "unreadable PDF" in corrupt.review_reason
        assert "JUNK-NOT-A-PDF-M5" not in corrupt.review_reason  # no raw bytes leak
        assert good.error is False

    def test_y_blank_only_pdf_review_row_no_calls(self, logger, tmp_path, monkeypatch):
        cfg = mixed_cfg()
        build_pdf(tmp_path / "blank.pdf", [("blank",), ("blank",)])
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))

        results = process_directory(tmp_path, cfg, logger)

        assert or_calls == []
        assert results[0].error is True
        assert "blank" in results[0].review_reason

    def test_x_encrypted_pdf_safe_controlled_result(
        self, logger, tmp_path, monkeypatch
    ):
        import fitz
        doc = fitz.open()
        page = doc.new_page()
        page.insert_text((50, 72), TEXT_BODY, fontsize=10)
        doc.save(str(tmp_path / "locked.pdf"),
                 encryption=fitz.PDF_ENCRYPT_AES_256, user_pw="secret")
        doc.close()
        cfg = mixed_cfg()
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))

        results = process_directory(tmp_path, cfg, logger)

        assert or_calls == []
        assert results[0].error is True
        assert "unreadable PDF" in results[0].review_reason


# --- AA: usage summary reconciliation --------------------------------------------

class TestUsageSummaryReconciliation:
    def test_aa_counts_costs_and_unknown_cost_labeling(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = mixed_cfg()
        pdf = mixed_pdf(pdf_factory)
        no_cost = envelope(line_items_only([item()]))
        del no_cost["usage"]["cost"]  # text attempt: cost unavailable
        rec = Recorder([no_cost, envelope(full_headers(), cost=0.0123)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        s = summarize_usage(result.usage_records)
        assert s.total_requests == len(rec.calls) == 2
        assert s.accepted_requests == 2
        assert s.total_cost_usd == Decimal("0.0123")
        assert s.unknown_cost_count == 1
        text = format_usage_summary(result.usage_records, processed_count=1)
        assert "Requests with unknown cost: 1" in text
        assert "incomplete" in text


# --- AB: privacy across mixed routes ---------------------------------------------

class TestPrivacy:
    def test_ab_no_sensitive_content_anywhere(
        self, logger, pdf_factory, monkeypatch, tmp_path, caplog
    ):
        cfg = mixed_cfg(n_text=1, n_vision=1,
                        openrouter_api_key="SECRET-OR-KEY-M5",
                        debug_artifact_dir=str(tmp_path / "debug"))
        pdf = mixed_pdf(pdf_factory)
        body_marker = "UNIQUE-FAKE-INVOICE-BODY-M5"
        b64_marker = "RkFLRUJBU0U2NC1NNQ=="
        error_marker = "FAKE-PROVIDER-ERROR-META-M5"
        rec = Recorder([
            envelope(f"not valid json {body_marker} {b64_marker}"),        # text primary
            envelope(f"still not valid json {body_marker} {b64_marker}"),  # text repair
            {"id": "gen-x", "error": {"code": 500, "message": error_marker,
                                      "metadata": {"raw": error_marker}}},  # vision
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        with caplog.at_level("DEBUG", logger="invoice_extractor"), \
             caplog.at_level("DEBUG", logger="invoice_extractor_tests"):
            result = process_file(pdf, cfg, logger)

        assert result.error is True
        forbidden = ["SECRET-OR-KEY-M5", body_marker, b64_marker, error_marker,
                     "data:image/png;base64"]

        messages = " ".join(r.message for r in caplog.records)
        for secret in forbidden:
            assert secret not in messages, f"leaked into logs: {secret}"
            assert secret not in (result.review_reason or ""), \
                f"leaked into review_reason: {secret}"
        for record in result.usage_records:
            for secret in forbidden:
                assert secret not in repr(record)

        wb_path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(wb_path)
        for sheet in ("Invoices", "LineItems", "NeedsReview"):
            cells = " ".join(
                str(c.value) for row in wb[sheet].iter_rows() for c in row if c.value is not None
            )
            for secret in forbidden:
                assert secret not in cells, f"leaked into {sheet}: {secret}"

        usage_path = write_usage_csv(result.usage_records, tmp_path / "out.usage.csv")
        content = usage_path.read_text()
        for secret in forbidden:
            assert secret not in content, f"leaked into usage CSV: {secret}"
        assert not (tmp_path / "debug").exists()  # debug artifacts stay off


# --- AC/AD/AE: other routes unchanged --------------------------------------------

class TestOtherRoutesUnchanged:
    def test_ac_direct_gateway_mixed_unchanged(self, logger, pdf_factory, monkeypatch):
        cfg = make_config(llm_gateway="direct")
        pdf = mixed_pdf(pdf_factory)
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))
        monkeypatch.setattr(gemini_client, "_generate",
                            lambda c, m, ct: invoice_json())

        result = process_file(pdf, cfg, logger)

        assert or_calls == []
        assert result.provider == "gemini"
        assert result.extraction_method == "mixed"
        assert result.usage_records == []

    def test_ad_openrouter_text_only_unchanged(self, logger, pdf_factory, monkeypatch):
        cfg = mixed_cfg()
        pdf = Path(pdf_factory([("text", TEXT_BODY)], name="text.pdf"))
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.extraction_method == "text"
        assert result.needs_review is False
        assert [r.route for r in result.usage_records] == ["text"]

    def test_ae_openrouter_vision_only_unchanged(self, logger, pdf_factory, monkeypatch):
        cfg = mixed_cfg()
        pdf = Path(pdf_factory([("image",)], name="scan.pdf"))
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.extraction_method == "vision"
        assert result.needs_review is False
        assert [r.route for r in result.usage_records] == ["vision"]


# --- Z: heterogeneous batch, end-to-end through the CLI --------------------------

_ENV_VARS = [
    "LLM_GATEWAY", "OPENROUTER_API_KEY", "OPENROUTER_BASE_URL",
    "OPENROUTER_TEXT_MODELS", "OPENROUTER_VISION_MODELS", "OPENROUTER_APP_NAME",
    "OPENROUTER_SITE_URL", "OPENROUTER_STRUCTURED_OUTPUT",
    "MAX_MODEL_ATTEMPTS_PER_FILE", "MAX_COST_USD_PER_FILE", "MAX_COST_USD_PER_RUN",
    "MAX_TEXT_PAGES", "MAX_VISION_PAGES", "MAX_RETRIES",
    "GEMINI_API_KEY", "ANTHROPIC_API_KEY", "ENABLE_CLAUDE_TEXT_FALLBACK",
    "SAVE_DEBUG_ARTIFACTS", "DEBUG_ARTIFACT_DIR",
    "GEMINI_MODEL", "CLAUDE_MODEL", "GEMINI_TEXT_MODEL", "GEMINI_VISION_MODEL",
    "CLAUDE_TEXT_MODEL", "CLAUDE_VISION_MODEL",
    "TEXT_QUALITY_THRESHOLD", "REQUEST_TIMEOUT_SECONDS",
]


class TestHeterogeneousBatch:
    def _setup_env(self, monkeypatch):
        for var in _ENV_VARS:
            monkeypatch.delenv(var, raising=False)
        monkeypatch.setenv("LLM_GATEWAY", "openrouter")
        monkeypatch.setenv("OPENROUTER_API_KEY", "test-or-key")
        monkeypatch.setenv("OPENROUTER_TEXT_MODELS", "tv/text-1")
        monkeypatch.setenv("OPENROUTER_VISION_MODELS", "tv/vis-1")
        monkeypatch.setenv("MAX_TEXT_PAGES", "2")
        monkeypatch.setenv("MAX_VISION_PAGES", "2")
        monkeypatch.setenv("MAX_RETRIES", "1")

    def test_z_full_batch_cli_end_to_end(self, tmp_path, monkeypatch):
        from click.testing import CliRunner
        from invoice_extractor.cli import cli

        self._setup_env(monkeypatch)
        samples = tmp_path / "samples"
        samples.mkdir()
        build_pdf(samples / "a_text.pdf", [("text", TEXT_BODY)])
        build_pdf(samples / "b_bigtext.pdf", [("text", TEXT_BODY)] * 6)  # 3 chunks
        build_pdf(samples / "c_scan.pdf", [("image",)])
        build_pdf(samples / "d_multiscan.pdf", [("image",)] * 3)          # 2 chunks
        build_pdf(samples / "e_mixed.pdf", [("text", TEXT_BODY), ("image",)])
        (samples / "f_corrupt.pdf").write_bytes(b"NOT-A-PDF")
        build_pdf(samples / "g_blank.pdf", [("blank",)])
        (samples / "notes.txt").write_text("not a pdf - must be ignored")

        # Call order = sorted file order, chunks in page order within a file:
        # a(1 text), b(3 text), c(1 vision), d(2 vision), e(text, vision) = 9.
        # Multi-chunk files return IDENTICAL headers but DISTINCT line items
        # per chunk (a real invoice's pages carry different rows), so no
        # dedup note fires and totals reconcile after aggregation.
        def chunk_payload(desc, amount, total):
            return json.dumps({**json.loads(full_headers(total_amount=total)),
                               "line_items": [item(desc, amount)]})

        rec = Recorder([
            envelope(invoice_json()),                        # a: complete invoice
            envelope(chunk_payload("B1", 10, 100)),          # b chunk 1-2
            envelope(chunk_payload("B2", 20, 100)),          # b chunk 3-4
            envelope(chunk_payload("B3", 70, 100)),          # b chunk 5-6
            envelope(invoice_json()),                        # c: complete invoice
            envelope(chunk_payload("D1", 10, 30)),           # d chunk 1-2
            envelope(chunk_payload("D2", 20, 30)),           # d chunk 3
            envelope(chunk_payload("E1", 30, 80)),           # e text page 1
            envelope(chunk_payload("E2", 50, 80)),           # e vision page 2
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        out = tmp_path / "out" / "results.xlsx"
        result = CliRunner().invoke(
            cli, ["run", "--input", str(samples), "--output", str(out)]
        )

        assert result.exit_code == 0, result.output
        assert len(rec.calls) == 9

        # Workbook: exactly three sheets; every PDF (not the .txt) has a row.
        wb = openpyxl.load_workbook(out)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]
        invoice_rows = list(wb["Invoices"].iter_rows(min_row=2, values_only=True))
        assert len(invoice_rows) == 7
        review_rows = list(wb["NeedsReview"].iter_rows(min_row=2, values_only=True))
        review_files = {row[1] for row in review_rows}
        assert review_files == {"f_corrupt.pdf", "g_blank.pdf"}

        # CLI summary reconciles with the workbook.
        assert "Files processed:      7" in result.output
        assert "Invoices extracted:   5" in result.output
        assert "Needs review:         2" in result.output
        assert "Failed/problem:       2" in result.output
        assert "- mixed: 1" in result.output
        assert "- text: 2" in result.output
        assert "- vision: 2" in result.output
        assert "- failed: 2" in result.output
        assert "Requests:             9" in result.output  # usage summary present

        # Usage CSV: only files that made calls, in deterministic
        # chronological order (file order, then chunk order within file).
        usage_path = out.parent / "results.usage.csv"
        with open(usage_path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))
        assert len(rows) == 9
        assert [(r["source_file"], r["route"], r["page_range"]) for r in rows] == [
            ("a_text.pdf", "text", "1"),
            ("b_bigtext.pdf", "text", "1-2"),
            ("b_bigtext.pdf", "text", "3-4"),
            ("b_bigtext.pdf", "text", "5-6"),
            ("c_scan.pdf", "vision", "1"),
            ("d_multiscan.pdf", "vision", "1-2"),
            ("d_multiscan.pdf", "vision", "3"),
            ("e_mixed.pdf", "text", "1"),
            ("e_mixed.pdf", "vision", "2"),
        ]
        assert {r["source_file"] for r in rows} == {
            "a_text.pdf", "b_bigtext.pdf", "c_scan.pdf", "d_multiscan.pdf", "e_mixed.pdf",
        }  # failed/blank files contributed no rows

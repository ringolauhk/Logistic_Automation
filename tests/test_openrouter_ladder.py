"""M3: application-controlled OpenRouter TEXT model ladder + per-request
usage/cost accounting. All offline - the autouse network-blocking fixture
guards; the only mocked seam is openrouter_client._chat_completion.
"""

import csv
import json
from decimal import Decimal
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from invoice_extractor import gemini_client, openrouter_client
from invoice_extractor.excel_export import export_workbook
from invoice_extractor.pipeline import process_directory, process_file
from invoice_extractor.provider import ProviderError
from invoice_extractor.usage import RunBudget, usage_csv_path, write_usage_csv

from .conftest import TEXT_BODY, build_pdf, invoice_json, make_config


def ladder_cfg(n_models=3, **overrides):
    models = tuple(f"test-vendor/tier-{i + 1}" for i in range(n_models))
    base = dict(
        llm_gateway="openrouter",
        openrouter_api_key="test-or-key",
        openrouter_text_models=models,
        max_retries=1,
    )
    base.update(overrides)
    return make_config(**base)


def envelope(content, *, model="test-vendor/actual-served", finish_reason="stop",
            native_finish_reason="STOP", generation_id="gen-1", **usage_overrides):
    usage = {
        "prompt_tokens": 500, "completion_tokens": 120, "total_tokens": 620,
        "cost": 0.0002, "completion_tokens_details": {"reasoning_tokens": 10},
    }
    usage.update(usage_overrides)
    return {
        "id": generation_id,
        "model": model,
        "choices": [{
            "finish_reason": finish_reason,
            "native_finish_reason": native_finish_reason,
            "message": {"content": content},
        }],
        "usage": usage,
    }


def error_envelope(code, generation_id="gen-err"):
    return {"id": generation_id, "error": {"code": code, "message": "synthetic"}}


class Recorder:
    """Callable seam replacement for openrouter_client._chat_completion."""

    def __init__(self, responses=None):
        self.calls = []
        self.responses = list(responses or [])

    def __call__(self, cfg, *, model, messages, response_format=None, max_tokens, timeout=None):
        self.calls.append({"model": model})
        if not self.responses:
            raise AssertionError("provider called more times than expected")
        item = self.responses.pop(0)
        if isinstance(item, BaseException):
            raise item
        return item


@pytest.fixture
def text_pdf(pdf_factory):
    return Path(pdf_factory([("text", TEXT_BODY)], name="text.pdf"))


def rate_limit_error():
    return ProviderError("OpenRouter request failed (HTTP 429)",
                         category="rate_limited", http_status=429)


# --- A: tier 1 succeeds, tiers 2+ never called -------------------------------

class TestTierOneSucceeds:
    def test_a_tier_1_succeeds_no_other_tier_called(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(3)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1"]
        assert result.provider == "openrouter"
        assert result.needs_review is False


# --- B-H: tier 1 fails in each way, tier 2 succeeds --------------------------

class TestEscalationCauses:
    def test_b_provider_error_escalates(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        rec = Recorder([rate_limit_error(), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-2"]
        assert result.needs_review is False
        assert result.model == "test-vendor/actual-served"

    def test_c_malformed_json_repair_succeeds_tier_2_not_called(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = ladder_cfg(2)
        rec = Recorder([envelope("not valid json"), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-1"]
        assert result.needs_review is False

    def test_d_repair_fails_tier_2_succeeds(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        rec = Recorder([
            envelope("not valid json"), envelope("still not valid json"),  # tier 1 + its repair
            envelope(invoice_json()),  # tier 2
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == [
            "test-vendor/tier-1", "test-vendor/tier-1", "test-vendor/tier-2",
        ]
        assert result.needs_review is False

    def test_e_schema_invalid_multiple_fields_escalates(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        rec = Recorder([envelope(bad), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-2"]
        assert result.needs_review is False

    def test_f_missing_single_required_field_escalates(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        bad = json.dumps({
            "invoice_number": "INV-9", "invoice_date": "2026-01-01",
            "seller_name": "Acme", "total_amount": 500,
        })  # currency omitted
        rec = Recorder([envelope(bad), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-2"]
        assert result.needs_review is False

    def test_g_truncated_escalates(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        rec = Recorder([
            envelope('{"invoice_number": "INV-1", "line_', finish_reason="length"),
            envelope(invoice_json()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-2"]
        assert result.needs_review is False

    def test_h_empty_content_escalates(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        rec = Recorder([envelope(None), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-2"]
        assert result.needs_review is False

    def test_i_embedded_error_envelope_escalates(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        rec = Recorder([error_envelope(429), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-2"]
        assert result.needs_review is False


# --- J: all models fail -------------------------------------------------------

class TestAllModelsFail:
    def test_j_all_models_fail_controlled_needs_review_batch_continues(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = ladder_cfg(3)
        build_pdf(tmp_path / "a_bad.pdf", [("text", TEXT_BODY)])
        build_pdf(tmp_path / "b_good.pdf", [("text", TEXT_BODY)])
        responses = [
            rate_limit_error(), rate_limit_error(), rate_limit_error(),  # a_bad: all 3 tiers fail
            envelope(invoice_json()),  # b_good: succeeds on tier 1
        ]
        monkeypatch.setattr(openrouter_client, "_chat_completion", Recorder(responses))

        results = process_directory(tmp_path, cfg, logger)

        assert len(results) == 2
        bad = next(r for r in results if r.source_file == "a_bad.pdf")
        good = next(r for r in results if r.source_file == "b_good.pdf")
        assert bad.error is True and bad.needs_review is True
        assert "NameError" not in bad.review_reason
        assert "tier-1" in bad.review_reason and "tier-3" in bad.review_reason
        assert good.error is False


# --- K: actual model differs from requested, both recorded ------------------

class TestProvenance:
    def test_k_actual_model_recorded_alongside_requested(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(1)
        rec = Recorder([envelope(invoice_json(), model="test-vendor/really-served")])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.model == "test-vendor/really-served"
        accepted = [r for r in result.usage_records if r.accepted]
        assert len(accepted) == 1
        assert accepted[0].requested_model == "test-vendor/tier-1"
        assert accepted[0].actual_model == "test-vendor/really-served"


# --- L: usage rows include primary, repair, and escalation -------------------

class TestUsageRowsAcrossAttemptTypes:
    def test_l_primary_repair_escalation_all_recorded(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        rec = Recorder([
            envelope("not valid json"),          # tier-1 primary (rejected)
            envelope("still not valid json"),     # tier-1 repair (rejected)
            envelope(invoice_json()),             # tier-2 escalation (accepted)
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        types = [r.attempt_type for r in result.usage_records]
        assert types == ["primary", "repair", "escalation"]
        assert [r.accepted for r in result.usage_records] == [False, False, True]
        assert [r.ladder_index for r in result.usage_records] == [0, 0, 1]


# --- M/N: cost sums and by-model summaries -----------------------------------

class TestCostAccounting:
    def test_m_cost_totals_are_exact_decimal_sums(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(2)
        rec = Recorder([
            envelope("not valid json", cost=0.0001),
            envelope(invoice_json(), cost=0.0002),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        total = sum((r.cost_usd or Decimal("0")) for r in result.usage_records)
        assert total == Decimal("0.0001") + Decimal("0.0002")
        assert isinstance(total, Decimal)

    def test_n_cost_by_model_and_accepted_by_model(self, logger, tmp_path, monkeypatch):
        from invoice_extractor.usage import summarize_usage
        cfg = ladder_cfg(2)
        build_pdf(tmp_path / "inv.pdf", [("text", TEXT_BODY)])
        # Schema-invalid (not malformed JSON) so tier-1 is rejected in ONE
        # call, no repair attempt to account for.
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        rec = Recorder([
            envelope(bad, cost=0.001, model="test-vendor/tier-1-served"),
            envelope(invoice_json(), cost=0.002, model="test-vendor/tier-2-served"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)
        all_records = [r for res in results for r in res.usage_records]
        s = summarize_usage(all_records)

        assert s.cost_by_model["test-vendor/tier-1"] == Decimal("0.001")
        assert s.cost_by_model["test-vendor/tier-2"] == Decimal("0.002")
        assert s.accepted_by_model == {"test-vendor/tier-2": 1}


# --- O/P/Q: budget and attempt controls --------------------------------------

class TestBudgetControls:
    def test_o_model_attempt_cap_stops_further_models(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(3, max_model_attempts_per_file=1)
        rec = Recorder([rate_limit_error()])  # only tier-1 should ever be attempted
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1"]
        assert result.error is True
        assert "attempt cap" in result.review_reason

    def test_p_file_cost_budget_stops_escalation_continues_next_file(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = ladder_cfg(2, max_cost_usd_per_file=Decimal("0.0005"))
        build_pdf(tmp_path / "a_expensive.pdf", [("text", TEXT_BODY)])
        build_pdf(tmp_path / "b_good.pdf", [("text", TEXT_BODY)])
        # Schema-invalid (not malformed JSON) so tier-1 is rejected in ONE
        # call - no repair attempt to account for - and costs more than the
        # whole per-file budget, so tier-2 must never be tried for file "a".
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        responses = [
            envelope(bad, cost=0.001),  # a: tier-1 alone costs more than the whole budget
            envelope(invoice_json()),  # b: succeeds on tier 1 (its own fresh per-file budget)
        ]
        monkeypatch.setattr(openrouter_client, "_chat_completion", Recorder(responses))

        results = process_directory(tmp_path, cfg, logger)

        a = next(r for r in results if r.source_file == "a_expensive.pdf")
        b = next(r for r in results if r.source_file == "b_good.pdf")
        assert a.error is True
        assert "cost budget" in a.review_reason
        assert b.error is False  # other files continue

    def test_q_run_wide_budget_stops_new_calls_writes_partial_outputs(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = ladder_cfg(1, max_cost_usd_per_run=Decimal("0.0001"))
        build_pdf(tmp_path / "a_first.pdf", [("text", TEXT_BODY)])
        build_pdf(tmp_path / "b_second.pdf", [("text", TEXT_BODY)])
        rec = Recorder([envelope(invoice_json(), cost=0.0005)])  # a alone exceeds run budget
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        assert len(rec.calls) == 1  # b never even attempted
        a = next(r for r in results if r.source_file == "a_first.pdf")
        b = next(r for r in results if r.source_file == "b_second.pdf")
        assert a.error is False  # a itself succeeded before the budget was hit
        assert b.error is True
        assert "run-wide" in b.review_reason

        # partial output still writes cleanly
        path = export_workbook(results, tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]
        usage_path = write_usage_csv(
            [r for res in results for r in res.usage_records],
            usage_csv_path(tmp_path / "out.xlsx"),
        )
        assert usage_path.exists()


# --- A-E: run-wide budget enforced MID-ladder, not just between files --------

class TestRunWideBudgetMidLadder:
    """The run-wide budget must be checked before every primary/repair/
    escalation call - not only once per file - so a single file's ladder can
    never overshoot past the point its own response revealed the crossing."""

    def test_a_run_budget_crossed_by_accepted_tier1_stops_further_calls(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = ladder_cfg(3, max_cost_usd_per_run=Decimal("0.0001"))
        build_pdf(tmp_path / "a_first.pdf", [("text", TEXT_BODY)])
        build_pdf(tmp_path / "b_second.pdf", [("text", TEXT_BODY)])
        # tier-1 succeeds outright but its own cost alone crosses the run
        # budget - no repair/escalation would have applied anyway (the ladder
        # already stops on first success), but file b must get a controlled,
        # no-call review row.
        rec = Recorder([envelope(invoice_json(), cost=0.0005)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1"]
        a = next(r for r in results if r.source_file == "a_first.pdf")
        b = next(r for r in results if r.source_file == "b_second.pdf")
        assert a.error is False
        assert b.error is True
        assert "run-wide" in b.review_reason

    def test_b_run_budget_crossed_by_rejected_tier1_stops_tier2(
        self, logger, text_pdf, monkeypatch
    ):
        # Note: process_file only ever consults a run_budget it is explicitly
        # given - process_directory is what constructs one for a real run
        # (see test_a); here we construct one directly to unit-test the
        # mid-ladder checkpoint in isolation.
        cfg = ladder_cfg(3)
        run_budget = RunBudget(Decimal("0.0001"))
        # Schema-invalid (not malformed JSON) so tier-1 rejects in exactly
        # ONE call, with cost already known and crossing the run budget.
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        rec = Recorder([
            envelope(bad, cost=0.0005),   # tier-1: rejected, crosses run budget
            envelope(invoice_json()),      # tier-2: must NEVER be called
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger, run_budget=run_budget)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1"]
        assert result.error is True
        assert "run-wide" in result.review_reason

    def test_c_run_budget_crossed_by_malformed_primary_stops_repair(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = ladder_cfg(2)
        run_budget = RunBudget(Decimal("0.0001"))
        rec = Recorder([
            envelope("not valid json", cost=0.0005),  # primary: malformed, crosses run budget
            envelope("still not valid json"),           # repair: must NEVER be called
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger, run_budget=run_budget)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1"]
        assert result.error is True
        assert "run-wide" in result.review_reason

    def test_d_run_budget_not_reached_normal_repair_and_escalation_continue(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = ladder_cfg(2)
        run_budget = RunBudget(Decimal("10"))  # generous, never reached
        rec = Recorder([
            envelope("not valid json", cost=0.0001),  # primary: malformed
            envelope(invoice_json(), cost=0.0001),      # repair: succeeds
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger, run_budget=run_budget)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1", "test-vendor/tier-1"]
        assert result.error is False
        assert result.needs_review is False

    def test_e_file_budget_and_run_budget_whichever_tighter_fires_first(
        self, logger, text_pdf, monkeypatch
    ):
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})

        # File budget tighter than run budget -> file-budget reason, not run-wide.
        cfg_file_tighter = ladder_cfg(2, max_cost_usd_per_file=Decimal("0.0001"))
        run_budget_loose = RunBudget(Decimal("10"))
        rec_file = Recorder([envelope(bad, cost=0.0005)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec_file)
        result_file = process_file(text_pdf, cfg_file_tighter, logger, run_budget=run_budget_loose)
        assert [c["model"] for c in rec_file.calls] == ["test-vendor/tier-1"]
        assert result_file.error is True
        assert "cost budget" in result_file.review_reason
        assert "run-wide" not in result_file.review_reason

        # Run budget tighter than file budget -> run-wide reason.
        cfg_run_tighter = ladder_cfg(2, max_cost_usd_per_file=Decimal("10"))
        run_budget_tight = RunBudget(Decimal("0.0001"))
        rec_run = Recorder([envelope(bad, cost=0.0005)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec_run)
        result_run = process_file(text_pdf, cfg_run_tighter, logger, run_budget=run_budget_tight)
        assert [c["model"] for c in rec_run.calls] == ["test-vendor/tier-1"]
        assert result_run.error is True
        assert "run-wide" in result_run.review_reason


# --- R: missing usage cost policy --------------------------------------------

class TestMissingCostPolicy:
    def test_r_missing_cost_does_not_block_escalation_or_budget(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = ladder_cfg(2, max_cost_usd_per_file=Decimal("0.01"))
        raw = envelope("not valid json")
        del raw["usage"]["cost"]  # OpenRouter did not report cost for this attempt
        rec = Recorder([raw, envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.needs_review is False  # missing cost never blocked escalation
        rejected = [r for r in result.usage_records if not r.accepted][0]
        assert rejected.cost_usd is None


# --- S: direct gateway unaffected ---------------------------------------------

class TestDirectGatewayUnaffected:
    def test_s_direct_gateway_never_touches_openrouter_or_ladder(
        self, cfg, logger, text_pdf, monkeypatch
    ):
        assert cfg.llm_gateway == "direct"
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        monkeypatch.setattr(gemini_client, "_generate", lambda c, m, ct: invoice_json())

        result = process_file(text_pdf, cfg, logger)

        assert result.provider == "gemini"
        assert rec.calls == []
        assert result.usage_records == []


# --- T: three-sheet workbook contract ----------------------------------------

class TestWorkbookContract:
    def test_t_three_sheet_contract_unchanged_with_ladder(
        self, logger, text_pdf, monkeypatch, tmp_path
    ):
        cfg = ladder_cfg(2)
        rec = Recorder([rate_limit_error(), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)
        path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)

        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]


# --- W: one failed PDF never crashes the batch --------------------------------

class TestBatchContinuation:
    def test_w_one_failed_pdf_never_crashes_batch(self, logger, tmp_path, monkeypatch):
        cfg = ladder_cfg(1)
        build_pdf(tmp_path / "a_bad.pdf", [("text", TEXT_BODY)])
        build_pdf(tmp_path / "b_good.pdf", [("text", TEXT_BODY)])
        responses = [rate_limit_error(), envelope(invoice_json())]
        monkeypatch.setattr(openrouter_client, "_chat_completion", Recorder(responses))

        results = process_directory(tmp_path, cfg, logger)  # must not raise

        assert len(results) == 2


# --- X: vision remains unsupported under openrouter in M3 --------------------

class TestVisionStillUnsupported:
    def test_x_vision_under_openrouter_still_explicitly_rejected(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(2)
        scan_pdf = Path(pdf_factory([("image",)], name="scan.pdf"))
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))

        result = process_file(scan_pdf, cfg, logger)

        assert or_calls == []
        assert result.error is True
        assert "not implemented" in result.review_reason


# --- Y: empty/malformed model list fails clearly ------------------------------

class TestConfigStillFailsClearly:
    def test_y_empty_model_list_fails_clearly(self, logger, text_pdf, monkeypatch):
        cfg = ladder_cfg(0)  # no models configured at all
        assert cfg.openrouter_text_models == ()

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert "OPENROUTER_TEXT_MODELS" in result.review_reason


# --- Z: maximum HTTP-call bound is enforced -----------------------------------

class TestMaxCallBound:
    def test_z_transient_only_failures_bounded_by_max_retries_one_model(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = ladder_cfg(1, max_retries=3)
        # Every attempt fails at the transport layer (never reaches JSON
        # parsing), so no repair is ever attempted - bound is exactly
        # MAX_RETRIES for a single model with no successful parse.
        rec = Recorder([rate_limit_error()] * 3)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 3  # exactly MAX_RETRIES, no more
        assert result.error is True

    def test_z_full_primary_plus_repair_bound_across_models(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = ladder_cfg(2, max_retries=1)
        # Worst case per model with max_retries=1: 1 primary + 1 repair = 2.
        # 2 models -> bound is 2*2 = 4 total calls when both fully exhaust.
        rec = Recorder([
            envelope("not valid json"), envelope("still not valid json"),
            envelope("not valid json"), envelope("still not valid json"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 4  # exactly the documented bound: 2 models * 2
        assert result.error is True


# --- Privacy regression: ladder + usage CSV -----------------------------------

class TestPrivacyRegressionAcrossLadderAndUsageCsv:
    """Distinctive fake API key, invoice text, malformed model output, and
    provider error message/metadata - verified absent from normal logs,
    review_reason, the usage CSV, and the exported workbook."""

    def test_no_sensitive_content_anywhere_across_a_full_escalation(
        self, logger, text_pdf, monkeypatch, caplog, tmp_path
    ):
        cfg = ladder_cfg(2, openrouter_api_key="SECRET-OR-KEY-M3-XYZ")
        body_marker = "UNIQUE-FAKE-INVOICE-BODY-M3-42"
        error_marker = "FAKE-PROVIDER-ERROR-META-M3-77"
        rec = Recorder([
            envelope(f"not valid json {body_marker}"),           # tier-1 primary
            envelope(f"still not valid json {body_marker}"),      # tier-1 repair
            {"id": "gen-x", "error": {"code": 500, "message": error_marker,
                                      "metadata": {"raw": error_marker}}},  # tier-2
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        with caplog.at_level("DEBUG", logger="invoice_extractor"), \
             caplog.at_level("DEBUG", logger="invoice_extractor_tests"):
            result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        forbidden = ["SECRET-OR-KEY-M3-XYZ", body_marker, error_marker]

        messages = " ".join(r.message for r in caplog.records)
        for secret in forbidden:
            assert secret not in messages, f"leaked into logs: {secret}"
            assert secret not in (result.review_reason or ""), f"leaked into review_reason: {secret}"
        for record in result.usage_records:
            record_repr = repr(record)
            for secret in forbidden:
                assert secret not in record_repr, f"leaked into UsageRecord repr: {secret}"

        wb_path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(wb_path)
        for sheet in ("Invoices", "NeedsReview"):
            cells = " ".join(
                str(c.value) for row in wb[sheet].iter_rows() for c in row if c.value is not None
            )
            for secret in forbidden:
                assert secret not in cells, f"leaked into {sheet}: {secret}"

        usage_path = write_usage_csv(result.usage_records, tmp_path / "out.usage.csv")
        csv_content = usage_path.read_text()
        for secret in forbidden:
            assert secret not in csv_content, f"leaked into usage CSV: {secret}"

    def test_h_run_budget_reason_carries_no_secrets(
        self, logger, text_pdf, monkeypatch
    ):
        # Schema-invalid (single-call) rejection so the crossing is caused by
        # a REJECTED response, not a successful one - the review_reason must
        # still be built only from safe values (model id, configured limit),
        # never from the response body or the API key.
        cfg = ladder_cfg(3, openrouter_api_key="SECRET-OR-KEY-M3-BUDGET")
        run_budget = RunBudget(Decimal("0.0001"))
        body_marker = "UNIQUE-FAKE-INVOICE-BODY-M3-BUDGET"
        bad = json.dumps({"invoice_number": body_marker, "total_amount": None, "currency": None})
        rec = Recorder([envelope(bad, cost=0.0005), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger, run_budget=run_budget)

        assert [c["model"] for c in rec.calls] == ["test-vendor/tier-1"]
        assert result.error is True
        assert "run-wide" in result.review_reason
        forbidden = ["SECRET-OR-KEY-M3-BUDGET", body_marker]
        for secret in forbidden:
            assert secret not in result.review_reason, f"leaked into review_reason: {secret}"

    def test_h_unknown_cost_summary_carries_no_secrets(self):
        from invoice_extractor.provider import ProviderResult
        from invoice_extractor.usage import format_usage_summary, usage_record_from_result

        secret_model = "vendor/SECRET-OR-KEY-M3-MODEL"
        result = ProviderResult(
            requested_model=secret_model, route="text", actual_model=f"{secret_model}-served",
            attempt_type="primary", structured_mode="json_schema",
            input_tokens=10, output_tokens=5, reasoning_tokens=0, total_tokens=15,
            cost_usd=None, finish_reason="stop", native_finish_reason="STOP",
            generation_id="gen-1", latency_ms=10.0,
        )
        record = usage_record_from_result(
            result, run_id="run-1", source_file="inv.pdf", page_range="1",
            ladder_index=0, accepted=True,
        )
        summary = format_usage_summary([record], processed_count=1)

        assert "Requests with unknown cost: 1" in summary
        assert "incomplete" in summary
        for forbidden in ("base64", "prompt", "Authorization", "Bearer"):
            assert forbidden.lower() not in summary.lower()

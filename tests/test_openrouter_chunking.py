"""M3.1: bounded text chunking (MAX_TEXT_PAGES) for large text-native
invoices under LLM_GATEWAY=openrouter - per-chunk ladder execution,
deterministic aggregation reuse, partial-failure retention, rejected-response
usage-metadata preservation, and file/run budget accounting across chunks.

All offline - the autouse network-blocking fixture guards; the only mocked
seam is openrouter_client._chat_completion (mirrors test_openrouter_ladder.py).
"""

import csv
import json
from decimal import Decimal
from pathlib import Path

import openpyxl

from invoice_extractor import gemini_client, openrouter_client
from invoice_extractor.excel_export import export_workbook
from invoice_extractor.pipeline import process_directory, process_file
from invoice_extractor.provider import ProviderError
from invoice_extractor.usage import write_usage_csv

from .conftest import TEXT_BODY, build_pdf, invoice_json, make_config


def ladder_cfg(n_models=3, **overrides):
    models = tuple(f"test-vendor/tier-{i + 1}" for i in range(n_models))
    base = dict(
        llm_gateway="openrouter",
        openrouter_api_key="test-or-key",
        openrouter_text_models=models,
        max_retries=1,
        max_text_pages=2,
    )
    base.update(overrides)
    return make_config(**base)


def envelope(content, *, model="test-vendor/actual-served", finish_reason="stop",
            native_finish_reason="STOP", generation_id="gen-1", **usage_overrides):
    usage = {
        "prompt_tokens": 500, "completion_tokens": 120, "total_tokens": 620,
        "cost": 0.0002, "completion_tokens_details": {"reasoning_tokens": 10},
    }
    usage.update(usage_overrides)
    return {
        "id": generation_id,
        "model": model,
        "choices": [{
            "finish_reason": finish_reason,
            "native_finish_reason": native_finish_reason,
            "message": {"content": content},
        }],
        "usage": usage,
    }


def error_envelope(code, generation_id="gen-err"):
    return {"id": generation_id, "error": {"code": code, "message": "synthetic"}}


class Recorder:
    def __init__(self, responses=None):
        self.calls = []
        self.responses = list(responses or [])

    def __call__(self, cfg, *, model, messages, response_format=None, max_tokens, timeout=None):
        self.calls.append({"model": model})
        if not self.responses:
            raise AssertionError("provider called more times than expected")
        item = self.responses.pop(0)
        if isinstance(item, BaseException):
            raise item
        return item


def rate_limit_error():
    return ProviderError("OpenRouter request failed (HTTP 429)",
                         category="rate_limited", http_status=429)


def multi_page_pdf(pdf_factory, n_pages, name="multi.pdf"):
    """Content is irrelevant here (the model seam is mocked directly) -
    page COUNT is what drives chunk boundaries."""
    specs = [("text", f"{TEXT_BODY} (page {i + 1})") for i in range(n_pages)]
    return Path(pdf_factory(specs, name=name))


def header_only(**overrides):
    data = {"line_items": []}
    data.update(overrides)
    return json.dumps(data)


def line_items_only(items):
    return json.dumps({"line_items": items})


# --- A: one-page document -> one chunk ---------------------------------------

class TestSingleChunk:
    def test_a_one_page_produces_one_chunk_one_call(self, logger, pdf_factory, monkeypatch):
        cfg = ladder_cfg(3)
        pdf = multi_page_pdf(pdf_factory, 1)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1
        assert result.text_chunk_count == 1
        assert result.needs_review is False


# --- Logging/provenance fix: accurate chunk counts in the completion log ----

class TestLogProvenanceChunkCounts:
    """Live-pilot bug: the final per-file log line hardcoded
    vision_chunk_count as 'chunks=%d', so a 3-text-chunk OpenRouter file
    always logged 'chunks=0'. Fixed to report vision_chunks and text_chunks
    separately and accurately."""

    def test_one_page_openrouter_file_logs_text_chunks_1(
        self, logger, pdf_factory, monkeypatch, caplog
    ):
        cfg = ladder_cfg(1)
        pdf = multi_page_pdf(pdf_factory, 1)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        with caplog.at_level("INFO", logger="invoice_extractor_tests"):
            result = process_file(pdf, cfg, logger)

        assert result.text_chunk_count == 1
        done_lines = [r.message for r in caplog.records if ": done in" in r.message]
        assert len(done_lines) == 1
        assert "text_chunks=1" in done_lines[0]
        assert "vision_chunks=0" in done_lines[0]

    def test_three_chunk_openrouter_file_logs_text_chunks_3(
        self, logger, pdf_factory, monkeypatch, caplog
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 6)
        rec = Recorder([envelope(invoice_json())] * 3)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        with caplog.at_level("INFO", logger="invoice_extractor_tests"):
            result = process_file(pdf, cfg, logger)

        assert result.text_chunk_count == 3
        done_lines = [r.message for r in caplog.records if ": done in" in r.message]
        assert "text_chunks=3" in done_lines[0]

    def test_vision_chunk_reporting_remains_accurate(
        self, logger, pdf_factory, monkeypatch, caplog
    ):
        cfg = make_config(llm_gateway="direct", max_vision_pages=2)
        pdf = Path(pdf_factory([("image",)] * 3, name="scan.pdf"))
        monkeypatch.setattr(gemini_client, "_generate", lambda c, m, ct: invoice_json())

        with caplog.at_level("INFO", logger="invoice_extractor_tests"):
            result = process_file(pdf, cfg, logger)

        assert result.vision_chunk_count == 2  # chunks of <=2: [1,2],[3]
        done_lines = [r.message for r in caplog.records if ": done in" in r.message]
        assert "vision_chunks=2" in done_lines[0]
        assert "text_chunks=0" in done_lines[0]

    def test_direct_gateway_text_route_logs_text_chunks_0(
        self, logger, pdf_factory, monkeypatch, caplog
    ):
        cfg = make_config(llm_gateway="direct")
        pdf = multi_page_pdf(pdf_factory, 6)
        monkeypatch.setattr(gemini_client, "_generate", lambda c, m, ct: invoice_json())

        with caplog.at_level("INFO", logger="invoice_extractor_tests"):
            result = process_file(pdf, cfg, logger)

        assert result.text_chunk_count == 0  # direct gateway never chunks text
        done_lines = [r.message for r in caplog.records if ": done in" in r.message]
        assert "text_chunks=0" in done_lines[0]


# --- B: page-boundary chunk ranges --------------------------------------------

class TestChunkPageRanges:
    def test_b_six_pages_max_two_produces_expected_ranges(self, logger, pdf_factory, monkeypatch):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 6)
        rec = Recorder([envelope(invoice_json())] * 3)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.text_chunk_count == 3
        page_ranges = [r.page_range for r in result.usage_records]
        assert page_ranges == ["1-2", "3-4", "5-6"]


# --- C: every chunk succeeds on tier 1 ----------------------------------------

class TestAllChunksSucceedTierOne:
    def test_c_all_tier1_aggregate_one_invoice_order_preserved_no_escalation(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(3, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 6)
        responses = [
            envelope(line_items_only([
                {"description": "Item A", "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
            envelope(line_items_only([
                {"description": "Item B", "quantity": 1, "unit_price": 20, "amount": 20},
            ])),
            envelope(header_only(
                invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=30,
            )),
        ]
        monkeypatch.setattr(openrouter_client, "_chat_completion", Recorder(responses))

        result = process_file(pdf, cfg, logger)

        descriptions = [it.description for it in result.invoice.line_items]
        assert descriptions == ["Item A", "Item B"]
        assert result.needs_review is False
        assert float(result.invoice.total_amount) == 30.0


# --- D: one chunk escalates, others stay tier 1 -------------------------------

class TestOneChunkEscalates:
    def test_d_one_chunk_escalates_others_tier1_final_succeeds(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(2, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            envelope(header_only(
                invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=10,
            ), model="tier-1-served"),  # chunk 1: tier-1 ok
            # chunk 2 tier-1: transport failure (NOT a missing-header
            # rejection - that's deliberately relaxed at the chunk level now,
            # so this must fail via a mechanism unrelated to header content).
            rate_limit_error(),
            envelope(line_items_only([
                {"description": "Item", "quantity": 1, "unit_price": 10, "amount": 10},
            ]), model="tier-2-served"),  # chunk 2 tier-2: ok
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == [
            "test-vendor/tier-1", "test-vendor/tier-1", "test-vendor/tier-2",
        ]
        assert result.needs_review is False
        actual_models = {r.actual_model for r in result.usage_records if r.accepted}
        assert actual_models == {"tier-1-served", "tier-2-served"}


# --- E: middle chunk fails all models -----------------------------------------

class TestMiddleChunkFails:
    def test_e_middle_chunk_fails_later_chunk_still_runs_partial_retained(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(2, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 6)  # chunks: 1-2, 3-4, 5-6
        rec = Recorder([
            envelope(line_items_only([
                {"description": "First", "quantity": 1, "unit_price": 1, "amount": 1},
            ])),  # chunk 1 ok
            rate_limit_error(), rate_limit_error(),  # chunk 2: both tiers fail
            envelope(header_only(
                invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=1,
            )),  # chunk 3 ok
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert result.error is False  # partial result, not a hard failure
        assert result.failed_pages == [3, 4]
        assert "pages 3-4" in result.review_reason
        descriptions = [it.description for it in result.invoice.line_items]
        assert descriptions == ["First"]  # survivors kept


# --- F: all chunks fail --------------------------------------------------------

class TestAllChunksFail:
    def test_f_all_chunks_fail_null_row_batch_continues(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        build_pdf(tmp_path / "a_bad.pdf", [("text", TEXT_BODY)] * 4)  # 2 chunks
        build_pdf(tmp_path / "b_good.pdf", [("text", TEXT_BODY)])  # 1 chunk
        rec = Recorder([
            rate_limit_error(), rate_limit_error(),  # a: both chunks fail
            envelope(invoice_json()),  # b: succeeds
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        bad = next(r for r in results if r.source_file == "a_bad.pdf")
        good = next(r for r in results if r.source_file == "b_good.pdf")
        assert bad.error is True and bad.needs_review is True
        assert good.error is False


# --- G/H: header distribution vs final hard-required gate (Correction 1) -----

class TestFinalHardRequiredGate:
    def test_g_headers_distributed_across_chunks_final_validation_succeeds(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 6)  # chunks: 1-2, 3-4, 5-6
        rec = Recorder([
            envelope(header_only(
                invoice_number="INV-1", seller_name="Acme", invoice_date="2026-01-01",
            )),  # chunk 1: some headers, no line items
            envelope(line_items_only([
                {"description": "Freight", "quantity": 1, "unit_price": 100, "amount": 100},
            ])),  # chunk 2: line-item-only, NO headers at all
            envelope(header_only(currency="USD", total_amount=100)),  # chunk 3: remaining headers
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.invoice.seller_name == "Acme"
        assert result.invoice.invoice_date == "2026-01-01"
        assert result.invoice.currency == "USD"
        assert float(result.invoice.total_amount) == 100.0
        assert result.needs_review is False  # final gate passes - all 4 required fields present
        assert result.error is False

    def test_h_line_item_only_chunk_accepted_but_final_validation_stays_strict(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            envelope(line_items_only([
                {"description": "A", "quantity": 1, "unit_price": 1, "amount": 1},
            ])),  # chunk 1: line-item-only, no headers - ACCEPTED at chunk level
            envelope(line_items_only([
                {"description": "B", "quantity": 1, "unit_price": 2, "amount": 2},
            ])),  # chunk 2: same - NO chunk ever supplies required headers
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        # Both chunks were accepted (not rejected for missing headers) -
        # confirmed by both line items surviving into the aggregated invoice.
        descriptions = [it.description for it in result.invoice.line_items]
        assert descriptions == ["A", "B"]
        # But the FINAL aggregated invoice still fails the hard-required
        # gate, unweakened - flagged, not silently downgraded, and NOT
        # discarded (existing batch contract: never null out real data).
        assert result.needs_review is True
        assert result.error is False
        assert "missing required fields" in result.review_reason
        for fld in ("invoice_date", "currency", "seller_name", "total_amount"):
            assert fld in result.review_reason


# --- I/J: header conflicts -----------------------------------------------------

class TestHeaderConflicts:
    def test_i_conflicting_header_values_flag_review_no_silent_merge(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(header_only(seller_name="Acme Ltd", total_amount=100,
                                 invoice_date="2026-01-01", currency="USD")),
            envelope(header_only(seller_name="Other Corp", total_amount=100,
                                 invoice_date="2026-01-01", currency="USD")),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "conflict in seller_name" in result.review_reason

    def test_j_conflicting_invoice_numbers_flagged_as_multi_invoice_signal(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(header_only(invoice_number="INV-1", seller_name="Acme",
                                 total_amount=100, invoice_date="2026-01-01", currency="USD")),
            envelope(header_only(invoice_number="INV-2", seller_name="Acme",
                                 total_amount=100, invoice_date="2026-01-01", currency="USD")),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "multiple invoices" in result.review_reason


# --- K: dedup across chunk boundaries ------------------------------------------

class TestDedup:
    def test_k_duplicate_lines_across_chunks_deduped_only_on_strong_evidence(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        dup_item = {"description": "Freight", "quantity": 1, "unit_price": 50, "amount": 50}
        rec = Recorder([
            envelope(line_items_only([dup_item])),
            envelope(line_items_only([dup_item, {  # exact repeat + one new item
                "description": "Handling", "quantity": 1, "unit_price": 5, "amount": 5,
            }])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        descriptions = [it.description for it in result.invoice.line_items]
        assert descriptions == ["Freight", "Handling"]  # exact duplicate dropped, once


# --- L/M: totals + suspicious-row guardrail after aggregation -----------------

class TestFinalValidationChecks:
    def test_l_totals_reconciliation_after_aggregation(self, logger, pdf_factory, monkeypatch):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "A", "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
            envelope(header_only(seller_name="Acme", invoice_date="2026-01-01",
                                 currency="USD", total_amount=999)),  # doesn't reconcile
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "totals inconclusive" in result.review_reason

    def test_m_suspicious_null_amount_line_guardrail_active(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "Description", "quantity": None, "unit_price": None, "amount": None},
                {"description": "Real item", "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
            envelope(header_only(seller_name="Acme", invoice_date="2026-01-01",
                                 currency="USD", total_amount=10)),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "missing an amount" in result.review_reason


# --- N/O/P: rejected-response usage-metadata preservation ---------------------

class TestRejectedResponseUsagePreservation:
    def test_n_truncated_response_preserves_metadata(self, logger, pdf_factory, monkeypatch):
        cfg = ladder_cfg(2, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 2)
        rec = Recorder([
            envelope('{"invoice_number": "INV-1", "line_', finish_reason="length",
                    model="tier-1-served", generation_id="gen-trunc",
                    cost=0.0007, prompt_tokens=111, completion_tokens=222),
            envelope(invoice_json()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is False
        rejected = next(r for r in result.usage_records if not r.accepted)
        assert rejected.rejection_category == "truncated"
        assert rejected.actual_model == "tier-1-served"
        assert rejected.finish_reason == "length"
        assert rejected.input_tokens == 111
        assert rejected.output_tokens == 222
        assert rejected.cost_usd == Decimal("0.0007")
        assert rejected.generation_id == "gen-trunc"

    def test_o_malformed_and_schema_invalid_preserve_usage_metadata(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(2, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 2)
        bad_schema = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        rec = Recorder([
            envelope(bad_schema, model="tier-1-served", generation_id="gen-schema", cost=0.0003),
            envelope(invoice_json()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        rejected = next(r for r in result.usage_records if not r.accepted)
        assert rejected.rejection_category == "missing_required_fields"
        assert rejected.actual_model == "tier-1-served"
        assert rejected.generation_id == "gen-schema"
        assert rejected.cost_usd == Decimal("0.0003")

    def test_p_http_failure_without_envelope_stays_blank_not_fabricated(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(2, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 2)
        rec = Recorder([rate_limit_error(), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        rejected = next(r for r in result.usage_records if not r.accepted)
        assert rejected.rejection_category == "rate_limited"
        for field in ("actual_model", "input_tokens", "output_tokens", "cost_usd",
                      "generation_id", "finish_reason"):
            assert getattr(rejected, field) is None


# --- Q: usage CSV page ranges --------------------------------------------------

class TestUsageCsvPageRanges:
    def test_q_usage_csv_contains_correct_chunk_page_ranges(
        self, logger, pdf_factory, monkeypatch, tmp_path
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 6)
        rec = Recorder([envelope(invoice_json())] * 3)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)
        path = write_usage_csv(result.usage_records, tmp_path / "out.usage.csv")
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))

        assert [r["page_range"] for r in rows] == ["1-2", "3-4", "5-6"]


# --- R: multi-model provenance (Correction 4) ----------------------------------

class TestProvenance:
    def test_r_single_actual_model_across_chunks_recorded_as_is(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "A", "quantity": 1, "unit_price": 1, "amount": 1},
            ]), model="same-served"),
            envelope(header_only(seller_name="Acme", invoice_date="2026-01-01",
                                 currency="USD", total_amount=1), model="same-served"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)
        assert result.model == "same-served"

    def test_r_multiple_actual_models_recorded_as_compact_multiple(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "A", "quantity": 1, "unit_price": 1, "amount": 1},
            ]), model="model-a-served"),
            envelope(header_only(seller_name="Acme", invoice_date="2026-01-01",
                                 currency="USD", total_amount=1), model="model-b-served"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)
        assert result.model == "multiple"
        assert result.provider == "openrouter"
        # Full per-chunk detail must still be recoverable from usage records.
        actual_models = {r.actual_model for r in result.usage_records if r.accepted}
        assert actual_models == {"model-a-served", "model-b-served"}


# --- S: file budget exhaustion mid-document (Correction 3) -------------------

class TestFileBudgetStopsCleanly:
    def test_s_file_budget_reached_stops_remaining_chunks_one_compact_entry(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2, max_cost_usd_per_file=Decimal("0.0004"))
        pdf = multi_page_pdf(pdf_factory, 6)  # chunks: 1-2, 3-4, 5-6
        rec = Recorder([
            envelope(invoice_json(), cost=0.0005),  # chunk 1: alone exceeds the file budget
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1  # chunks 2 and 3 never attempted at all
        assert result.needs_review is True
        assert result.failed_pages == [3, 4, 5, 6]
        # Exactly ONE compact entry for both skipped chunks combined, not one
        # per remaining chunk.
        assert result.review_reason.count("text chunks pages 3-6") == 1
        assert "cost budget" in result.review_reason


# --- T: run budget exhaustion mid-chunk (Correction 3) ------------------------

class TestRunBudgetStopsCleanly:
    def test_t_run_budget_reached_mid_chunk_stops_remaining_chunks_and_files(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2, max_cost_usd_per_run=Decimal("0.0004"))
        build_pdf(tmp_path / "a_first.pdf", [("text", TEXT_BODY)] * 6)  # 3 chunks
        build_pdf(tmp_path / "b_second.pdf", [("text", TEXT_BODY)])  # 1 chunk
        rec = Recorder([
            envelope(invoice_json(), cost=0.0005),  # a, chunk 1: crosses run budget
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        assert len(rec.calls) == 1  # a's chunks 2-3 AND b: zero further calls
        a = next(r for r in results if r.source_file == "a_first.pdf")
        b = next(r for r in results if r.source_file == "b_second.pdf")
        assert a.failed_pages == [3, 4, 5, 6]
        assert a.review_reason.count("text chunks pages 3-6") == 1
        assert b.error is True
        assert "run-wide" in b.review_reason


# --- U: model-attempt cap bounded across the whole file -----------------------

class TestModelAttemptCapAcrossChunks:
    def test_u_attempt_cap_shared_across_chunks_not_reset_per_chunk(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(3, max_text_pages=2, max_model_attempts_per_file=2)
        pdf = multi_page_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            rate_limit_error(),  # chunk 1, tier-1: fails
            rate_limit_error(),  # chunk 1, tier-2: fails (2 attempts used file-wide)
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 2  # chunk 2 never gets its own fresh 2 attempts
        assert result.error is True
        assert "attempt cap" in result.review_reason


# --- V: direct gateway unaffected ----------------------------------------------

class TestDirectGatewayUnaffected:
    def test_v_direct_gateway_never_chunks_text(self, logger, pdf_factory, monkeypatch):
        cfg = make_config(llm_gateway="direct", max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 6)
        calls = []
        monkeypatch.setattr(
            gemini_client, "_generate",
            lambda cfg_, model, contents: (calls.append(1), invoice_json())[1],
        )
        result = process_file(pdf, cfg, logger)
        assert len(calls) == 1  # one combined request for all 6 pages
        assert result.text_chunk_count == 0
        assert result.provider == "gemini"


# --- W: OpenRouter vision requires OPENROUTER_VISION_MODELS (M4) --------------

class TestVisionNeedsOwnModelList:
    def test_w_openrouter_vision_without_models_rejected_cleanly(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1)  # text models only - no vision list configured
        pdf = Path(pdf_factory([("image",)], name="scan.pdf"))
        result = process_file(pdf, cfg, logger)
        assert result.error is True
        assert "OPENROUTER_VISION_MODELS" in result.review_reason


# --- X: three-sheet workbook contract unchanged -------------------------------

class TestWorkbookContract:
    def test_x_three_sheets_with_chunked_results(self, logger, pdf_factory, monkeypatch, tmp_path):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)
        rec = Recorder([envelope(invoice_json())] * 2)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = process_file(pdf, cfg, logger)
        path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]


# --- Y: privacy regression across chunking + budgets + truncation ------------

class TestPrivacyAcrossChunking:
    def test_y_no_sensitive_content_leaks_anywhere(
        self, logger, pdf_factory, monkeypatch, tmp_path, caplog
    ):
        cfg = ladder_cfg(2, max_text_pages=2, openrouter_api_key="SECRET-M31-KEY")
        pdf = multi_page_pdf(pdf_factory, 4)
        body_marker = "UNIQUE-FAKE-BODY-M31"
        error_marker = "FAKE-ERROR-META-M31"
        rec = Recorder([
            envelope(f"not valid json {body_marker}", finish_reason="length"),
            {"id": "gen-x", "error": {"code": 500, "message": error_marker,
                                      "metadata": {"raw": error_marker}}},
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        with caplog.at_level("DEBUG", logger="invoice_extractor"), \
             caplog.at_level("DEBUG", logger="invoice_extractor_tests"):
            result = process_file(pdf, cfg, logger)

        forbidden = ["SECRET-M31-KEY", body_marker, error_marker]
        messages = " ".join(r.message for r in caplog.records)
        for secret in forbidden:
            assert secret not in messages
            assert secret not in (result.review_reason or "")
        for record in result.usage_records:
            for secret in forbidden:
                assert secret not in repr(record)
        wb_path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(wb_path)
        for sheet in ("Invoices", "LineItems", "NeedsReview"):
            cells = " ".join(
                str(c.value) for row in wb[sheet].iter_rows() for c in row if c.value is not None
            )
            for secret in forbidden:
                assert secret not in cells
        usage_path = write_usage_csv(result.usage_records, tmp_path / "out.usage.csv")
        content = usage_path.read_text()
        for secret in forbidden:
            assert secret not in content


# --- Z: one PDF failure never crashes the batch -------------------------------

class TestBatchContinuation:
    def test_z_one_file_all_chunks_fail_batch_continues(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        build_pdf(tmp_path / "a_bad.pdf", [("text", TEXT_BODY)] * 4)
        build_pdf(tmp_path / "b_good.pdf", [("text", TEXT_BODY)])
        rec = Recorder([
            rate_limit_error(), rate_limit_error(),
            envelope(invoice_json()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)
        assert len(results) == 2
        good = next(r for r in results if r.source_file == "b_good.pdf")
        assert good.error is False


# --- Correction 2 / live-pilot fix: line_no vs item_code are distinct fields -

class TestLineNoField:
    def test_printed_line_number_populates_line_no_not_description(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1)
        pdf = multi_page_pdf(pdf_factory, 1)
        payload = invoice_json(line_items=[
            {"line_no": "1", "description": "Ocean freight", "quantity": 1,
             "unit_price": 100.0, "amount": 100.0},
        ])
        rec = Recorder([envelope(payload)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        item = result.invoice.line_items[0]
        assert item.line_no == "1"
        assert item.description == "Ocean freight"


class TestLineNoAndItemCode:
    """Regression suite for the live-pilot bug: a real model put SKU/product
    codes (e.g. '31C207') into line_no because its description was
    ambiguous. item_code is now a dedicated field; line_no's description was
    reworded to explicitly exclude product codes."""

    def test_a_product_code_stored_in_item_code_line_no_null(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1)
        pdf = multi_page_pdf(pdf_factory, 1)
        payload = invoice_json(line_items=[
            {"item_code": "31C207", "description": "Widget", "quantity": 1,
             "unit_price": 10, "amount": 10},
        ])
        rec = Recorder([envelope(payload)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        item = result.invoice.line_items[0]
        assert item.item_code == "31C207"
        assert item.line_no is None
        assert item.description == "Widget"

    def test_b_printed_sequence_plus_product_code_both_populated(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1)
        pdf = multi_page_pdf(pdf_factory, 1)
        payload = invoice_json(line_items=[
            {"line_no": "12", "item_code": "31C207", "description": "Widget",
             "quantity": 1, "unit_price": 10, "amount": 10},
        ])
        rec = Recorder([envelope(payload)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        item = result.invoice.line_items[0]
        assert item.line_no == "12"
        assert item.item_code == "31C207"
        assert item.description == "Widget"

    def test_c_chunked_aggregation_preserves_both_fields(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            envelope(line_items_only([
                {"line_no": "1", "item_code": "31C207", "description": "Widget A",
                 "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
            envelope(line_items_only([
                {"line_no": "2", "item_code": "73SA041601", "description": "Widget B",
                 "quantity": 1, "unit_price": 20, "amount": 20},
            ])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        items = result.invoice.line_items
        assert [(it.line_no, it.item_code) for it in items] == [
            ("1", "31C207"), ("2", "73SA041601"),
        ]

    def test_d_excel_lineitems_contains_both_columns(
        self, logger, pdf_factory, monkeypatch, tmp_path
    ):
        cfg = ladder_cfg(1)
        pdf = multi_page_pdf(pdf_factory, 1)
        payload = invoice_json(line_items=[
            {"line_no": "1", "item_code": "31C207", "description": "Widget",
             "quantity": 1, "unit_price": 10, "amount": 10},
        ])
        rec = Recorder([envelope(payload)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = process_file(pdf, cfg, logger)

        path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        header = [c.value for c in next(wb["LineItems"].iter_rows(max_row=1))]
        assert "line_no" in header
        assert "item_code" in header
        row = next(wb["LineItems"].iter_rows(min_row=2, max_row=2, values_only=True))
        row_dict = dict(zip(header, row))
        assert row_dict["line_no"] == "1"
        assert row_dict["item_code"] == "31C207"

    def test_e_description_does_not_absorb_line_no_or_item_code(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1)
        pdf = multi_page_pdf(pdf_factory, 1)
        payload = invoice_json(line_items=[
            {"line_no": "12", "item_code": "31C207", "description": "Widget",
             "quantity": 1, "unit_price": 10, "amount": 10},
        ])
        rec = Recorder([envelope(payload)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        description = result.invoice.line_items[0].description
        assert "12" not in description
        assert "31C207" not in description

    def test_f_dedup_ignores_line_no_and_item_code_not_broadened(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1, max_text_pages=2)
        pdf = multi_page_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            envelope(line_items_only([
                {"line_no": "1", "item_code": "31C207", "description": "Widget",
                 "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
            # Same strong-evidence fields (description/qty/unit_price/amount)
            # but DIFFERENT line_no/item_code - must still dedup (those two
            # fields are not part of the duplicate key; this is not broadened
            # to require them to match too).
            envelope(line_items_only([
                {"line_no": "99", "item_code": "DIFFERENT-CODE", "description": "Widget",
                 "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(result.invoice.line_items) == 1  # deduped despite differing line_no/item_code
        assert result.invoice.line_items[0].item_code == "31C207"  # first-seen kept

    def test_g_invoices_without_line_numbers_remain_valid(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = ladder_cfg(1)
        pdf = multi_page_pdf(pdf_factory, 1)
        rec = Recorder([envelope(invoice_json())])  # no line_no/item_code keys at all
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        item = result.invoice.line_items[0]
        assert item.line_no is None
        assert item.item_code is None
        assert result.needs_review is False

    def test_h_privacy_unchanged_with_item_code_present(
        self, logger, pdf_factory, monkeypatch, tmp_path
    ):
        cfg = ladder_cfg(2, openrouter_api_key="SECRET-ITEMCODE-KEY")
        pdf = multi_page_pdf(pdf_factory, 1)
        body_marker = "UNIQUE-FAKE-BODY-ITEMCODE"
        rec = Recorder([
            envelope(f"not valid json {body_marker}"),
            envelope(f"still not valid json {body_marker}"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        forbidden = ["SECRET-ITEMCODE-KEY", body_marker]
        for secret in forbidden:
            assert secret not in (result.review_reason or "")
        for record in result.usage_records:
            for secret in forbidden:
                assert secret not in repr(record)

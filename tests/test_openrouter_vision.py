"""M4: OpenRouter VISION extraction - per-chunk vision-model ladder over
multimodal (base64 PNG data-URL) requests, with the same escalation, repair,
budget, usage, aggregation, and partial-failure semantics as the text route.

All offline - the autouse network-blocking fixture guards; the only mocked
seams are openrouter_client._chat_completion (provider boundary) and, where
a test must prove rendering did NOT happen, pdf_utils.render_pages_png.
"""

import csv
import json
from decimal import Decimal
from pathlib import Path

import openpyxl

from invoice_extractor import gemini_client, openrouter_client, pdf_utils
from invoice_extractor.excel_export import export_workbook
from invoice_extractor.pipeline import process_directory, process_file
from invoice_extractor.provider import ProviderError
from invoice_extractor.usage import write_usage_csv

from .conftest import TEXT_BODY, build_pdf, invoice_json, make_config


def vision_cfg(n_models=3, **overrides):
    """openrouter gateway with SEPARATE text and vision ladders configured."""
    vision = tuple(f"test-vendor/vis-{i + 1}" for i in range(n_models))
    base = dict(
        llm_gateway="openrouter",
        openrouter_api_key="test-or-key",
        openrouter_text_models=("test-vendor/text-1",),
        openrouter_vision_models=vision,
        max_retries=1,
    )
    base.update(overrides)
    return make_config(**base)


def envelope(content, *, model="test-vendor/vis-served", finish_reason="stop",
            native_finish_reason="STOP", generation_id="gen-1", **usage_overrides):
    usage = {
        "prompt_tokens": 900, "completion_tokens": 150, "total_tokens": 1050,
        "cost": 0.0002, "completion_tokens_details": {"reasoning_tokens": 5},
    }
    usage.update(usage_overrides)
    return {
        "id": generation_id,
        "model": model,
        "choices": [{
            "finish_reason": finish_reason,
            "native_finish_reason": native_finish_reason,
            "message": {"content": content},
        }],
        "usage": usage,
    }


class Recorder:
    """Seam replacement for _chat_completion, recording model AND messages
    so tests can assert on payload structure (image blocks vs text-only)."""

    def __init__(self, responses=None):
        self.calls = []
        self.responses = list(responses or [])

    def __call__(self, cfg, *, model, messages, response_format=None, max_tokens, timeout=None):
        self.calls.append({"model": model, "messages": messages})
        if not self.responses:
            raise AssertionError("provider called more times than expected")
        item = self.responses.pop(0)
        if isinstance(item, BaseException):
            raise item
        return item


def rate_limit_error():
    return ProviderError("OpenRouter request failed (HTTP 429)",
                         category="rate_limited", http_status=429)


def scan_pdf(pdf_factory, n_pages, name="scan.pdf"):
    return Path(pdf_factory([("image",)] * n_pages, name=name))


def header_only(**overrides):
    data = {"line_items": []}
    data.update(overrides)
    return json.dumps(data)


def line_items_only(items):
    return json.dumps({"line_items": items})


def image_blocks(call) -> list:
    """The image_url content blocks of one recorded request's messages."""
    blocks = []
    for message in call["messages"]:
        content = message.get("content")
        if isinstance(content, list):
            blocks.extend(b for b in content if b.get("type") == "image_url")
    return blocks


def render_spy(monkeypatch):
    """Spy on the render boundary; returns the list of per-call page lists."""
    rendered = []
    real = pdf_utils.render_pages_png

    def spy(path, page_numbers, dpi=200):
        rendered.append(list(page_numbers))
        return real(path, page_numbers, dpi)

    monkeypatch.setattr(pdf_utils, "render_pages_png", spy)
    return rendered


# --- A: one image page, one chunk, one tier-1 call ----------------------------

class TestSingleImagePage:
    def test_a_one_page_one_chunk_one_tier1_call_accepted(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(3)
        pdf = scan_pdf(pdf_factory, 1)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/vis-1"]
        assert len(image_blocks(rec.calls[0])) == 1  # exactly the one page image
        assert result.vision_chunk_count == 1
        assert result.provider == "openrouter"
        assert result.extraction_method == "vision"
        assert result.needs_review is False
        assert result.usage_records[0].route == "vision"
        assert result.usage_records[0].accepted is True


# --- B: multi-page image PDF - ranges and original page numbers ---------------

class TestMultiPageRanges:
    def test_b_five_pages_max_two_ranges_and_page_numbers(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 5)
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(invoice_json())] * 3)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert rendered == [[1, 2], [3, 4], [5]]  # original page numbers, in order
        assert result.vision_chunk_count == 3
        assert [r.page_range for r in result.usage_records] == ["1-2", "3-4", "5"]
        assert [len(image_blocks(c)) for c in rec.calls] == [2, 2, 1]


# --- C: all chunks succeed on tier 1 ------------------------------------------

class TestAllChunksTierOne:
    def test_c_aggregate_order_preserved_no_higher_model(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(3, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            envelope(line_items_only([
                {"description": "Item A", "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
            envelope(header_only(
                invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=10,
            )),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/vis-1"] * 2
        descriptions = [it.description for it in result.invoice.line_items]
        assert descriptions == ["Item A"]
        assert result.needs_review is False


# --- D: one chunk escalates, others stay tier 1 -------------------------------

class TestOneChunkEscalates:
    def test_d_only_failing_chunk_uses_tier2(self, logger, pdf_factory, monkeypatch):
        cfg = vision_cfg(2, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            envelope(line_items_only([
                {"description": "Item", "quantity": 1, "unit_price": 10, "amount": 10},
            ]), model="vis-1-served"),        # chunk 1: tier-1 ok
            rate_limit_error(),                 # chunk 2 tier-1: transport failure
            envelope(header_only(
                invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=10,
            ), model="vis-2-served"),           # chunk 2 tier-2: ok
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == [
            "test-vendor/vis-1", "test-vendor/vis-1", "test-vendor/vis-2",
        ]
        assert result.needs_review is False
        accepted_models = {r.actual_model for r in result.usage_records if r.accepted}
        assert accepted_models == {"vis-1-served", "vis-2-served"}


# --- E: one chunk fails all models, later chunks still run --------------------

class TestMiddleChunkFails:
    def test_e_later_chunks_run_partial_retained_safe_reason(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(2, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 6)  # chunks: 1-2, 3-4, 5-6
        rec = Recorder([
            envelope(line_items_only([
                {"description": "First", "quantity": 1, "unit_price": 1, "amount": 1},
            ])),
            rate_limit_error(), rate_limit_error(),  # chunk 2: both tiers fail
            envelope(header_only(
                invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=1,
            )),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert result.error is False  # partial, not a hard failure
        assert result.failed_pages == [3, 4]
        assert "pages 3-4" in result.review_reason
        assert [it.description for it in result.invoice.line_items] == ["First"]


# --- F: all chunks fail -> null review row, batch continues -------------------

class TestAllChunksFail:
    def test_f_null_row_batch_continues(self, logger, tmp_path, monkeypatch):
        cfg = vision_cfg(1, max_vision_pages=2)
        build_pdf(tmp_path / "a_bad.pdf", [("image",)] * 4)   # 2 chunks
        build_pdf(tmp_path / "b_good.pdf", [("image",)])       # 1 chunk
        rec = Recorder([
            rate_limit_error(), rate_limit_error(),  # a: both chunks fail
            envelope(invoice_json()),                 # b: succeeds
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        bad = next(r for r in results if r.source_file == "a_bad.pdf")
        good = next(r for r in results if r.source_file == "b_good.pdf")
        assert bad.error is True and bad.needs_review is True
        assert good.error is False


# --- G/H: chunk-level relaxation vs final hard-required gate ------------------

class TestFinalHardRequiredGate:
    def test_g_headers_distributed_across_chunks_final_passes(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 6)  # chunks: 1-2, 3-4, 5-6
        rec = Recorder([
            envelope(header_only(invoice_number="INV-1", seller_name="Acme",
                                 invoice_date="2026-01-01")),
            envelope(line_items_only([
                {"description": "Freight", "quantity": 1, "unit_price": 100, "amount": 100},
            ])),  # line-item-only middle chunk, NO headers
            envelope(header_only(currency="USD", total_amount=100)),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.invoice.seller_name == "Acme"
        assert result.invoice.currency == "USD"
        assert float(result.invoice.total_amount) == 100.0
        assert result.needs_review is False

    def test_h_line_item_only_chunks_accepted_final_stays_strict(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rec = Recorder([
            envelope(line_items_only([
                {"description": "A", "quantity": 1, "unit_price": 1, "amount": 1},
            ])),
            envelope(line_items_only([
                {"description": "B", "quantity": 1, "unit_price": 2, "amount": 2},
            ])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        # Both chunks accepted (line items retained)...
        assert [it.description for it in result.invoice.line_items] == ["A", "B"]
        # ...but the aggregated invoice still fails the unweakened final gate.
        assert result.needs_review is True
        assert result.error is False
        assert "missing required fields" in result.review_reason
        for fld in ("invoice_date", "currency", "seller_name", "total_amount"):
            assert fld in result.review_reason


# --- I/J: header conflicts across vision chunks -------------------------------

class TestHeaderConflicts:
    def test_i_conflicting_headers_flagged_never_silent(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(header_only(seller_name="Acme Ltd", total_amount=100,
                                 invoice_date="2026-01-01", currency="USD")),
            envelope(header_only(seller_name="Other Corp", total_amount=100,
                                 invoice_date="2026-01-01", currency="USD")),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "conflict in seller_name" in result.review_reason

    def test_j_conflicting_invoice_numbers_multi_invoice_signal(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(header_only(invoice_number="INV-1", seller_name="Acme",
                                 total_amount=100, invoice_date="2026-01-01", currency="USD")),
            envelope(header_only(invoice_number="INV-2", seller_name="Acme",
                                 total_amount=100, invoice_date="2026-01-01", currency="USD")),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is True
        assert "multiple invoices" in result.review_reason


# --- K: strong-evidence dedup only --------------------------------------------

class TestDedup:
    def test_k_exact_duplicate_across_chunks_dropped_once(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)
        dup = {"description": "Freight", "quantity": 1, "unit_price": 50, "amount": 50}
        rec = Recorder([
            envelope(line_items_only([dup])),
            envelope(line_items_only([dup, {
                "description": "Handling", "quantity": 1, "unit_price": 5, "amount": 5,
            }])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [it.description for it in result.invoice.line_items] == ["Freight", "Handling"]


# --- L: line_no and item_code stay separate ------------------------------------

class TestLineNoItemCode:
    def test_l_no_field_crossover(self, logger, pdf_factory, monkeypatch):
        cfg = vision_cfg(1)
        pdf = scan_pdf(pdf_factory, 1)
        payload = invoice_json(line_items=[
            {"line_no": "12", "item_code": "31C207", "description": "Widget",
             "quantity": 1, "unit_price": 100.0, "amount": 100.0},
        ])
        rec = Recorder([envelope(payload)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        item = result.invoice.line_items[0]
        assert item.line_no == "12"
        assert item.item_code == "31C207"
        assert item.description == "Widget"


# --- M/N/O: rejected-response usage preservation ------------------------------

class TestUsagePreservation:
    def test_m_truncated_vision_completion_preserves_metadata(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(2)
        pdf = scan_pdf(pdf_factory, 1)
        rec = Recorder([
            envelope('{"invoice_number": "INV-1", "line_', finish_reason="length",
                    model="vis-1-served", generation_id="gen-trunc",
                    cost=0.0009, prompt_tokens=800, completion_tokens=150),
            envelope(invoice_json()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        rejected = next(r for r in result.usage_records if not r.accepted)
        assert rejected.route == "vision"
        assert rejected.rejection_category == "truncated"
        assert rejected.actual_model == "vis-1-served"
        assert rejected.finish_reason == "length"
        assert rejected.input_tokens == 800
        assert rejected.output_tokens == 150
        assert rejected.cost_usd == Decimal("0.0009")
        assert rejected.generation_id == "gen-trunc"

    def test_n_schema_invalid_and_malformed_preserve_metadata(
        self, logger, pdf_factory, monkeypatch
    ):
        # Single chunk -> hard-required enforced at chunk level, so a
        # schema-invalid response rejects in ONE call with metadata intact.
        cfg = vision_cfg(2)
        pdf = scan_pdf(pdf_factory, 1)
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        rec = Recorder([
            envelope(bad, model="vis-1-served", generation_id="gen-schema", cost=0.0003),
            envelope(invoice_json()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        rejected = next(r for r in result.usage_records if not r.accepted)
        assert rejected.rejection_category == "missing_required_fields"
        assert rejected.actual_model == "vis-1-served"
        assert rejected.generation_id == "gen-schema"
        assert rejected.cost_usd == Decimal("0.0003")

    def test_n_malformed_json_both_attempts_preserve_metadata(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(2)
        pdf = scan_pdf(pdf_factory, 1)
        rec = Recorder([
            envelope("not valid json", model="vis-1-served", cost=0.0001),
            envelope("still not valid json", model="vis-1-served", cost=0.0002),
            envelope(invoice_json()),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        malformed = [r for r in result.usage_records
                     if r.rejection_category == "malformed_json"]
        assert len(malformed) == 2
        assert all(r.actual_model == "vis-1-served" for r in malformed)
        assert {r.cost_usd for r in malformed} == {Decimal("0.0001"), Decimal("0.0002")}

    def test_o_http_failure_stays_blank_never_fabricated(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(2)
        pdf = scan_pdf(pdf_factory, 1)
        rec = Recorder([rate_limit_error(), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        rejected = next(r for r in result.usage_records if not r.accepted)
        assert rejected.rejection_category == "rate_limited"
        assert rejected.http_status == 429
        for field in ("actual_model", "input_tokens", "output_tokens", "cost_usd",
                      "generation_id", "finish_reason"):
            assert getattr(rejected, field) is None


# --- P: repair - one max, text-only, page range retained -----------------------

class TestRepair:
    def test_p_one_repair_text_only_no_image_resend(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 2)
        rec = Recorder([
            envelope("not valid json"),      # primary: malformed
            envelope(invoice_json()),         # repair: recovers
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/vis-1"] * 2
        assert len(image_blocks(rec.calls[0])) == 2   # primary carries the images
        assert image_blocks(rec.calls[1]) == []        # repair is TEXT-ONLY
        assert result.needs_review is False
        repair_record = next(r for r in result.usage_records if r.attempt_type == "repair")
        assert repair_record.page_range == "1-2"
        assert repair_record.accepted is True


# --- Q: usage CSV - route, ranges, no content ----------------------------------

class TestUsageCsv:
    def test_q_route_vision_ranges_no_image_or_prompt_content(
        self, logger, pdf_factory, monkeypatch, tmp_path
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)
        rec = Recorder([envelope(invoice_json())] * 2)
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = process_file(pdf, cfg, logger)

        path = write_usage_csv(result.usage_records, tmp_path / "out.usage.csv")
        with open(path, newline="", encoding="utf-8") as f:
            rows = list(csv.DictReader(f))

        assert [r["route"] for r in rows] == ["vision", "vision"]
        assert [r["page_range"] for r in rows] == ["1-2", "3-4"]
        content = path.read_text().lower()
        for forbidden in ("base64", "data:image", "image_url",
                          "extraction engine", "invoice text start"):
            assert forbidden not in content


# --- R: multiple accepted models -> compact workbook provenance ----------------

class TestProvenance:
    def test_r_multiple_actual_models_workbook_multiple_csv_full(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "A", "quantity": 1, "unit_price": 1, "amount": 1},
            ]), model="vis-model-a"),
            envelope(header_only(seller_name="Acme", invoice_date="2026-01-01",
                                 currency="USD", total_amount=1), model="vis-model-b"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.model == "multiple"
        assert result.provider == "openrouter"
        accepted_models = {r.actual_model for r in result.usage_records if r.accepted}
        assert accepted_models == {"vis-model-a", "vis-model-b"}

    def test_r_single_actual_model_recorded_as_is(self, logger, pdf_factory, monkeypatch):
        cfg = vision_cfg(1, max_vision_pages=2)
        pdf = scan_pdf(pdf_factory, 4)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "A", "quantity": 1, "unit_price": 1, "amount": 1},
            ]), model="same-vis-served"),
            envelope(header_only(seller_name="Acme", invoice_date="2026-01-01",
                                 currency="USD", total_amount=1), model="same-vis-served"),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)
        assert result.model == "same-vis-served"


# --- S/T/U: budgets across vision chunks --------------------------------------

class TestBudgets:
    def test_s_file_attempt_cap_spans_all_vision_chunks(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(3, max_vision_pages=2, max_model_attempts_per_file=2)
        pdf = scan_pdf(pdf_factory, 4)  # chunks: 1-2, 3-4
        rendered = render_spy(monkeypatch)
        rec = Recorder([rate_limit_error(), rate_limit_error()])  # chunk 1: 2 attempts
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 2      # cap is file-wide: chunk 2 gets NO fresh attempts
        assert rendered == [[1, 2]]     # chunk 2 never rendered either
        assert result.error is True
        assert "attempt cap" in result.review_reason
        assert result.review_reason.count("vision chunks pages 3-4") == 1

    def test_t_file_cost_budget_stops_remaining_chunks_retains_earlier(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2, max_cost_usd_per_file=Decimal("0.0004"))
        pdf = scan_pdf(pdf_factory, 6)  # chunks: 1-2, 3-4, 5-6
        rendered = render_spy(monkeypatch)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "Kept", "quantity": 1, "unit_price": 1, "amount": 1},
            ]), cost=0.0005),  # chunk 1 alone crosses the file budget
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1
        assert rendered == [[1, 2]]  # chunks 2-3 never rendered
        assert [it.description for it in result.invoice.line_items] == ["Kept"]
        assert result.needs_review is True
        assert result.failed_pages == [3, 4, 5, 6]
        assert result.review_reason.count("vision chunks pages 3-6") == 1
        assert "cost budget" in result.review_reason

    def test_u_run_budget_stops_later_chunks_and_files(
        self, logger, tmp_path, monkeypatch
    ):
        cfg = vision_cfg(1, max_vision_pages=2, max_cost_usd_per_run=Decimal("0.0004"))
        build_pdf(tmp_path / "a_first.pdf", [("image",)] * 4)  # 2 chunks
        build_pdf(tmp_path / "b_second.pdf", [("image",)])
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(invoice_json(), cost=0.0005)])  # a chunk 1 crosses run budget
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        assert len(rec.calls) == 1
        assert rendered == [[1, 2]]  # a's chunk 2 and all of b: zero renders
        a = next(r for r in results if r.source_file == "a_first.pdf")
        b = next(r for r in results if r.source_file == "b_second.pdf")
        assert a.review_reason.count("vision chunks pages 3-4") == 1
        assert b.error is True
        assert "run-wide" in b.review_reason


# --- V: mixed text/image file under openrouter --------------------------------

class TestMixedDocument:
    def test_v_text_and_vision_routes_one_aggregate(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(2)
        pdf = Path(pdf_factory([("text", TEXT_BODY), ("image",)], name="mixed.pdf"))
        rec = Recorder([
            envelope(line_items_only([
                {"description": "Text item", "quantity": 1, "unit_price": 10, "amount": 10},
            ]), model="shared-served"),  # text chunk (page 1)
            envelope(header_only(
                invoice_number="INV-1", invoice_date="2026-01-01", currency="USD",
                seller_name="Acme", total_amount=10,
            ), model="shared-served"),   # vision chunk (page 2)
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/text-1", "test-vendor/vis-1"]
        assert image_blocks(rec.calls[0]) == []       # text request: no images
        assert len(image_blocks(rec.calls[1])) == 1   # vision request: page 2 image
        assert result.extraction_method == "mixed"
        assert result.provider == "openrouter"
        assert result.needs_review is False
        assert {r.route for r in result.usage_records} == {"text", "vision"}
        assert result.invoice.seller_name == "Acme"
        assert [it.description for it in result.invoice.line_items] == ["Text item"]

    def test_v_shared_file_budget_spans_text_then_vision(
        self, logger, pdf_factory, monkeypatch
    ):
        # attempt cap 1: the text chunk consumes the file's ONLY attempt, so
        # the vision chunk must be skipped with zero rendering and zero calls.
        cfg = vision_cfg(2, max_model_attempts_per_file=1)
        pdf = Path(pdf_factory([("text", TEXT_BODY), ("image",)], name="mixed.pdf"))
        rendered = render_spy(monkeypatch)
        rec = Recorder([
            envelope(line_items_only([
                {"description": "Text item", "quantity": 1, "unit_price": 10, "amount": 10},
            ])),
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert len(rec.calls) == 1   # vision never called
        assert rendered == []         # vision never rendered
        assert [it.description for it in result.invoice.line_items] == ["Text item"]
        assert result.needs_review is True
        assert result.review_reason.count("vision chunks pages 2") == 1
        assert "attempt cap" in result.review_reason


# --- W/X: other routes unchanged ------------------------------------------------

class TestOtherRoutesUnchanged:
    def test_w_direct_gateway_vision_untouched(self, logger, pdf_factory, monkeypatch):
        cfg = make_config(llm_gateway="direct")
        pdf = scan_pdf(pdf_factory, 1)
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))
        monkeypatch.setattr(gemini_client, "_generate", lambda c, m, ct: invoice_json())

        result = process_file(pdf, cfg, logger)

        assert or_calls == []
        assert result.provider == "gemini"
        assert result.usage_records == []

    def test_x_openrouter_text_route_unchanged_with_vision_models_configured(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(3)
        pdf = Path(pdf_factory([("text", TEXT_BODY)], name="text.pdf"))
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert [c["model"] for c in rec.calls] == ["test-vendor/text-1"]
        assert result.usage_records[0].route == "text"
        assert result.needs_review is False


# --- Y: workbook contract --------------------------------------------------------

class TestWorkbookContract:
    def test_y_exactly_three_sheets(self, logger, pdf_factory, monkeypatch, tmp_path):
        cfg = vision_cfg(1)
        pdf = scan_pdf(pdf_factory, 1)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)
        result = process_file(pdf, cfg, logger)

        path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]


# --- Z: privacy across the vision path ------------------------------------------

class TestPrivacy:
    def test_z_no_sensitive_content_leaks_anywhere(
        self, logger, pdf_factory, monkeypatch, tmp_path, caplog
    ):
        cfg = vision_cfg(2, openrouter_api_key="SECRET-OR-KEY-M4-VIS",
                         debug_artifact_dir=str(tmp_path / "debug"))
        pdf = scan_pdf(pdf_factory, 1)
        body_marker = "UNIQUE-FAKE-INVOICE-BODY-M4-VIS"
        b64_marker = "RkFLRUJBU0U2NC1NNC1WSVM="  # fake base64 fragment in provider body
        error_marker = "FAKE-PROVIDER-ERROR-META-M4-VIS"
        rec = Recorder([
            envelope(f"not valid json {body_marker} {b64_marker}"),        # primary
            envelope(f"still not valid json {body_marker} {b64_marker}"),  # repair
            {"id": "gen-x", "error": {"code": 500, "message": error_marker,
                                      "metadata": {"raw": error_marker}}},  # tier 2
        ])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        with caplog.at_level("DEBUG", logger="invoice_extractor"), \
             caplog.at_level("DEBUG", logger="invoice_extractor_tests"):
            result = process_file(pdf, cfg, logger)

        assert result.error is True
        forbidden = ["SECRET-OR-KEY-M4-VIS", body_marker, b64_marker, error_marker,
                     "data:image/png;base64"]

        messages = " ".join(r.message for r in caplog.records)
        for secret in forbidden:
            assert secret not in messages, f"leaked into logs: {secret}"
            assert secret not in (result.review_reason or ""), \
                f"leaked into review_reason: {secret}"
        for record in result.usage_records:
            for secret in forbidden:
                assert secret not in repr(record), f"leaked into UsageRecord repr: {secret}"

        wb_path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(wb_path)
        for sheet in ("Invoices", "LineItems", "NeedsReview"):
            cells = " ".join(
                str(c.value) for row in wb[sheet].iter_rows() for c in row if c.value is not None
            )
            for secret in forbidden:
                assert secret not in cells, f"leaked into {sheet}: {secret}"

        usage_path = write_usage_csv(result.usage_records, tmp_path / "out.usage.csv")
        content = usage_path.read_text()
        for secret in forbidden:
            assert secret not in content, f"leaked into usage CSV: {secret}"

    def test_z_no_debug_artifact_written_by_default(
        self, logger, pdf_factory, monkeypatch, tmp_path
    ):
        # SAVE_DEBUG_ARTIFACTS defaults OFF (make_config sets False): a
        # missing-required rejection reaches the save_debug_artifact call
        # site, which must write NOTHING - no raw responses, no images.
        cfg = vision_cfg(1, debug_artifact_dir=str(tmp_path / "debug"))
        assert cfg.save_debug_artifacts is False
        pdf = scan_pdf(pdf_factory, 1)
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        rec = Recorder([envelope(bad)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.error is True
        assert not (tmp_path / "debug").exists()


# --- AA: one PDF failure never crashes the batch --------------------------------

class TestBatchResilience:
    def test_aa_unreadable_pdf_plus_good_vision_pdf(self, logger, tmp_path, monkeypatch):
        cfg = vision_cfg(1)
        (tmp_path / "a_corrupt.pdf").write_bytes(b"not a real pdf at all")
        build_pdf(tmp_path / "b_good.pdf", [("image",)])
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        assert len(results) == 2
        corrupt = next(r for r in results if r.source_file == "a_corrupt.pdf")
        good = next(r for r in results if r.source_file == "b_good.pdf")
        assert corrupt.error is True
        assert good.error is False


# --- Adjustment 2: config/budget checks precede rendering -----------------------

class TestChecksPrecedeRendering:
    def test_no_vision_models_means_zero_renders_zero_calls(
        self, logger, pdf_factory, monkeypatch
    ):
        cfg = vision_cfg(0)  # no vision models; text models present but irrelevant
        assert cfg.openrouter_vision_models == ()
        pdf = scan_pdf(pdf_factory, 2)
        rendered = render_spy(monkeypatch)
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))

        result = process_file(pdf, cfg, logger)

        assert rendered == []
        assert or_calls == []
        assert result.error is True
        assert "OPENROUTER_VISION_MODELS" in result.review_reason
        # ONE compact reason for the whole file, not one per chunk/page.
        assert result.review_reason.count("OPENROUTER_VISION_MODELS") == 1

    def test_image_only_file_does_not_demand_text_models(
        self, logger, pdf_factory, monkeypatch
    ):
        # Vision-only configuration must be sufficient for an image-only PDF.
        cfg = vision_cfg(1, openrouter_text_models=())
        pdf = scan_pdf(pdf_factory, 1)
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(pdf, cfg, logger)

        assert result.needs_review is False
        assert "OPENROUTER_TEXT_MODELS" not in (result.review_reason or "")

    def test_pre_exhausted_run_budget_means_zero_renders(
        self, logger, tmp_path, monkeypatch
    ):
        # File a (text) crosses the run budget; file b (image-only) must then
        # be skipped with zero rendering work and zero provider calls.
        cfg = vision_cfg(1, max_cost_usd_per_run=Decimal("0.0004"))
        build_pdf(tmp_path / "a_text.pdf", [("text", TEXT_BODY)])
        build_pdf(tmp_path / "b_scan.pdf", [("image",)])
        rendered = render_spy(monkeypatch)
        rec = Recorder([envelope(invoice_json(), cost=0.0005)])  # a crosses the budget
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        results = process_directory(tmp_path, cfg, logger)

        assert len(rec.calls) == 1  # only a's text call
        assert rendered == []        # b never rendered
        b = next(r for r in results if r.source_file == "b_scan.pdf")
        assert b.error is True
        assert "run-wide" in b.review_reason

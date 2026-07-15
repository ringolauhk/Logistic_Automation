"""M2: one OpenRouter text model wired to the pipeline behind
LLM_GATEWAY=openrouter. Vision under openrouter is explicitly rejected (a
later milestone). All offline - the autouse network-blocking fixture guards;
the only mocked seam is openrouter_client._chat_completion, mirroring how
gemini_client._generate / claude_client._request are mocked elsewhere.
"""

import json
import logging
from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from invoice_extractor import gemini_client, openrouter_client
from invoice_extractor.config import ConfigurationError
from invoice_extractor.excel_export import export_workbook
from invoice_extractor.pipeline import process_directory, process_file
from invoice_extractor.provider import ProviderError

from .conftest import TEXT_BODY, build_pdf, invoice_json, make_config


def openrouter_cfg(**overrides):
    base = dict(
        llm_gateway="openrouter",
        openrouter_api_key="test-or-key",
        openrouter_text_models=("test-vendor/tier-1",),
        max_retries=1,
    )
    base.update(overrides)
    return make_config(**base)


def envelope(content, *, model="test-vendor/actual-served", finish_reason="stop",
            native_finish_reason="STOP", generation_id="gen-1", **usage_overrides):
    usage = {
        "prompt_tokens": 500, "completion_tokens": 120, "total_tokens": 620,
        "cost": 0.0002, "completion_tokens_details": {"reasoning_tokens": 10},
    }
    usage.update(usage_overrides)
    return {
        "id": generation_id,
        "model": model,
        "choices": [{
            "finish_reason": finish_reason,
            "native_finish_reason": native_finish_reason,
            "message": {"content": content},
        }],
        "usage": usage,
    }


class Recorder:
    """Callable seam replacement for openrouter_client._chat_completion."""

    def __init__(self, responses=None):
        self.calls = []
        self.responses = list(responses or [])

    def __call__(self, cfg, *, model, messages, response_format=None, max_tokens, timeout=None):
        self.calls.append(
            {"model": model, "messages": messages, "response_format": response_format}
        )
        if not self.responses:
            raise AssertionError("provider called more times than expected")
        item = self.responses.pop(0)
        if isinstance(item, BaseException):
            raise item
        return item


@pytest.fixture
def text_pdf(pdf_factory):
    return Path(pdf_factory([("text", TEXT_BODY)], name="text.pdf"))


@pytest.fixture
def scan_pdf(pdf_factory):
    return Path(pdf_factory([("image",)], name="scan.pdf"))


# --- A: direct gateway is unaffected -----------------------------------------

class TestDirectGatewayUnaffected:
    def test_direct_gateway_never_touches_openrouter(self, cfg, logger, text_pdf, monkeypatch):
        assert cfg.llm_gateway == "direct"
        or_recorder = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", or_recorder)
        monkeypatch.setattr(gemini_client, "_generate", lambda c, m, ct: invoice_json())

        result = process_file(text_pdf, cfg, logger)

        assert result.provider == "gemini"
        assert result.needs_review is False
        assert or_recorder.calls == []  # OpenRouter boundary never invoked


# --- B/C: success, actual model differs from requested -----------------------

class TestOpenRouterTextSuccess:
    def test_b_success_through_one_model(self, logger, text_pdf, monkeypatch):
        cfg = openrouter_cfg()
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 1
        assert rec.calls[0]["model"] == "test-vendor/tier-1"
        assert result.provider == "openrouter"
        assert result.error is False
        assert result.needs_review is False
        assert result.invoice.invoice_number == "INV-1001"

    def test_c_actual_model_differs_and_provenance_records_both(self):
        cfg = openrouter_cfg()
        rec = Recorder([envelope(invoice_json(), model="test-vendor/actually-served")])
        import unittest.mock as mock
        with mock.patch.object(openrouter_client, "_chat_completion", rec):
            inv, provider_result = openrouter_client.extract_from_text(cfg, TEXT_BODY, label="t")
        assert provider_result.requested_model == "test-vendor/tier-1"
        assert provider_result.actual_model == "test-vendor/actually-served"
        assert provider_result.requested_model != provider_result.actual_model

    def test_c_pipeline_result_model_reflects_actual_served_model(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        rec = Recorder([envelope(invoice_json(), model="test-vendor/actually-served")])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.model == "test-vendor/actually-served"


# --- D: json_schema request shape --------------------------------------------

class TestStructuredOutputShape:
    def test_json_schema_mode_is_default_and_shaped_correctly(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        assert cfg.openrouter_structured_output == "json_schema"
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        process_file(text_pdf, cfg, logger)

        rf = rec.calls[0]["response_format"]
        assert rf["type"] == "json_schema"
        assert rf["json_schema"]["strict"] is True
        assert rf["json_schema"]["name"] == "invoice"
        assert isinstance(rf["json_schema"]["schema"], dict)

    def test_json_object_mode_configurable(self, logger, text_pdf, monkeypatch):
        cfg = openrouter_cfg(openrouter_structured_output="json_object")
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        process_file(text_pdf, cfg, logger)

        assert rec.calls[0]["response_format"] == {"type": "json_object"}

    def test_prompt_only_mode_sends_no_response_format(self, logger, text_pdf, monkeypatch):
        cfg = openrouter_cfg(openrouter_structured_output="prompt_only")
        rec = Recorder([envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        process_file(text_pdf, cfg, logger)

        assert rec.calls[0]["response_format"] is None


# --- E/F/G: local cleanup, one repair, repair-fails ---------------------------

class TestJsonRepairPath:
    def test_e_fenced_json_recovers_locally_no_repair_call(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        rec = Recorder([envelope(f"```json\n{invoice_json()}\n```")])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 1  # local cleanup only, no repair
        assert result.needs_review is False
        assert result.provider == "openrouter"

    def test_e_prose_wrapped_json_recovers_locally_no_repair_call(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        wrapped = f"Here is the extraction:\n{invoice_json()}\nLet me know if anything else is needed."
        rec = Recorder([envelope(wrapped)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 1
        assert result.needs_review is False

    def test_f_malformed_json_triggers_exactly_one_repair_then_succeeds(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        rec = Recorder([envelope("not valid json at all"), envelope(invoice_json())])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 2
        assert result.error is False
        assert result.needs_review is False
        assert result.provider == "openrouter"

    def test_g_repair_also_malformed_is_controlled_failure(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        rec = Recorder([envelope("not valid json"), envelope("still not valid json")])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 2  # exactly one repair, no loop
        assert result.error is True
        assert result.needs_review is True
        assert result.extraction_method == "failed"
        assert "NameError" not in (result.review_reason or "")


# --- H/I: schema-invalid / missing required field -----------------------------

class TestSchemaValidationFailures:
    def test_h_schema_invalid_missing_multiple_fields_fails_safely(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        bad = json.dumps({"invoice_number": "X", "total_amount": None, "currency": None})
        rec = Recorder([envelope(bad), envelope(bad)])  # repair attempted, still bad
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert result.needs_review is True
        assert "missing required fields" in result.review_reason

    def test_i_single_missing_hard_required_field_is_named(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        bad = json.dumps({
            "invoice_number": "INV-9", "invoice_date": "2026-01-01",
            "seller_name": "Acme", "total_amount": 500,
        })  # currency omitted - the only missing hard-required field
        rec = Recorder([envelope(bad), envelope(bad)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert "missing required fields: currency" in result.review_reason


# --- J/K: truncation and empty content ----------------------------------------

class TestTruncationAndEmptyContent:
    def test_j_finish_reason_length_reports_truncation_no_repair_attempt(
        self, logger, text_pdf, monkeypatch
    ):
        cfg = openrouter_cfg()
        rec = Recorder([envelope('{"invoice_number": "INV-1", "line_items": [{"desc',
                                 finish_reason="length")])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert len(rec.calls) == 1  # truncation is never treated as a repair case
        assert result.error is True
        assert result.needs_review is True
        assert "truncated" in result.review_reason
        assert "finish_reason=length" in result.review_reason

    def test_k_empty_content_fails_safely(self, logger, text_pdf, monkeypatch):
        cfg = openrouter_cfg()
        rec = Recorder([envelope(None)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert result.needs_review is True
        assert "empty response" in result.review_reason


# --- L: OpenRouter 429/5xx/transport errors sanitized -------------------------

class TestProviderErrorSanitization:
    def test_l_429_error_sanitized_in_review_reason(self, logger, text_pdf, monkeypatch):
        cfg = openrouter_cfg(max_retries=1)
        rec = Recorder([ProviderError("OpenRouter request failed (HTTP 429)",
                                      category="rate_limited", http_status=429)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert "HTTP 429" in result.review_reason

    def test_l_5xx_error_sanitized(self, logger, text_pdf, monkeypatch):
        cfg = openrouter_cfg(max_retries=1)
        rec = Recorder([ProviderError("OpenRouter request failed (HTTP 503)",
                                      category="server_error", http_status=503)])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert "HTTP 503" in result.review_reason

    def test_l_transport_error_sanitized(self, logger, text_pdf, monkeypatch):
        cfg = openrouter_cfg(max_retries=1)
        rec = Recorder([ProviderError("OpenRouter transport error", category="transport")])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert "transport" in result.review_reason.lower()


# --- M/N: missing key affects only live openrouter extraction -----------------

class TestMissingKeyScope:
    def test_m_missing_key_is_clean_failure_not_crash(self, logger, text_pdf):
        cfg = openrouter_cfg(openrouter_api_key=None)

        result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        assert result.needs_review is True
        assert "OPENROUTER_API_KEY" in result.review_reason

    def test_n_classify_and_doctor_stay_key_free(self, tmp_path, monkeypatch):
        from click.testing import CliRunner
        from invoice_extractor.cli import cli

        monkeypatch.setenv("LLM_GATEWAY", "openrouter")
        monkeypatch.delenv("OPENROUTER_API_KEY", raising=False)
        monkeypatch.delenv("GEMINI_API_KEY", raising=False)
        monkeypatch.delenv("ANTHROPIC_API_KEY", raising=False)
        samples = tmp_path / "samples"
        samples.mkdir()
        build_pdf(samples / "inv.pdf", [("text", TEXT_BODY)])

        classify_result = CliRunner().invoke(cli, ["classify", "--input", str(samples)])
        assert classify_result.exit_code == 0, classify_result.output

        doctor_result = CliRunner().invoke(
            cli, ["doctor", "--input", str(samples), "--output", str(tmp_path / "out")]
        )
        assert doctor_result.exit_code == 0, doctor_result.output
        assert "offline mode" in doctor_result.output


# --- O: privacy regression -----------------------------------------------------

class TestPrivacyRegression:
    def test_o_no_sensitive_content_leaks_anywhere(
        self, logger, text_pdf, monkeypatch, caplog, tmp_path
    ):
        cfg = openrouter_cfg(openrouter_api_key="SECRET-OR-KEY-SHOULD-NOT-LEAK")
        marker = "UNIQUE-FAKE-INVOICE-BODY-MARKER-OR-1"
        rec = Recorder([envelope(f"not valid json {marker}"),
                        envelope(f"still not valid json {marker}")])
        monkeypatch.setattr(openrouter_client, "_chat_completion", rec)

        with caplog.at_level("DEBUG", logger="invoice_extractor"), \
             caplog.at_level("DEBUG", logger="invoice_extractor_tests"):
            result = process_file(text_pdf, cfg, logger)

        assert result.error is True
        forbidden = [marker, "SECRET-OR-KEY-SHOULD-NOT-LEAK"]

        messages = " ".join(r.message for r in caplog.records)
        for secret in forbidden:
            assert secret not in messages, f"leaked into logs: {secret}"
            assert secret not in (result.review_reason or ""), f"leaked into review_reason: {secret}"

        path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)
        for sheet in ("Invoices", "NeedsReview"):
            cells = " ".join(
                str(c.value) for row in wb[sheet].iter_rows() for c in row if c.value is not None
            )
            for secret in forbidden:
                assert secret not in cells, f"leaked into {sheet}: {secret}"


# --- P/Q: batch continuation + three-sheet contract ---------------------------

class TestBatchContinuationAndWorkbookContract:
    def test_p_one_file_failure_does_not_crash_batch(self, logger, tmp_path, monkeypatch):
        cfg = openrouter_cfg()
        build_pdf(tmp_path / "a_bad.pdf", [("text", TEXT_BODY)])
        build_pdf(tmp_path / "b_good.pdf", [("text", TEXT_BODY)])
        responses = [
            envelope("not valid json"), envelope("still not valid json"),  # a_bad: fails
            envelope(invoice_json()),  # b_good: succeeds
        ]
        monkeypatch.setattr(openrouter_client, "_chat_completion", Recorder(responses))

        results = process_directory(tmp_path, cfg, logger)

        assert len(results) == 2
        bad = next(r for r in results if r.source_file == "a_bad.pdf")
        good = next(r for r in results if r.source_file == "b_good.pdf")
        assert bad.error is True
        assert good.error is False and good.provider == "openrouter"

    def test_q_three_sheet_contract_unchanged(self, logger, tmp_path, monkeypatch):
        cfg = openrouter_cfg()
        pdf = Path(build_pdf(tmp_path / "inv.pdf", [("text", TEXT_BODY)]))
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            Recorder([envelope(invoice_json())]))

        result = process_file(pdf, cfg, logger)
        path = export_workbook([result], tmp_path / "out.xlsx")
        wb = openpyxl.load_workbook(path)

        assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]


# --- R: vision under openrouter requires OPENROUTER_VISION_MODELS (M4) --------

class TestOpenRouterVisionNeedsModels:
    def test_r_no_vision_models_fails_clearly_before_rendering(
        self, logger, scan_pdf, monkeypatch
    ):
        # openrouter_cfg() configures TEXT models only - a required vision
        # route with no OPENROUTER_VISION_MODELS must fail safely BEFORE any
        # rendering or HTTP work, with a clean review row and never a silent
        # fallback to direct Gemini/Claude.
        cfg = openrouter_cfg()
        rendered = []
        from invoice_extractor import pdf_utils
        real_render = pdf_utils.render_pages_png

        def spy(path, page_numbers, dpi=200):
            rendered.append(list(page_numbers))
            return real_render(path, page_numbers, dpi)

        monkeypatch.setattr(pdf_utils, "render_pages_png", spy)
        or_calls = []
        monkeypatch.setattr(openrouter_client, "_chat_completion",
                            lambda *a, **k: or_calls.append(1))

        result = process_file(scan_pdf, cfg, logger)

        assert rendered == []  # never rendered - fails before spending anything
        assert or_calls == []  # OpenRouter never called without a vision ladder
        assert result.error is True
        assert result.needs_review is True
        assert "OPENROUTER_VISION_MODELS" in result.review_reason

    def test_r_missing_vision_models_is_a_configuration_error_not_a_crash(
        self, logger, scan_pdf
    ):
        cfg = openrouter_cfg()
        # No mocking at all - process_file must still complete cleanly.
        result = process_file(scan_pdf, cfg, logger)
        assert result.error is True
        assert result.extraction_method == "failed"

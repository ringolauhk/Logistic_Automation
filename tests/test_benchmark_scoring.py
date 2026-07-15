"""M6 benchmark: scoring metrics, review quality, cost/runtime, report
output, privacy, determinism (tests J-AK, AO-AU)."""

import json
from decimal import Decimal

import openpyxl

from invoice_extractor.benchmark.dataset import load_manifest
from invoice_extractor.benchmark.report import (
    REPORT_SHEETS,
    build_json_summary,
    write_json_summary,
    write_report_workbook,
)
from invoice_extractor.benchmark.scoring import parse_review_categories, score_benchmark

from .benchmark_helpers import (
    gt,
    invoice_row,
    line_row,
    manifest_entry,
    usage_csv,
    usage_record,
    write_manifest,
    write_workbook,
)


def _score(tmp_path, cases, ground_truths, invoices, line_items,
           usage_records=None, thresholds=None):
    manifest = write_manifest(tmp_path, cases, ground_truths, thresholds=thresholds)
    wb = write_workbook(tmp_path / "results.xlsx", invoices, line_items)
    usage_path = None
    if usage_records is not None:
        usage_path = usage_csv(tmp_path / "results.usage.csv", usage_records)
    ds = load_manifest(manifest)
    return score_benchmark(ds, wb, usage_path=usage_path)


def case_result(report, case_id):
    return next(c for c in report.cases if c.case_id == case_id)


# --- J/K/L/M/N: header scoring ------------------------------------------------

def test_j_header_exact_match(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"invoice_number": "INV-1", "seller_name": "Acme GmbH",
                                 "currency": "EUR"})},
        [invoice_row("INV-1", "a.pdf", invoice_number="INV-1",
                     seller_name="acme  gmbh", currency="eur")],  # normalized-equal
        [],
    )
    cr = case_result(report, "c1")
    assert cr.exact_header_match is True
    assert report.aggregates["header_micro_accuracy"] == Decimal("1")


def test_k_header_wrong_missing_unexpected_distinctions(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={
            "invoice_number": "INV-1",     # actual differs -> incorrect
            "seller_name": "Acme",         # actual null -> missing
            "po_number": None,             # actual has value -> unexpected
            "currency": "EUR",             # correct
        })},
        [invoice_row("INV-1", "a.pdf", invoice_number="INV-9", seller_name=None,
                     po_number="PO-123", currency="EUR")],
        [],
    )
    cr = case_result(report, "c1")
    assert cr.header_fields["invoice_number"].outcome() == "incorrect"
    assert cr.header_fields["seller_name"].outcome() == "missing"
    assert cr.header_fields["po_number"].outcome() == "unexpected"
    assert cr.header_fields["currency"].outcome() == "correct"


def test_l_ignored_field_excluded_from_denominator(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"invoice_number": "INV-1", "seller_name": "Acme"},
                  ignored_fields=["seller_name"])},
        [invoice_row("INV-1", "a.pdf", invoice_number="INV-1", seller_name="WRONG")],
        [],
    )
    cr = case_result(report, "c1")
    assert "seller_name" not in cr.header_fields        # excluded entirely
    assert "seller_name" in cr.ignored_fields
    assert report.aggregates["header_micro_accuracy"] == Decimal("1")


def test_m_decimal_tolerance_default_one_cent(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"total_amount": "100.00"})},
        [invoice_row("INV-1", "a.pdf", total_amount=100.009)],  # within 0.01
        [],
    )
    assert case_result(report, "c1").header_fields["total_amount"].outcome() == "correct"


def test_m_decimal_beyond_tolerance_incorrect(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"total_amount": "100.00"})},
        [invoice_row("INV-1", "a.pdf", total_amount=100.5)],
        [],
    )
    assert case_result(report, "c1").header_fields["total_amount"].outcome() == "incorrect"


def test_n_date_normalization(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"invoice_date": "2026-07-01"})},
        [invoice_row("INV-1", "a.pdf", invoice_date="2026-07-01")],
        [],
    )
    assert case_result(report, "c1").header_fields["invoice_date"].outcome() == "correct"


# --- not-extractable surfaced (adjustment 3) ----------------------------------

def test_not_extractable_field_surfaced_and_counted(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"invoice_number": "INV-1", "ship_to": "Rotterdam",
                                 "incoterms": "FOB"})},
        [invoice_row("INV-1", "a.pdf", invoice_number="INV-1")],
        [],
    )
    cr = case_result(report, "c1")
    assert set(cr.not_extractable_fields) >= {"ship_to", "incoterms"}
    assert "ship_to" not in cr.header_fields   # never scored
    assert set(report.aggregates["not_extractable_fields"]) >= {"ship_to", "incoterms"}
    assert report.aggregates["not_extractable_field_count"] >= 2


# --- V/W: line detection + per-field ------------------------------------------

def test_v_line_precision_recall_f1(tmp_path):
    # expected 2 lines, actual 2 lines, 1 matched -> prec 0.5, rec 0.5, f1 0.5.
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", line_items=[{"line_no": "1", "amount": "10.00"},
                                    {"line_no": "2", "amount": "20.00"}])},
        [invoice_row("INV-1", "a.pdf")],
        [line_row("INV-1", "a.pdf", 1, line_no="1", amount=10.0),
         line_row("INV-1", "a.pdf", 2, line_no="9", amount=99.0)],
    )
    a = report.aggregates
    assert a["line_precision"] == Decimal("0.5")
    assert a["line_recall"] == Decimal("0.5")
    assert a["line_f1"] == Decimal("0.5")


def test_w_per_field_matched_line_accuracy(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", line_items=[
            {"line_no": "1", "description": "Freight", "amount": "10.00"}])},
        [invoice_row("INV-1", "a.pdf")],
        [line_row("INV-1", "a.pdf", 1, line_no="1", description="freight", amount=10.0)],
    )
    cr = case_result(report, "c1")
    assert cr.line_field_tallies["description"].outcome() == "correct"
    assert cr.line_field_tallies["amount"].outcome() == "correct"
    assert report.aggregates["matched_line_field_accuracy"] == Decimal("1")


# --- X/Y/Z: totals -------------------------------------------------------------

def test_x_totals_correct(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"total_amount": "30.00"},
                  line_items=[{"amount": "10.00"}, {"amount": "20.00"}])},
        [invoice_row("INV-1", "a.pdf", total_amount=30.0)],
        [line_row("INV-1", "a.pdf", 1, amount=10.0), line_row("INV-1", "a.pdf", 2, amount=20.0)],
    )
    assert case_result(report, "c1").totals["totals_flag_class"] == "correct_no_flag"


def test_y_totals_inconclusive_correctly_flagged(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf", expected_outcome="needs_review")],
        {"c1": gt("c1", invoice={"total_amount": "999.00"},
                  line_items=[{"amount": "10.00"}], expected_needs_review=True)},
        [invoice_row("INV-1", "a.pdf", total_amount=999.0, needs_review=True,
                     review_reason="totals inconclusive: sum(line_items)=10 total=999")],
        [line_row("INV-1", "a.pdf", 1, amount=10.0)],
    )
    assert case_result(report, "c1").totals["totals_flag_class"] == "correctly_flagged"


def test_z_false_totals_warning(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"total_amount": "10.00"}, line_items=[{"amount": "10.00"}])},
        [invoice_row("INV-1", "a.pdf", total_amount=10.0, needs_review=True,
                     review_reason="totals inconclusive: spurious")],
        [line_row("INV-1", "a.pdf", 1, amount=10.0)],
    )
    assert case_result(report, "c1").totals["totals_flag_class"] == "false_flag"


# --- AA-AD: needs_review confusion --------------------------------------------

def _review_case(tmp_path, expected_nr, actual_nr, reason=None):
    return _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"invoice_number": "INV-1"}, expected_needs_review=expected_nr)},
        [invoice_row("INV-1", "a.pdf", invoice_number="INV-1",
                     needs_review=actual_nr, review_reason=reason)],
        [],
    )


def test_aa_needs_review_true_positive(tmp_path):
    report = _review_case(tmp_path, True, True, reason="conflict in seller_name: ...")
    assert case_result(report, "c1").review_class == "TP"
    assert report.aggregates["review_tp"] == 1


def test_ab_needs_review_true_negative(tmp_path):
    report = _review_case(tmp_path, False, False)
    assert case_result(report, "c1").review_class == "TN"
    assert report.aggregates["review_tn"] == 1


def test_ac_needs_review_false_positive(tmp_path):
    report = _review_case(tmp_path, False, True, reason="totals inconclusive: ...")
    cr = case_result(report, "c1")
    assert cr.review_class == "FP"
    assert report.aggregates["false_review_rate"] == Decimal("1")


def test_ad_needs_review_false_negative(tmp_path):
    report = _review_case(tmp_path, True, False)
    cr = case_result(report, "c1")
    assert cr.review_class == "FN"
    assert report.aggregates["missed_problem_rate"] == Decimal("1")


# --- AE/AF: review category parsing -------------------------------------------

def test_ae_review_category_parsing():
    cats, unknown = parse_review_categories(
        "conflict in seller_name: 'A' vs 'B'; totals inconclusive: diff=5")
    assert cats == ["header_conflict", "totals_inconclusive"]
    assert unknown == []


def test_ae_multi_invoice_and_missing_fields():
    cats, _ = parse_review_categories(
        "possible multiple invoices in one PDF; missing required fields: currency")
    assert "invoice_number_conflict" in cats
    assert "missing_required_fields" in cats


def test_af_unknown_review_category_preserved():
    cats, unknown = parse_review_categories("some brand new reason we never mapped")
    assert cats == []
    assert unknown == ["some brand new reason we never mapped"]


def test_af_unknown_clause_surfaced_on_case(tmp_path):
    report = _review_case(tmp_path, True, True, reason="totally novel clause here")
    cr = case_result(report, "c1")
    assert cr.unknown_review_clauses == ["totally novel clause here"]


# --- AG/AH/AI/AJ: cost + runtime ----------------------------------------------

def test_ag_usage_aggregation_by_case(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf"), manifest_entry("c2", "b.pdf")],
        {"c1": gt("c1"), "c2": gt("c2")},
        [invoice_row("INV-1", "a.pdf"), invoice_row("INV-2", "b.pdf")],
        [],
        usage_records=[
            usage_record("a.pdf", cost_usd="0.0010"),
            usage_record("a.pdf", route="vision", cost_usd="0.0020"),
            usage_record("b.pdf", cost_usd="0.0005"),
        ],
    )
    assert case_result(report, "c1").cost["requests"] == 2
    assert case_result(report, "c1").cost["reported_cost"] == "0.0030"
    assert case_result(report, "c2").cost["reported_cost"] == "0.0005"
    assert report.aggregates["total_reported_cost"] == "0.0035"


def test_ah_unknown_cost_makes_aggregate_incomplete(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")},
        [invoice_row("INV-1", "a.pdf")], [],
        usage_records=[
            usage_record("a.pdf", cost_usd="0.0010"),
            usage_record("a.pdf", cost_usd=""),   # unknown
        ],
    )
    a = report.aggregates
    assert a["cost_incomplete"] is True
    assert a["unknown_cost_requests"] == 1
    assert case_result(report, "c1").cost["reported_cost"] == "0.0010"  # unknown not fabricated


def test_ai_repair_and_escalation_counts(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")},
        [invoice_row("INV-1", "a.pdf")], [],
        usage_records=[
            usage_record("a.pdf", attempt_type="primary", accepted="False"),
            usage_record("a.pdf", attempt_type="repair", accepted="False"),
            usage_record("a.pdf", attempt_type="escalation", accepted="True"),
        ],
    )
    cost = case_result(report, "c1").cost
    assert (cost["primary"], cost["repair"], cost["escalation"]) == (1, 1, 1)


def test_aj_runtime_from_metadata_vs_latency_labeling(tmp_path):
    # No run metadata -> runtime derives from usage latency, labeled provider_latency_only.
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")},
        [invoice_row("INV-1", "a.pdf")], [],
        usage_records=[usage_record("a.pdf", latency_ms="2500")],
    )
    cr = case_result(report, "c1")
    assert cr.runtime_basis == "provider_latency_only"
    assert cr.runtime_seconds == Decimal("2.500")


def test_aj_runtime_end_to_end_from_metadata(tmp_path):
    manifest = write_manifest(tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")})
    wb = write_workbook(tmp_path / "results.xlsx", [invoice_row("INV-1", "a.pdf")], [])
    meta = tmp_path / "run.json"
    meta.write_text(json.dumps({"run_id": "r", "files": [
        {"source_file": "a.pdf", "elapsed_seconds": 4.2}]}), encoding="utf-8")
    ds = load_manifest(manifest)
    report = score_benchmark(ds, wb, run_metadata_path=meta)
    cr = case_result(report, "c1")
    assert cr.runtime_basis == "end_to_end"
    assert cr.runtime_seconds == Decimal("4.2")


# --- AK: doc-type breakdown ---------------------------------------------------

def test_ak_document_type_breakdown(tmp_path):
    report = _score(
        tmp_path,
        [manifest_entry("c1", "a.pdf", "text_single_page"),
         manifest_entry("c2", "b.pdf", "vision_single_page")],
        {"c1": gt("c1"), "c2": gt("c2")},
        [invoice_row("INV-1", "a.pdf"), invoice_row("INV-2", "b.pdf")],
        [],
        usage_records=[usage_record("a.pdf", cost_usd="0.0010"),
                       usage_record("b.pdf", cost_usd="0.0030")],
    )
    by_type = {r["document_type"]: r for r in report.doc_type_table}
    assert by_type["text_single_page"]["avg_cost"] == "0.001000"
    assert by_type["vision_single_page"]["avg_cost"] == "0.003000"


# --- model table: request stats only, basis identified ------------------------

def test_model_table_request_stats_only_with_basis(tmp_path):
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")},
        [invoice_row("INV-1", "a.pdf")], [],
        usage_records=[
            usage_record("a.pdf", requested_model="req-x", actual_model="act-x", cost_usd="0.001"),
            usage_record("a.pdf", requested_model="req-y", actual_model="", cost_usd="0.002"),
        ],
    )
    rows = {r["model"]: r for r in report.model_table}
    assert rows["act-x"]["model_basis"] == "actual_model"
    assert rows["req-y"]["model_basis"] == "requested_model"
    # Model rows carry request/token/cost stats, never invoice-outcome fields.
    assert "needs_review" not in rows["act-x"]
    assert "passed" not in rows["act-x"]


# --- AO/AP: report outputs ----------------------------------------------------

def test_ao_benchmark_workbook_has_expected_sheets(tmp_path):
    report = _score(tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")},
                    [invoice_row("INV-1", "a.pdf")], [])
    path = write_report_workbook(report, tmp_path / "bench.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == REPORT_SHEETS
    assert len(wb.sheetnames) == 8


def test_ap_json_summary_deterministic(tmp_path):
    report = _score(
        tmp_path,
        [manifest_entry("c2", "b.pdf"), manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"invoice_number": "INV-1"}),
         "c2": gt("c2", invoice={"invoice_number": "INV-2"})},
        [invoice_row("INV-2", "b.pdf", invoice_number="INV-2"),
         invoice_row("INV-1", "a.pdf", invoice_number="INV-1")],
        [],
    )
    p1 = write_json_summary(report, tmp_path / "r1.json")
    p2 = write_json_summary(report, tmp_path / "r2.json")
    assert p1.read_text() == p2.read_text()             # byte-identical
    payload = json.loads(p1.read_text())
    assert [c["case_id"] for c in payload["cases"]] == ["c1", "c2"]  # sorted


# --- AR: normal extraction workbook stays three sheets ------------------------

def test_ar_normal_extraction_workbook_still_three_sheets(tmp_path):
    from invoice_extractor.excel_export import export_workbook
    from invoice_extractor.pipeline import InvoiceResult
    path = export_workbook([InvoiceResult(source_file="x.pdf")], tmp_path / "out.xlsx")
    wb = openpyxl.load_workbook(path)
    assert wb.sheetnames == ["Invoices", "LineItems", "NeedsReview"]


# --- AS: scoring makes zero network calls -------------------------------------

def test_as_scoring_offline_under_network_block(tmp_path):
    # The autouse block_network fixture raises on any socket use; a clean score
    # proves scoring never touches the network.
    report = _score(tmp_path, [manifest_entry("c1", "a.pdf")], {"c1": gt("c1")},
                    [invoice_row("INV-1", "a.pdf")], [])
    assert report.aggregates["num_cases"] == 1


# --- AT: anomalies recorded to Errors, not a crash ----------------------------

def test_at_missing_and_duplicate_and_extra_recorded(tmp_path):
    report = _score(
        tmp_path,
        [manifest_entry("c1", "a.pdf"), manifest_entry("c2", "dup.pdf"),
         manifest_entry("c3", "gone.pdf")],
        {"c1": gt("c1"), "c2": gt("c2"), "c3": gt("c3")},
        [invoice_row("INV-1", "a.pdf"),
         invoice_row("INV-2a", "dup.pdf"), invoice_row("INV-2b", "dup.pdf"),
         invoice_row("INV-X", "unexpected.pdf")],
        [],
    )
    cats = {(e["source_file"], e["category"].split(" ")[0]) for e in report.errors}
    assert ("dup.pdf", "duplicate_workbook_rows") in cats
    assert ("gone.pdf", "missing_workbook_result") in cats
    assert ("unexpected.pdf", "extra_workbook_source_file") in cats
    # Other cases still scored - one anomaly doesn't corrupt the whole report.
    assert case_result(report, "c1").invoice_status == "matched"


# --- AU: deterministic case ordering ------------------------------------------

def test_au_case_ordering_deterministic(tmp_path):
    report = _score(
        tmp_path,
        [manifest_entry("zebra", "z.pdf"), manifest_entry("alpha", "a.pdf")],
        {"zebra": gt("zebra"), "alpha": gt("alpha")},
        [invoice_row("INV-Z", "z.pdf"), invoice_row("INV-A", "a.pdf")],
        [],
    )
    assert [c.case_id for c in report.cases] == ["alpha", "zebra"]


# --- AQ: privacy --------------------------------------------------------------

def test_aq_no_sensitive_content_in_report_or_json(tmp_path):
    # Fake markers planted in usage rejection_category (a free-ish field) and in
    # a review reason; report outputs must not surface secrets/base64/etc.
    secret = "SECRET-OR-KEY-M6"
    b64 = "RkFLRUJBU0U2NC1NNg=="
    body = "FAKE-PROVIDER-BODY-M6"
    report = _score(
        tmp_path, [manifest_entry("c1", "a.pdf")],
        {"c1": gt("c1", invoice={"invoice_number": "INV-1"})},
        [invoice_row("INV-1", "a.pdf", invoice_number="INV-1", needs_review=True,
                     review_reason="totals inconclusive: safe reason only")],
        [],
        usage_records=[usage_record("a.pdf", rejection_category="rate_limited")],
    )
    xlsx = write_report_workbook(report, tmp_path / "bench.xlsx")
    js = write_json_summary(report, tmp_path / "bench.json")
    wb = openpyxl.load_workbook(xlsx)
    blob = js.read_text()
    for ws in wb.worksheets:
        blob += " ".join(str(c.value) for row in ws.iter_rows() for c in row
                         if c.value is not None)
    for forbidden in (secret, b64, body, "data:image", "base64"):
        assert forbidden not in blob

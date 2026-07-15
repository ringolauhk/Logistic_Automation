"""Benchmark scoring: header/line/totals/review/cost/runtime metrics (M6).

Consumes an already-produced extraction workbook + usage CSV (+ optional run
metadata) and compares against validated ground truth. Zero network calls,
zero provider calls. All numeric scoring uses Decimal.
"""

import csv
from dataclasses import dataclass, field
from decimal import Decimal
from pathlib import Path

import openpyxl

from invoice_extractor.benchmark.dataset import (
    NUMERIC_HEADER_GT_FIELDS,
    SCORED_HEADER_FIELDS,
    SCORED_LINE_FIELDS,
    BenchmarkDataset,
)
from invoice_extractor.benchmark.matching import (
    DEFAULT_AMOUNT_TOLERANCE,
    as_decimal,
    match_invoices,
    match_lines,
    norm_date,
    norm_identifier,
    norm_string,
)

# Header GT fields whose comparison is case-insensitive string equality.
_CASE_INSENSITIVE_HEADER = frozenset(
    {"seller_name", "seller_address", "buyer_name", "buyer_address", "currency"}
)
_IDENTIFIER_HEADER = frozenset({"invoice_number", "po_number", "reference"})
_NUMERIC_LINE_MONEY = frozenset({"unit_price", "amount"})  # quantity is exact
_REQUIRED_ACTUAL_FIELDS = ("invoice_date", "currency", "seller_name", "total_amount")

# Review-reason substring -> stable benchmark category. Order matters: the
# FIRST matching phrase wins for a given clause. Parsed conservatively from
# the production review strings (never changes them).
_REVIEW_PATTERNS = [
    ("invoice_number conflict", "invoice_number_conflict"),
    ("possible multiple invoices", "invoice_number_conflict"),
    ("conflict in invoice_number", "invoice_number_conflict"),
    ("conflict in", "header_conflict"),
    ("missing required fields", "missing_required_fields"),
    ("totals inconclusive", "totals_inconclusive"),
    ("partial extraction", "partial_extraction"),
    ("missing an amount", "suspicious_line"),
    ("hallucinated", "suspicious_line"),
    ("no line items", "suspicious_line"),
    ("run-wide", "budget_exhausted"),
    ("cost budget", "budget_exhausted"),
    ("attempt cap", "budget_exhausted"),
    ("unreadable pdf", "malformed_pdf"),
    ("no meaningful pages", "blank_document"),
    ("document is blank", "blank_document"),
    ("openrouter_text_models", "provider_failure"),
    ("openrouter_vision_models", "provider_failure"),
    ("openrouter_api_key", "provider_failure"),
    ("failed on all", "provider_failure"),
    ("http 4", "provider_failure"),
    ("http 5", "provider_failure"),
    ("rate_limited", "provider_failure"),
    ("truncated", "provider_failure"),
]


def parse_review_categories(review_reason: str | None) -> tuple[list[str], list[str]]:
    """Return (categories, unknown_clauses) parsed from a review reason.

    Splits on '; ' (the production reason joiner), maps each clause via
    _REVIEW_PATTERNS, and records any clause that matched nothing as an
    unknown clause label (its leading phrase only - never a value)."""
    if not review_reason:
        return [], []
    cats: list[str] = []
    unknown: list[str] = []
    for clause in review_reason.split("; "):
        low = clause.lower()
        hit = next((cat for pat, cat in _REVIEW_PATTERNS if pat in low), None)
        if hit:
            if hit not in cats:
                cats.append(hit)
        else:
            label = clause.split(":", 1)[0].strip()[:40]
            unknown.append(label)
    return cats, unknown


# --- Readers ------------------------------------------------------------------

def _norm_basename(source_file) -> str:
    if source_file is None:
        return ""
    return str(source_file).replace("\\", "/").rsplit("/", 1)[-1]


def read_workbook(path: Path) -> dict[str, list[dict]]:
    """Read the extraction workbook's Invoices + LineItems sheets into
    {basename(source_file) -> [actual_row, ...]}. actual_row carries header
    fields, needs_review, review_reason, provenance, and its line_items."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        invoices = _sheet_rows(wb["Invoices"])
        line_rows = _sheet_rows(wb["LineItems"]) if "LineItems" in wb.sheetnames else []
    finally:
        wb.close()

    lines_by_invoice: dict[str, list[dict]] = {}
    for lr in line_rows:
        lines_by_invoice.setdefault(lr.get("invoice_id"), []).append(lr)

    by_source: dict[str, list[dict]] = {}
    for row in invoices:
        src = _norm_basename(row.get("source_file"))
        actual = dict(row)
        actual["source_file"] = src
        actual["line_items"] = lines_by_invoice.get(row.get("invoice_id"), [])
        by_source.setdefault(src, []).append(actual)
    return by_source


def _sheet_rows(ws) -> list[dict]:
    rows = ws.iter_rows(values_only=True)
    try:
        header = list(next(rows))
    except StopIteration:
        return []
    out = []
    for values in rows:
        if values is None or all(v is None for v in values):
            continue
        out.append({header[i]: values[i] for i in range(len(header))})
    return out


def read_usage(path: Path | None) -> dict[str, list[dict]]:
    """Read the usage CSV into {basename(source_file) -> [record, ...]} in file
    (chronological) order. Missing path -> empty mapping."""
    if path is None:
        return {}
    path = Path(path)
    if not path.exists():
        return {}
    by_source: dict[str, list[dict]] = {}
    with open(path, newline="", encoding="utf-8") as f:
        for rec in csv.DictReader(f):
            src = _norm_basename(rec.get("source_file"))
            by_source.setdefault(src, []).append(rec)
    return by_source


def read_run_metadata(path: Path | None) -> dict[str, dict]:
    """Read optional run-metadata JSON into {basename -> {elapsed_seconds,...}}."""
    if path is None:
        return {}
    import json
    path = Path(path)
    if not path.exists():
        return {}
    raw = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, dict] = {}
    for entry in raw.get("files", []):
        out[_norm_basename(entry.get("source_file"))] = entry
    return out


# --- Per-field scoring --------------------------------------------------------

@dataclass
class FieldTally:
    correct: int = 0
    incorrect: int = 0
    missing: int = 0
    unexpected: int = 0

    @property
    def evaluated(self) -> int:
        return self.correct + self.incorrect + self.missing + self.unexpected

    def add(self, other: "FieldTally"):
        self.correct += other.correct
        self.incorrect += other.incorrect
        self.missing += other.missing
        self.unexpected += other.unexpected

    def outcome(self) -> str:
        for name in ("correct", "incorrect", "missing", "unexpected"):
            if getattr(self, name):
                return name
        return "none"


def _cmp_value(gt_name, expected, actual, tolerance: Decimal, *, numeric, ci, ident, date):
    """One-cell comparison -> a single-slot FieldTally."""
    t = FieldTally()
    exp_null = expected is None
    act_null = actual is None or actual == ""
    if exp_null and act_null:
        t.correct = 1
    elif exp_null and not act_null:
        t.unexpected = 1
    elif not exp_null and act_null:
        t.missing = 1
    else:
        if numeric:
            e, a = as_decimal(expected), as_decimal(actual)
            ok = e is not None and a is not None and abs(e - a) <= tolerance
        elif date:
            ok = norm_date(expected) == norm_date(actual)
        elif ident:
            ok = norm_identifier(expected) == norm_identifier(actual)
        elif ci:
            ok = norm_string(expected) == norm_string(actual)
        else:
            ok = norm_string(expected) == norm_string(actual)
        t.correct, t.incorrect = (1, 0) if ok else (0, 1)
    return t


def _tolerance_for(gt_name, case, default_numeric: Decimal) -> Decimal:
    return case.field_tolerances.get(gt_name, default_numeric)


# --- Case result --------------------------------------------------------------

@dataclass
class CaseResult:
    case_id: str
    source_file: str
    document_type: str
    expected_outcome: str
    invoice_status: str                     # matched|missing_result|duplicate_result
    actual_outcome: str = "none"            # extracted|needs_review|failed|none
    header_fields: dict = field(default_factory=dict)   # gt_name -> FieldTally
    not_extractable_fields: list = field(default_factory=list)
    ignored_fields: list = field(default_factory=list)
    exact_header_match: bool | None = None
    required_complete: bool | None = None
    line_counts: dict = field(default_factory=dict)     # expected/actual/matched/...
    line_field_tallies: dict = field(default_factory=dict)  # field -> FieldTally
    all_lines_correct: bool | None = None
    line_matches: list = field(default_factory=list)    # detail rows for LineMatches sheet
    totals: dict = field(default_factory=dict)
    expected_needs_review: bool | None = None
    actual_needs_review: bool | None = None
    review_class: str = "n/a"               # TP|TN|FP|FN|n/a
    actual_review_categories: list = field(default_factory=list)
    unknown_review_clauses: list = field(default_factory=list)
    cost: dict = field(default_factory=dict)
    runtime_seconds: Decimal | None = None
    runtime_basis: str = "unknown"          # end_to_end|provider_latency_only|unknown
    accepted_models: list = field(default_factory=list)
    routes: list = field(default_factory=list)
    passed: bool = True
    notes: str = ""


def _header_tally(case, actual_row) -> tuple[dict, list, list]:
    tallies: dict[str, FieldTally] = {}
    for gt_name, ext_name in SCORED_HEADER_FIELDS.items():
        if gt_name not in case.invoice:      # not annotated -> excluded
            continue
        if gt_name in case.ignored_fields:   # human-excluded -> excluded
            continue
        expected = case.invoice[gt_name]
        actual = actual_row.get(ext_name)
        tallies[gt_name] = _cmp_value(
            gt_name, expected, actual, _tolerance_for(gt_name, case, DEFAULT_AMOUNT_TOLERANCE),
            numeric=gt_name in NUMERIC_HEADER_GT_FIELDS,
            ci=gt_name in _CASE_INSENSITIVE_HEADER,
            ident=gt_name in _IDENTIFIER_HEADER,
            date=gt_name == "invoice_date",
        )
    return tallies, list(case.not_extractable_header_fields()), list(case.ignored_fields)


def _line_tally(case, actual_row, dataset):
    expected = case.line_items
    actual = actual_row.get("line_items", [])
    outcome = match_lines(
        expected, actual,
        amount_tolerance=DEFAULT_AMOUNT_TOLERANCE,
        fuzzy_enabled=dataset.fuzzy_enabled,
        fuzzy_threshold=dataset.fuzzy_threshold,
    )
    matched = outcome.matched
    counts = {
        "expected": len(expected), "actual": len(actual),
        "matched": len(matched), "missing": len(outcome.missing),
        "extra": len(outcome.extra), "ambiguous": len(outcome.ambiguous),
        "fuzzy": sum(1 for p in matched if p.method == "fuzzy"),
    }
    field_tallies: dict[str, FieldTally] = {f: FieldTally() for f in SCORED_LINE_FIELDS}
    detail_rows = []
    all_correct = True
    for pair in matched:
        exp = expected[pair.expected_index]
        act = actual[pair.actual_index]
        row_correct = True
        for gt_name in SCORED_LINE_FIELDS:
            if gt_name not in exp:
                continue
            numeric = gt_name in ("quantity", "unit_price", "amount")
            tol = (Decimal("0") if gt_name == "quantity"
                   else _tolerance_for(gt_name, case, DEFAULT_AMOUNT_TOLERANCE))
            t = _cmp_value(
                gt_name, exp[gt_name], act.get(gt_name), tol,
                numeric=numeric, ci=False,
                ident=gt_name in ("line_no", "item_code"),
                date=False,
            )
            field_tallies[gt_name].add(t)
            if t.outcome() != "correct":
                row_correct = False
        if not row_correct:
            all_correct = False
        detail_rows.append({
            "case_id": case.case_id, "expected_index": pair.expected_index,
            "actual_index": pair.actual_index, "method": pair.method,
            "confidence": str(pair.confidence), "row_correct": row_correct,
        })
    for pair in outcome.missing:
        detail_rows.append({"case_id": case.case_id, "expected_index": pair.expected_index,
                            "actual_index": None, "method": "missing",
                            "confidence": "0", "row_correct": False})
        all_correct = False
    for pair in outcome.extra:
        detail_rows.append({"case_id": case.case_id, "expected_index": None,
                            "actual_index": pair.actual_index, "method": "extra",
                            "confidence": "0", "row_correct": False})
        all_correct = False
    for pair in outcome.ambiguous:
        detail_rows.append({"case_id": case.case_id, "expected_index": pair.expected_index,
                            "actual_index": None, "method": "ambiguous",
                            "confidence": str(pair.confidence), "row_correct": False})
        all_correct = False
    all_lines_correct = (all_correct and counts["missing"] == 0
                         and counts["extra"] == 0 and counts["ambiguous"] == 0
                         and counts["expected"] == counts["matched"])
    ne_line = list(case.not_extractable_line_fields())
    return counts, field_tallies, detail_rows, all_lines_correct, ne_line


def _totals_result(case, actual_row) -> dict:
    """Compare totals + classify the extractor's totals-flag against GT."""
    exp_total = as_decimal(case.invoice.get("total_amount"))
    exp_tax = as_decimal(case.invoice.get("tax"))
    exp_line_sum = _line_sum(case.line_items)
    act_line_sum = _line_sum(actual_row.get("line_items", []))
    tol = _tolerance_for("total_amount", case, DEFAULT_AMOUNT_TOLERANCE)

    # Is GT internally consistent (a genuine "no totals problem")?
    gt_consistent = None
    if exp_total is not None and exp_line_sum is not None:
        gt_consistent = (abs(exp_line_sum - exp_total) <= tol or
                         (exp_tax is not None and
                          abs(exp_line_sum + exp_tax - exp_total) <= tol))
    reason = actual_row.get("review_reason") or ""
    flagged = "totals inconclusive" in reason.lower()
    if gt_consistent is None:
        flag_class = "undeterminable"
    elif gt_consistent and flagged:
        flag_class = "false_flag"
    elif gt_consistent and not flagged:
        flag_class = "correct_no_flag"
    elif not gt_consistent and flagged:
        flag_class = "correctly_flagged"
    else:
        flag_class = "missed_problem"
    return {
        "expected_total": str(exp_total) if exp_total is not None else None,
        "expected_line_sum": str(exp_line_sum) if exp_line_sum is not None else None,
        "actual_line_sum": str(act_line_sum) if act_line_sum is not None else None,
        "line_sum_reconciles": (exp_line_sum is not None and act_line_sum is not None
                                and abs(exp_line_sum - act_line_sum) <= tol),
        "totals_flag_class": flag_class,
    }


def _line_sum(items) -> Decimal | None:
    vals = [as_decimal(it.get("amount")) for it in items]
    vals = [v for v in vals if v is not None]
    return sum(vals, Decimal("0")) if vals else None


def _actual_outcome(actual_row) -> str:
    reason = (actual_row.get("review_reason") or "").lower()
    if "unreadable pdf" in reason or "no meaningful pages" in reason or \
       "failed on all" in reason or actual_row.get("extraction_method") == "failed":
        return "failed"
    if actual_row.get("needs_review"):
        return "needs_review"
    return "extracted"


def _cost_and_runtime(case, usage_by_source, meta_by_source):
    recs = usage_by_source.get(case.source_file, [])
    primary = sum(1 for r in recs if r.get("attempt_type") == "primary")
    repair = sum(1 for r in recs if r.get("attempt_type") == "repair")
    escalation = sum(1 for r in recs if r.get("attempt_type") == "escalation")
    unknown_cost = 0
    total_cost = Decimal("0")
    tokens = {"input": 0, "output": 0, "reasoning": 0, "total": 0}
    accepted_models, routes, page_ranges = [], [], []
    latency_ms = Decimal("0")
    have_latency = False
    for r in recs:
        cost = r.get("cost_usd")
        if cost is None or cost == "":
            unknown_cost += 1
        else:
            total_cost += Decimal(cost)
        for tk, col in (("input", "input_tokens"), ("output", "output_tokens"),
                        ("reasoning", "reasoning_tokens"), ("total", "total_tokens")):
            v = r.get(col)
            if v not in (None, ""):
                tokens[tk] += int(v)
        lat = r.get("latency_ms")
        if lat not in (None, ""):
            latency_ms += Decimal(lat)
            have_latency = True
        if str(r.get("accepted")).lower() == "true":
            model = r.get("actual_model") or r.get("requested_model")
            if model and model not in accepted_models:
                accepted_models.append(model)
        route = r.get("route")
        if route and route not in routes:
            routes.append(route)
        pr = r.get("page_range")
        if pr:
            page_ranges.append(pr)

    cost = {
        "requests": len(recs), "primary": primary, "repair": repair,
        "escalation": escalation, "input_tokens": tokens["input"],
        "output_tokens": tokens["output"], "reasoning_tokens": tokens["reasoning"],
        "total_tokens": tokens["total"], "reported_cost": str(total_cost),
        "unknown_cost_requests": unknown_cost, "page_ranges": page_ranges,
    }
    meta = meta_by_source.get(case.source_file)
    if meta and meta.get("elapsed_seconds") is not None:
        runtime = Decimal(str(meta["elapsed_seconds"]))
        basis = "end_to_end"
    elif have_latency:
        runtime = (latency_ms / Decimal("1000")).quantize(Decimal("0.001"))
        basis = "provider_latency_only"
    else:
        runtime = None
        basis = "unknown"
    return cost, accepted_models, routes, runtime, basis, unknown_cost


def score_case(case, invoice_match, dataset, usage_by_source, meta_by_source) -> CaseResult:
    cr = CaseResult(
        case_id=case.case_id, source_file=case.source_file,
        document_type=case.document_type, expected_outcome=case.expected_outcome,
        invoice_status=invoice_match.status,
        expected_needs_review=case.expected_needs_review,
    )
    cost, models, routes, runtime, basis, unknown_cost = _cost_and_runtime(
        case, usage_by_source, meta_by_source)
    cr.cost = cost
    cr.accepted_models = models
    cr.routes = routes
    cr.runtime_seconds = runtime
    cr.runtime_basis = basis

    if invoice_match.status != "matched":
        cr.passed = False
        cr.notes = invoice_match.status
        return cr

    row = invoice_match.actual_row
    cr.actual_outcome = _actual_outcome(row)
    cr.actual_needs_review = bool(row.get("needs_review"))
    cr.actual_review_categories, cr.unknown_review_clauses = parse_review_categories(
        row.get("review_reason"))

    cr.header_fields, cr.not_extractable_fields, cr.ignored_fields = _header_tally(case, row)
    evaluated = [t for t in cr.header_fields.values() if t.evaluated]
    cr.exact_header_match = bool(evaluated) and all(
        t.correct == t.evaluated for t in evaluated)
    cr.required_complete = all(row.get(f) not in (None, "") for f in _REQUIRED_ACTUAL_FIELDS)

    counts, line_tallies, detail, all_lines_correct, ne_line = _line_tally(case, row, dataset)
    cr.line_counts = counts
    cr.line_field_tallies = line_tallies
    cr.line_matches = detail
    cr.all_lines_correct = all_lines_correct
    cr.not_extractable_fields += ne_line
    cr.totals = _totals_result(case, row)

    # needs_review classification (matched cases only).
    exp, act = case.expected_needs_review, cr.actual_needs_review
    cr.review_class = {(True, True): "TP", (False, False): "TN",
                       (False, True): "FP", (True, False): "FN"}[(exp, act)]

    # Case pass/fail: an honest per-case verdict for the CaseResults sheet.
    cr.passed = (
        cr.actual_outcome != "failed"
        and (counts["missing"] == 0 and counts["extra"] == 0 and counts["ambiguous"] == 0)
        and cr.exact_header_match is not False
        and cr.review_class in ("TP", "TN")
    )
    return cr


# --- Aggregation + thresholds -------------------------------------------------

@dataclass
class BenchmarkReport:
    cases: list = field(default_factory=list)          # CaseResult
    aggregates: dict = field(default_factory=dict)
    model_table: list = field(default_factory=list)
    doc_type_table: list = field(default_factory=list)
    errors: list = field(default_factory=list)         # {case_id, source_file, category}
    threshold_results: list = field(default_factory=list)
    fuzzy_enabled: bool = False

    @property
    def thresholds_passed(self) -> bool:
        return all(t["passed"] for t in self.threshold_results)


def _ratio(num, den) -> Decimal | None:
    if den == 0:
        return None
    return (Decimal(num) / Decimal(den)).quantize(Decimal("0.0001"))


def _f1(precision, recall) -> Decimal | None:
    if precision is None or recall is None or (precision + recall) == 0:
        return None
    return (2 * precision * recall / (precision + recall)).quantize(Decimal("0.0001"))


def _median(values):
    if not values:
        return None
    ordered = sorted(values)
    n = len(ordered)
    mid = n // 2
    if n % 2:
        return ordered[mid]
    return (ordered[mid - 1] + ordered[mid]) / 2


def _aggregate(cases) -> tuple[dict, list, list]:
    matched = [c for c in cases if c.invoice_status == "matched"]

    # Header field aggregation.
    header_fields: dict[str, FieldTally] = {}
    for c in matched:
        for name, tally in c.header_fields.items():
            header_fields.setdefault(name, FieldTally()).add(tally)
    h_correct = sum(t.correct for t in header_fields.values())
    h_eval = sum(t.evaluated for t in header_fields.values())
    field_accuracies = {n: _ratio(t.correct, t.evaluated)
                        for n, t in header_fields.items() if t.evaluated}
    header_macro = (
        (sum(field_accuracies.values(), Decimal("0")) / len(field_accuracies))
        .quantize(Decimal("0.0001")) if field_accuracies else None)

    # Line field aggregation.
    line_fields: dict[str, FieldTally] = {}
    for c in matched:
        for name, tally in c.line_field_tallies.items():
            line_fields.setdefault(name, FieldTally()).add(tally)
    l_correct = sum(t.correct for t in line_fields.values())
    l_eval = sum(t.evaluated for t in line_fields.values())
    line_field_acc = {n: _ratio(t.correct, t.evaluated)
                     for n, t in line_fields.items() if t.evaluated}

    exp_lines = sum(c.line_counts.get("expected", 0) for c in matched)
    act_lines = sum(c.line_counts.get("actual", 0) for c in matched)
    matched_lines = sum(c.line_counts.get("matched", 0) for c in matched)
    fuzzy_lines = sum(c.line_counts.get("fuzzy", 0) for c in matched)
    line_prec = _ratio(matched_lines, act_lines)
    line_rec = _ratio(matched_lines, exp_lines)

    numeric_fields = ("quantity", "unit_price", "amount")
    num_correct = sum(line_fields[f].correct for f in numeric_fields if f in line_fields)
    num_eval = sum(line_fields[f].evaluated for f in numeric_fields if f in line_fields)

    # needs_review confusion matrix.
    conf = {k: sum(1 for c in matched if c.review_class == k)
            for k in ("TP", "TN", "FP", "FN")}
    rev_prec = _ratio(conf["TP"], conf["TP"] + conf["FP"])
    rev_rec = _ratio(conf["TP"], conf["TP"] + conf["FN"])

    # Cost / runtime.
    total_cost = sum((Decimal(c.cost.get("reported_cost", "0")) for c in cases), Decimal("0"))
    unknown_cost = sum(c.cost.get("unknown_cost_requests", 0) for c in cases)
    runtimes = [c.runtime_seconds for c in cases if c.runtime_seconds is not None]
    bases = sorted({c.runtime_basis for c in cases if c.runtime_seconds is not None})

    outcomes = {k: sum(1 for c in matched if c.actual_outcome == k)
                for k in ("extracted", "needs_review", "failed")}

    # Weakest fields (header + line), ascending accuracy, then by evaluated desc.
    weakest = sorted(
        ([("header", n, field_accuracies[n], header_fields[n].evaluated)
          for n in field_accuracies]
         + [("line", n, line_field_acc[n], line_fields[n].evaluated)
            for n in line_field_acc]),
        key=lambda t: (t[2], -t[3], t[1]),
    )[:5]

    not_extractable = sorted({f for c in cases for f in c.not_extractable_fields})

    aggregates: dict = {
        "num_cases": len(cases),
        "matched_cases": len(matched),
        "outcome_extracted": outcomes["extracted"],
        "outcome_needs_review": outcomes["needs_review"],
        "outcome_failed": outcomes["failed"],
        "header_micro_accuracy": _ratio(h_correct, h_eval),
        "header_macro_accuracy": header_macro,
        "header_cells_evaluated": h_eval,
        "exact_header_match_rate": _ratio(
            sum(1 for c in matched if c.exact_header_match), len(matched)),
        "required_field_completeness": _ratio(
            sum(1 for c in matched if c.required_complete), len(matched)),
        "line_precision": line_prec,
        "line_recall": line_rec,
        "line_f1": _f1(line_prec, line_rec),
        "matched_line_field_accuracy": _ratio(l_correct, l_eval),
        "numeric_field_accuracy": _ratio(num_correct, num_eval),
        "description_accuracy": line_field_acc.get("description"),
        "item_code_accuracy": line_field_acc.get("item_code"),
        "invoice_all_lines_correct_rate": _ratio(
            sum(1 for c in matched if c.all_lines_correct), len(matched)),
        "review_tp": conf["TP"], "review_tn": conf["TN"],
        "review_fp": conf["FP"], "review_fn": conf["FN"],
        "review_precision": rev_prec, "review_recall": rev_rec,
        "review_f1": _f1(rev_prec, rev_rec),
        "false_review_rate": _ratio(conf["FP"], conf["FP"] + conf["TN"]),
        "missed_problem_rate": _ratio(conf["FN"], conf["FN"] + conf["TP"]),
        "total_reported_cost": str(total_cost),
        "cost_incomplete": unknown_cost > 0,
        "unknown_cost_requests": unknown_cost,
        "average_runtime_seconds": (
            str((sum(runtimes) / len(runtimes)).quantize(Decimal("0.001")))
            if runtimes else None),
        "runtime_bases": bases,
        "fuzzy_line_matches": fuzzy_lines,
        "not_extractable_fields": not_extractable,
        "not_extractable_field_count": len(not_extractable),
        "weakest_fields": [
            {"scope": scope, "field": name, "accuracy": str(acc), "evaluated": ev}
            for scope, name, acc, ev in weakest],
        "header_field_accuracy": {n: str(a) for n, a in sorted(field_accuracies.items())},
        "line_field_accuracy": {n: str(a) for n, a in sorted(line_field_acc.items())},
    }
    doc_table = _doc_type_table(cases)
    return aggregates, doc_table


def _doc_type_table(cases) -> list:
    by_type: dict[str, list] = {}
    for c in cases:
        by_type.setdefault(c.document_type, []).append(c)
    rows = []
    for dtype in sorted(by_type):
        group = by_type[dtype]
        costs = [Decimal(c.cost.get("reported_cost", "0")) for c in group]
        runtimes = [c.runtime_seconds for c in group if c.runtime_seconds is not None]
        matched = [c for c in group if c.invoice_status == "matched"]
        lines = sum(c.line_counts.get("actual", 0) for c in matched)
        cost_sum = sum(costs, Decimal("0"))
        runtime_sum = sum(runtimes, Decimal("0")) if runtimes else None
        rows.append({
            "document_type": dtype, "cases": len(group),
            "avg_cost": str((cost_sum / len(group)).quantize(Decimal("0.000001"))) if group else None,
            "median_cost": str(_median(costs)) if costs else None,
            "avg_runtime": (str((runtime_sum / len(runtimes)).quantize(Decimal("0.001")))
                            if runtimes else None),
            "median_runtime": str(_median(runtimes)) if runtimes else None,
            "p95_runtime": (str(_percentile(runtimes, 95)) if len(runtimes) >= 20 else None),
            "extracted": sum(1 for c in matched if c.actual_outcome == "extracted"),
            "needs_review": sum(1 for c in matched if c.actual_outcome == "needs_review"),
            "failed": sum(1 for c in matched if c.actual_outcome == "failed"),
            "lines_per_dollar": (str((Decimal(lines) / cost_sum).quantize(Decimal("0.01")))
                                 if cost_sum > 0 else None),
            "lines_per_minute": (
                str((Decimal(lines) / (runtime_sum / Decimal("60"))).quantize(Decimal("0.01")))
                if runtime_sum and runtime_sum > 0 else None),
        })
    return rows


def _percentile(values, pct):
    ordered = sorted(values)
    k = (len(ordered) - 1) * (Decimal(pct) / Decimal("100"))
    lo = int(k)
    hi = min(lo + 1, len(ordered) - 1)
    frac = k - lo
    return (ordered[lo] + (ordered[hi] - ordered[lo]) * frac).quantize(Decimal("0.001"))


def _build_model_table(usage_by_source) -> list:
    from collections import defaultdict
    agg = defaultdict(lambda: {"requests": 0, "primary": 0, "repair": 0,
                               "escalation": 0, "input_tokens": 0, "output_tokens": 0,
                               "total_tokens": 0, "cost": Decimal("0"),
                               "unknown_cost": 0, "basis": set()})
    for recs in usage_by_source.values():
        for r in recs:
            actual = r.get("actual_model")
            model = actual if actual not in (None, "") else r.get("requested_model")
            basis = "actual_model" if actual not in (None, "") else "requested_model"
            if not model:
                continue
            a = agg[model]
            a["requests"] += 1
            at = r.get("attempt_type")
            if at in ("primary", "repair", "escalation"):
                a[at] += 1
            for tk, col in (("input_tokens", "input_tokens"),
                            ("output_tokens", "output_tokens"),
                            ("total_tokens", "total_tokens")):
                v = r.get(col)
                if v not in (None, ""):
                    a[tk] += int(v)
            cost = r.get("cost_usd")
            if cost in (None, ""):
                a["unknown_cost"] += 1
            else:
                a["cost"] += Decimal(cost)
            a["basis"].add(basis)
    rows = []
    for model in sorted(agg):
        a = agg[model]
        rows.append({
            "model": model, "model_basis": "+".join(sorted(a["basis"])),
            "requests": a["requests"], "primary": a["primary"],
            "repair": a["repair"], "escalation": a["escalation"],
            "input_tokens": a["input_tokens"], "output_tokens": a["output_tokens"],
            "total_tokens": a["total_tokens"], "reported_cost": str(a["cost"]),
            "unknown_cost_requests": a["unknown_cost"],
        })
    return rows


# Threshold key -> (aggregate key, comparison). "min" => value must be >=;
# "max" => value must be <=.
_THRESHOLD_SPECS = {
    "minimum_header_micro_accuracy": ("header_micro_accuracy", "min"),
    "minimum_line_recall": ("line_recall", "min"),
    "minimum_line_precision": ("line_precision", "min"),
    "minimum_numeric_accuracy": ("numeric_field_accuracy", "min"),
    "maximum_false_review_rate": ("false_review_rate", "max"),
    "maximum_missed_problem_rate": ("missed_problem_rate", "max"),
    "maximum_average_cost_per_invoice": ("_avg_cost_per_invoice", "max"),
    "maximum_average_runtime_seconds": ("average_runtime_seconds", "max"),
}


def evaluate_thresholds(thresholds: dict, aggregates: dict) -> list:
    if not thresholds:
        return []
    results = []
    num_cases = aggregates.get("num_cases", 0) or 1
    avg_cost = (Decimal(aggregates["total_reported_cost"]) / num_cases)
    for key, target in sorted(thresholds.items()):
        if key.startswith("__"):   # documentation keys (e.g. "__doc__")
            continue
        if key not in _THRESHOLD_SPECS:
            results.append({"threshold": key, "target": str(target),
                            "actual": None, "passed": False,
                            "note": "unknown threshold key"})
            continue
        agg_key, mode = _THRESHOLD_SPECS[key]
        if agg_key == "_avg_cost_per_invoice":
            actual = avg_cost
        else:
            raw = aggregates.get(agg_key)
            actual = Decimal(raw) if raw is not None else None
        target_dec = Decimal(str(target))
        if actual is None:
            passed, note = False, "metric unavailable (no evaluated data)"
        elif mode == "min":
            passed, note = actual >= target_dec, ""
        else:
            passed, note = actual <= target_dec, ""
        results.append({"threshold": key, "target": str(target_dec),
                        "actual": str(actual) if actual is not None else None,
                        "passed": passed, "note": note})
    return results


def score_benchmark(dataset: BenchmarkDataset, workbook_path, usage_path=None,
                    run_metadata_path=None) -> BenchmarkReport:
    """Full offline scoring pass. Never makes a network/provider call."""
    actual_by_source = read_workbook(Path(workbook_path))
    usage_by_source = read_usage(usage_path)
    meta_by_source = read_run_metadata(run_metadata_path)

    invoice_matches, extra_sources = match_invoices(dataset.cases, actual_by_source)
    match_by_case = {m.case_id: m for m in invoice_matches}

    cases: list[CaseResult] = []
    errors: list[dict] = []
    for case in dataset.cases:
        m = match_by_case[case.case_id]
        if m.status == "duplicate_result":
            errors.append({"case_id": case.case_id, "source_file": case.source_file,
                           "category": f"duplicate_workbook_rows ({m.duplicate_count})"})
        elif m.status == "missing_result":
            errors.append({"case_id": case.case_id, "source_file": case.source_file,
                           "category": "missing_workbook_result"})
        cases.append(score_case(case, m, dataset, usage_by_source, meta_by_source))
    for src in extra_sources:
        errors.append({"case_id": None, "source_file": src,
                       "category": "extra_workbook_source_file"})

    aggregates, doc_table = _aggregate(cases)
    model_table = _build_model_table(usage_by_source)
    threshold_results = evaluate_thresholds(dataset.thresholds, aggregates)
    aggregates["fuzzy_enabled"] = dataset.fuzzy_enabled
    return BenchmarkReport(
        cases=cases, aggregates=aggregates, model_table=model_table,
        doc_type_table=doc_table, errors=errors,
        threshold_results=threshold_results, fuzzy_enabled=dataset.fuzzy_enabled,
    )

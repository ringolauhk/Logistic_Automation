"""Packing-list workbook generation (Build 7).

Generates ONE validated .xlsx per destination from the CURRENT Build 6
packing artifact (consumed verbatim - groups, carton numbers, consolidation
and delivery invoice numbers are never recalculated), plus a ZIP when
multiple destinations exist. Layout mirrors the legacy IMAGINEX packing
list (verified against the local sample read-only): company header, boxed
title, delivery invoice number/date, TN# remarks, Form Of Delivery,
CTN./Description/SKU/PLU/IMX columns, per-carton subtotals, final totals.

Customer-specific columns (Customer Style / Color Code / Color Desc) stay
BLANK unless a configuration mapping names a normalized product attribute -
Analysis Code meanings are not confirmed and are never hard-coded.

Safety: five fixed sheets; identifier cells written as text (leading zeros
preserved); every workbook is reopened and validated before being recorded;
outputs live only under <job>/output/ with sanitized deterministic names;
metadata (output/result.json) stores relative paths + SHA-256 and the
packing checksum for staleness; no API call, no printing, no email.

The Needs Review and Source Documents audit sheets additionally READ the
product-lookup and extraction artifacts (never written) as sanctioned
audit context.
"""

import hashlib
import json
import os
import re
import zipfile
from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from apps.web.job_manager import JobError, utc_now
from apps.web.transfer import jobs
from apps.web.transfer import packing as pk
from apps.web.transfer.models import (
    JOB_PACKING_PREPARATION_COMPLETE,
    JOB_PACKING_PREPARATION_WITH_ISSUES,
    JOB_WORKBOOK_GENERATION_COMPLETE,
    JOB_WORKBOOK_GENERATION_FAILED,
    JOB_WORKBOOK_GENERATION_IN_PROGRESS,
    JOB_WORKBOOK_GENERATION_WITH_ISSUES,
)

OUTPUT_SCHEMA_VERSION = 1
OUTPUT_DIR = "output"
OUTPUT_META_NAME = "result.json"

SHEET_PACKING_LIST = "Packing List"
SHEET_DETAIL = "Detail"
SHEET_CARTON_MAPPING = "Carton Mapping"
SHEET_NEEDS_REVIEW = "Needs Review"
SHEET_SOURCE_DOCUMENTS = "Source Documents"
REQUIRED_SHEETS = (SHEET_PACKING_LIST, SHEET_DETAIL, SHEET_CARTON_MAPPING,
                   SHEET_NEEDS_REVIEW, SHEET_SOURCE_DOCUMENTS)

DEFAULT_COMPANY_NAME = "IMAGINEX BG LIMITED"
DEFAULT_DOCUMENT_TITLE = "Packing List"
DEFAULT_FORM_OF_DELIVERY = "Warehouse Shipment"
DEFAULT_RETENTION_HOURS = 24.0

# --- validation issue codes -------------------------------------------------------

WORKBOOK_FILE_MISSING = "WORKBOOK_FILE_MISSING"
WORKBOOK_FILE_EMPTY = "WORKBOOK_FILE_EMPTY"
WORKBOOK_OPEN_FAILED = "WORKBOOK_OPEN_FAILED"
WORKBOOK_SHEET_MISSING = "WORKBOOK_SHEET_MISSING"
WORKBOOK_SHEET_DUPLICATE = "WORKBOOK_SHEET_DUPLICATE"
WORKBOOK_METADATA_MISMATCH = "WORKBOOK_METADATA_MISMATCH"
WORKBOOK_CARTON_COUNT_MISMATCH = "WORKBOOK_CARTON_COUNT_MISMATCH"
WORKBOOK_LINE_COUNT_MISMATCH = "WORKBOOK_LINE_COUNT_MISMATCH"
WORKBOOK_TOTAL_MISMATCH = "WORKBOOK_TOTAL_MISMATCH"
WORKBOOK_IDENTIFIER_FORMAT_ERROR = "WORKBOOK_IDENTIFIER_FORMAT_ERROR"
WORKBOOK_EXTERNAL_LINK_FOUND = "WORKBOOK_EXTERNAL_LINK_FOUND"
WORKBOOK_VALIDATION_FAILED = "WORKBOOK_VALIDATION_FAILED"

SEV_BLOCKING = "blocking"
SEV_WARNING = "warning"

GENERABLE_STATUSES = (JOB_PACKING_PREPARATION_COMPLETE,
                      JOB_PACKING_PREPARATION_WITH_ISSUES,
                      JOB_WORKBOOK_GENERATION_IN_PROGRESS,
                      JOB_WORKBOOK_GENERATION_COMPLETE,
                      JOB_WORKBOOK_GENERATION_WITH_ISSUES,
                      JOB_WORKBOOK_GENERATION_FAILED)

# Normalized attribute names a customer-mapping field may reference.
ALLOWED_MAPPING_FIELDS = tuple(
    [f"analysis_code_{i:02d}" for i in range(1, 16)]
    + [f"composition_{i:02d}" for i in range(1, 5)]
    + ["item_code", "plu", "ean", "color_code", "color_desc", "size_code",
       "brand", "brand_name", "season", "subcat", "supplier_item_code"])


# --- configuration ----------------------------------------------------------------

def _env(name: str) -> str:
    return (os.environ.get(name) or "").strip()


@dataclass(frozen=True)
class PackingWorkbookConfig:
    company_name: str = DEFAULT_COMPANY_NAME
    document_title: str = DEFAULT_DOCUMENT_TITLE
    form_of_delivery: str = DEFAULT_FORM_OF_DELIVERY
    retention_hours: float = DEFAULT_RETENTION_HOURS
    create_zip_for_multiple: bool = True
    customer_style_field: str = ""
    customer_color_code_field: str = ""
    customer_color_desc_field: str = ""

    def summary(self) -> dict:
        return {"company_name": self.company_name,
                "document_title": self.document_title,
                "form_of_delivery": self.form_of_delivery,
                "retention_hours": self.retention_hours,
                "create_zip_for_multiple": self.create_zip_for_multiple,
                "customer_style_field": self.customer_style_field or None,
                "customer_color_code_field":
                    self.customer_color_code_field or None,
                "customer_color_desc_field":
                    self.customer_color_desc_field or None}


def workbook_config_problems() -> list[str]:
    problems = []
    raw = _env("PACKING_OUTPUT_RETENTION_HOURS")
    if raw:
        try:
            if float(raw) <= 0:
                problems.append("PACKING_OUTPUT_RETENTION_HOURS must be "
                                "positive.")
        except ValueError:
            problems.append("PACKING_OUTPUT_RETENTION_HOURS must be a "
                            "number.")
    raw = _env("PACKING_CREATE_ZIP_FOR_MULTIPLE")
    if raw and raw.lower() not in ("1", "0", "true", "false", "yes", "no",
                                   "on", "off"):
        problems.append("PACKING_CREATE_ZIP_FOR_MULTIPLE must be a boolean.")
    for name in ("PACKING_CUSTOMER_STYLE_FIELD",
                 "PACKING_CUSTOMER_COLOR_CODE_FIELD",
                 "PACKING_CUSTOMER_COLOR_DESC_FIELD"):
        raw = _env(name)
        if raw and raw.lower() not in ALLOWED_MAPPING_FIELDS:
            problems.append(f"{name} must be one of the normalized product "
                            "attribute names (e.g. analysis_code_01).")
    title = _env("PACKING_DOCUMENT_TITLE")
    if title and ("/" in title or "\\" in title):
        problems.append("PACKING_DOCUMENT_TITLE must not contain path "
                        "separators.")
    return problems


def load_workbook_config() -> PackingWorkbookConfig:
    problems = workbook_config_problems()
    if problems:
        raise JobError("Workbook configuration invalid: "
                       + " ".join(problems))
    raw_zip = _env("PACKING_CREATE_ZIP_FOR_MULTIPLE")
    return PackingWorkbookConfig(
        company_name=_env("PACKING_COMPANY_NAME") or DEFAULT_COMPANY_NAME,
        document_title=_env("PACKING_DOCUMENT_TITLE")
        or DEFAULT_DOCUMENT_TITLE,
        form_of_delivery=_env("PACKING_FORM_OF_DELIVERY")
        or DEFAULT_FORM_OF_DELIVERY,
        retention_hours=float(_env("PACKING_OUTPUT_RETENTION_HOURS")
                              or DEFAULT_RETENTION_HOURS),
        create_zip_for_multiple=(raw_zip.lower()
                                 in ("1", "true", "yes", "on")
                                 if raw_zip else True),
        customer_style_field=_env("PACKING_CUSTOMER_STYLE_FIELD").lower(),
        customer_color_code_field=_env(
            "PACKING_CUSTOMER_COLOR_CODE_FIELD").lower(),
        customer_color_desc_field=_env(
            "PACKING_CUSTOMER_COLOR_DESC_FIELD").lower(),
    )


# --- paths, filenames, checksums --------------------------------------------------

_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._-]+")


def sanitize_component(value: str) -> str:
    """Filesystem-safe filename component; business values stay intact in
    cells/metadata."""
    cleaned = _SAFE_NAME_RE.sub("-", (value or "").strip()).strip("-.")
    return cleaned[:80] or "UNKNOWN"


def output_dir(job_id: str) -> Path:
    return jobs.transfer_job_dir_for(job_id) / OUTPUT_DIR


def output_meta_path(job_id: str) -> Path:
    return output_dir(job_id) / OUTPUT_META_NAME


def workbook_filename(destination: str, invoice_no: str) -> str:
    return (f"Packing_List_{sanitize_component(destination)}_"
            f"{sanitize_component(invoice_no)}.xlsx")


def zip_filename(job_id: str) -> str:
    return f"Packing_Lists_{sanitize_component(job_id)}.zip"


def packing_checksum(job_id: str) -> str | None:
    try:
        return hashlib.sha256(
            pk.result_path(job_id).read_bytes()).hexdigest()
    except (JobError, OSError):
        return None


def _sha256(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


# --- input boundary ---------------------------------------------------------------

def load_generation_inputs(job_id: str):
    """(job, packing_dict) or JobError. Groups are consumed verbatim."""
    job = jobs.load_transfer_job(job_id)
    if job is None:
        raise JobError("Unknown transfer job id.")
    if job.status not in GENERABLE_STATUSES:
        raise JobError(f"Job in state {job.status} cannot generate "
                       "workbooks.")
    prepared = pk.load_preparation(job_id)
    if prepared is None:
        raise JobError("No packing preparation exists; prepare packing "
                       "groups first.")
    if prepared.get("stale"):
        raise JobError("The packing preparation is stale (upstream data "
                       "changed); rerun packing preparation first.")
    if prepared.get("status") not in ("complete", "complete_with_issues"):
        raise JobError("The packing preparation did not complete.")
    blocking = (prepared.get("summary") or {}).get("blocking_issues", 0)
    if blocking:
        raise JobError(f"{blocking} blocking packing issue(s) remain; "
                       "resolve them before generating workbooks.")
    groups = prepared.get("destinations") or []
    if not groups:
        raise JobError("No destination groups exist.")
    for group in groups:
        if not group.get("destination_code"):
            raise JobError("A destination group is missing its code.")
        if not group.get("delivery_invoice_number"):
            raise JobError(f"Destination {group['destination_code']} has no "
                           "delivery invoice number.")
        if not group.get("carton_mappings"):
            raise JobError(f"Destination {group['destination_code']} has no "
                           "cartons.")
        if not group.get("prepared_lines"):
            raise JobError(f"Destination {group['destination_code']} has no "
                           "prepared lines.")
    return job, prepared


# --- styling helpers --------------------------------------------------------------

_BOLD = Font(bold=True)
_TITLE_FONT = Font(bold=True, size=16)
_HEADER_FONT = Font(bold=True, size=9)
_CELL_FONT = Font(size=9)
_THIN = Side(style="thin")
_HEADER_BORDER = Border(bottom=Side(style="medium"))
_BOX_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_WRAP = Alignment(wrap_text=True, vertical="center")

TEXT_FORMAT = "@"
QTY_FORMAT = "0"
PRICE_FORMAT = "0.00"


def _text_cell(ws, row, col, value, *, font=_CELL_FONT, bold=False):
    cell = ws.cell(row=row, column=col,
                   value="" if value is None else str(value))
    cell.number_format = TEXT_FORMAT
    cell.font = _BOLD if bold else font
    return cell


def _num_cell(ws, row, col, value, *, fmt=QTY_FORMAT, bold=False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.number_format = fmt
    if bold:
        cell.font = _BOLD
    return cell


def _header_row(ws, row, headers, *, widths=None):
    for index, title in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=index, value=title)
        cell.font = _HEADER_FONT
        cell.border = _HEADER_BORDER
        cell.alignment = _WRAP
    if widths:
        for index, width in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(index)].width = width
    ws.freeze_panes = ws.cell(row=row + 1, column=1)
    ws.auto_filter.ref = (f"A{row}:{get_column_letter(len(headers))}{row}")
    return row + 1


def _mapped(product: dict, field_name: str) -> str:
    return str(product.get(field_name) or "") if field_name else ""


# --- sheet builders ---------------------------------------------------------------

def _build_packing_list_sheet(ws, group, config, generated_at):
    """Business-facing printable sheet mirroring the legacy IMAGINEX
    layout. Values are written as fixed calculated cells - no formulas."""
    ws.sheet_view.showGridLines = False
    cell = ws.cell(row=1, column=1, value=config.company_name)
    cell.font = _TITLE_FONT
    title = ws.cell(row=2, column=7, value=config.document_title)
    title.font = Font(bold=True, size=12)
    title.border = _BOX_BORDER
    title.alignment = Alignment(horizontal="center")

    labels = [
        ("To :", ""),                       # customer unknown - never invented
        ("Delivery Invoice No. :", group["delivery_invoice_number"]),
        ("Invoice Date :", _invoice_date_display(group)),
        ("Destination :", f"{group['destination_code']} "
                          f"{group.get('destination_name') or ''}".strip()),
        ("Remarks :", "TN# " + ", ".join(group.get("source_delivery_notes")
                                         or []) if
         group.get("source_delivery_notes") else "Remarks :"),
        ("Deliver to :", ""),
        ("Contact Person:", ""),
        ("Tel:", ""),
        ("Form Of Delivery :", config.form_of_delivery),
        ("Generated :", generated_at),
    ]
    row = 4
    for label, value in labels:
        _text_cell(ws, row, 1, label, bold=True)
        _text_cell(ws, row, 3, value)
        row += 1

    headers = ["CTN. No.", "Description", "SKU Number / EAN", "PLU",
               "IMX Item Code", "IMX Color Code", "Color Description",
               "Size", "Customer Style", "Customer Color Code",
               "Customer Color Description", "Qty"]
    widths = [9, 36, 17, 22, 18, 13, 16, 8, 14, 14, 18, 7]
    table_header_row = row + 1
    data_row = _header_row(ws, table_header_row, headers, widths=widths)
    ws.freeze_panes = None                 # printable sheet: no freeze

    total_units = 0
    for mapping in group["carton_mappings"]:
        carton_no = mapping["generated_carton_number"]
        carton_lines = [l for l in group["prepared_lines"]
                        if l["generated_carton_number"] == carton_no]
        subtotal = 0
        for line in carton_lines:
            product = line.get("product") or {}
            _text_cell(ws, data_row, 1, carton_no)
            _text_cell(ws, data_row, 2, product.get("item_desc")
                       or (line.get("source") or {}).get("description"))
            _text_cell(ws, data_row, 3, product.get("ean"))
            _text_cell(ws, data_row, 4, product.get("plu"))
            _text_cell(ws, data_row, 5, product.get("item_code"))
            _text_cell(ws, data_row, 6, product.get("color_code"))
            _text_cell(ws, data_row, 7, product.get("color_desc"))
            _text_cell(ws, data_row, 8, product.get("size_code"))
            _text_cell(ws, data_row, 9,
                       _mapped(product, config.customer_style_field))
            _text_cell(ws, data_row, 10,
                       _mapped(product, config.customer_color_code_field))
            _text_cell(ws, data_row, 11,
                       _mapped(product, config.customer_color_desc_field))
            _num_cell(ws, data_row, 12, line["quantity"])
            subtotal += line["quantity"]
            data_row += 1
        _text_cell(ws, data_row, 1, f"{carton_no} Total", bold=True)
        _num_cell(ws, data_row, 12, subtotal, bold=True)
        ws.cell(row=data_row, column=12).border = Border(top=_THIN)
        total_units += subtotal
        data_row += 1

    data_row += 1
    _text_cell(ws, data_row, 1, "Total Cartons", bold=True)
    _num_cell(ws, data_row, 2, group["generated_carton_count"], bold=True)
    _text_cell(ws, data_row, 5, "Total Prepared Lines", bold=True)
    _num_cell(ws, data_row, 7, group["prepared_line_count"], bold=True)
    _text_cell(ws, data_row, 10, "Total Units", bold=True)
    total_cell = _num_cell(ws, data_row, 12, total_units, bold=True)
    total_cell.border = Border(top=Side(style="double"))

    # print layout: landscape, one page wide, repeat the table header
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = f"{table_header_row}:{table_header_row}"
    ws.print_area = f"A1:L{data_row}"
    ws.oddHeader.right.text = (f"{group['delivery_invoice_number']} - "
                               f"{group['destination_code']}")
    ws.oddFooter.center.text = "Page &P of &N"
    return total_units


def _invoice_date_display(group) -> str:
    raw = str(group.get("invoice_date") or "")
    if re.fullmatch(r"\d{8}", raw):
        return f"{raw[0:4]}-{raw[4:6]}-{raw[6:8]}"
    return raw


DETAIL_HEADERS = (
    ["Destination Code", "Destination Name", "Delivery Invoice No.",
     "Generated Carton", "Original Carton", "Carton Sequence",
     "Source Carton Key", "Source File", "Upload Sequence", "Source Page",
     "Delivery Note No.", "Reviewed Line IDs", "Source Row Count",
     "Src Item", "Src EAN", "Src Description", "Src Color", "Src Size",
     "Src Retail Price", "Qty",
     "Org ID", "Location Code", "Brand", "Brand Name", "Currency",
     "IMX Item Code", "PLU", "EAN", "Item Description",
     "Long Item Description", "Color Code", "Color Description",
     "Size Code", "Season", "Original Retail Price", "Discount Price"]
    + [f"Analysis Code {i:02d}" for i in range(1, 16)]
    + [f"Composition #{i}" for i in range(1, 5)]
    + ["Lookup Identifier", "Lookup Identifier Type", "Match Status",
       "Comparison Issues", "Consolidated Source Count"])

_TEXT_DETAIL_COLUMNS = {1, 3, 4, 5, 7, 11, 12, 14, 15, 17, 18, 21, 22, 26,
                        27, 28, 31, 33}


def _build_detail_sheet(ws, group, enrichment_by_line):
    data_row = _header_row(ws, 1, DETAIL_HEADERS,
                           widths=[14] * len(DETAIL_HEADERS))
    carton_seq = {m["generated_carton_number"]: m["sequence_index"]
                  for m in group["carton_mappings"]}
    carton_key = {m["generated_carton_number"]: m["source_carton_key"]
                  for m in group["carton_mappings"]}
    for line in group["prepared_lines"]:
        product = line.get("product") or {}
        source = line.get("source") or {}
        first = (line.get("sources") or [{}])[0]
        gen = line["generated_carton_number"]
        enrich = enrichment_by_line.get(
            (line.get("source_line_ids") or [None])[0], {})
        attempts = enrich.get("attempts") or []
        last_attempt = attempts[-1] if attempts else {}
        values = [
            group["destination_code"], group.get("destination_name"),
            group["delivery_invoice_number"], gen,
            line.get("original_carton_number"), carton_seq.get(gen),
            (carton_key.get(gen) or {}).get("carton_entity_id"),
            first.get("source_file"), first.get("upload_sequence"),
            first.get("source_page"), first.get("delivery_note_number"),
            ", ".join(line.get("source_line_ids") or []),
            line.get("source_rows"),
            source.get("item_code"), source.get("ean"),
            source.get("description"), source.get("color_code"),
            source.get("size_code"), source.get("retail_price"),
            line["quantity"],
            product.get("org_id"), product.get("location_code"),
            product.get("brand"), product.get("brand_name"),
            product.get("currency"), product.get("item_code"),
            product.get("plu"), product.get("ean"),
            product.get("item_desc"), product.get("long_item_desc"),
            product.get("color_code"), product.get("color_desc"),
            product.get("size_code"), product.get("season"),
            product.get("original_retail_price"),
            product.get("discount_price"),
        ] + [product.get(f"analysis_code_{i:02d}") for i in range(1, 16)] \
          + [product.get(f"composition_{i:02d}") for i in range(1, 5)] \
          + [last_attempt.get("identifier"),
             last_attempt.get("identifier_type"),
             enrich.get("status"),
             enrich.get("comparison_issue_count"),
             line.get("source_rows")]
        for col, value in enumerate(values, start=1):
            if col == 20:
                _num_cell(ws, data_row, col, value)
            elif col in (35, 36):
                cell = ws.cell(row=data_row, column=col,
                               value="" if value is None else str(value))
                cell.number_format = TEXT_FORMAT   # price strings from API
                cell.font = _CELL_FONT
            elif col in _TEXT_DETAIL_COLUMNS:
                _text_cell(ws, data_row, col, value)
            else:
                cell = ws.cell(row=data_row, column=col,
                               value="" if value is None else value)
                cell.font = _CELL_FONT
                if isinstance(value, str):
                    cell.number_format = TEXT_FORMAT
        data_row += 1
    return data_row


def _build_carton_mapping_sheet(ws, group):
    headers = ["Destination Code", "Destination Name", "Generated Carton",
               "Original Carton", "Upload Sequence", "Source File",
               "First Source Page", "Delivery Note No.",
               "Source Line Count", "Prepared Line Count", "Total Units"]
    data_row = _header_row(ws, 1, headers,
                           widths=[14, 20, 12, 12, 10, 24, 10, 26, 10, 10,
                                   10])
    prepared_by_carton: dict[str, list] = {}
    for line in group["prepared_lines"]:
        prepared_by_carton.setdefault(
            line["generated_carton_number"], []).append(line)
    total_units = 0
    for mapping in group["carton_mappings"]:      # exact Build 6 order
        gen = mapping["generated_carton_number"]
        lines = prepared_by_carton.get(gen, [])
        units = sum(l["quantity"] for l in lines)
        total_units += units
        key = mapping["source_carton_key"]
        for col, value in enumerate(
                [group["destination_code"], group.get("destination_name"),
                 gen, mapping.get("original_carton_number"),
                 key.get("upload_sequence"), key.get("source_file"),
                 key.get("first_source_page"),
                 key.get("delivery_note_number"),
                 mapping.get("line_count"), len(lines)], start=1):
            if col in (3, 4, 6, 8):
                _text_cell(ws, data_row, col, value)
            else:
                ws.cell(row=data_row, column=col,
                        value=value).font = _CELL_FONT
        _num_cell(ws, data_row, 11, units)
        data_row += 1
    _text_cell(ws, data_row, 1, "Total", bold=True)
    _num_cell(ws, data_row, 9,
              sum(m.get("line_count") or 0
                  for m in group["carton_mappings"]), bold=True)
    _num_cell(ws, data_row, 10, group["prepared_line_count"], bold=True)
    _num_cell(ws, data_row, 11, total_units, bold=True)


def _build_needs_review_sheet(ws, group, packing_issues, product_warnings):
    headers = ["Severity", "Issue Code", "Destination", "Generated Carton",
               "Original Carton", "Source File", "Source Page",
               "Delivery Note No.", "Reviewed Line ID", "Message"]
    data_row = _header_row(ws, 1, headers,
                           widths=[10, 30, 12, 12, 12, 24, 10, 26, 18, 60])
    dest = group["destination_code"]
    rows = 0
    for issue in packing_issues:
        if issue.get("destination") not in (None, dest):
            continue
        for col, value in enumerate(
                [issue.get("severity"), issue.get("code"),
                 issue.get("destination") or dest, "",
                 issue.get("original_carton_number"),
                 issue.get("source_file"), issue.get("source_page"),
                 issue.get("delivery_note_number"), issue.get("line_id"),
                 issue.get("message")], start=1):
            _text_cell(ws, data_row, col, value)
        data_row += 1
        rows += 1
    line_ids = {lid for l in group["prepared_lines"]
                for lid in (l.get("source_line_ids") or [])}
    for warning in product_warnings:
        if warning.get("line_id") and warning["line_id"] not in line_ids:
            continue
        for col, value in enumerate(
                [warning.get("severity"), warning.get("code"), dest, "", "",
                 "", "", "", warning.get("line_id"),
                 warning.get("message")], start=1):
            _text_cell(ws, data_row, col, value)
        data_row += 1
        rows += 1
    if rows == 0:
        _text_cell(ws, data_row, 1, "No unresolved review items.")


def _build_source_documents_sheet(ws, group, extraction_docs):
    headers = ["Upload Sequence", "Source File", "Page Count",
               "Delivery Note No.", "Delivery Date", "From Location",
               "To Location", "Original Carton Count",
               "Extracted Line Count", "Total Units",
               "Extraction Methods", "Extraction Issue Count",
               "In This Workbook"]
    data_row = _header_row(ws, 1, headers,
                           widths=[10, 26, 8, 28, 12, 20, 20, 10, 10, 10,
                                   22, 10, 12])
    group_files = {key.get("source_file") for key in
                   (m["source_carton_key"] for m in
                    group["carton_mappings"])}
    for doc in extraction_docs:
        header = doc.get("header") or {}
        cartons = doc.get("cartons") or []
        methods = {}
        for method in doc.get("page_methods") or []:
            methods[method] = methods.get(method, 0) + 1
        method_text = ", ".join(f"{k}: {v}" for k, v in
                                sorted(methods.items()))
        values = [doc.get("upload_sequence"), doc.get("source_file"),
                  doc.get("page_count"),
                  header.get("delivery_note_number"),
                  header.get("delivery_date"),
                  header.get("from_location_code"),
                  header.get("to_location_code"), len(cartons),
                  sum(len(c.get("lines") or []) for c in cartons),
                  doc.get("calculated_grand_total"), method_text,
                  len(doc.get("issues") or []),
                  "yes" if doc.get("source_file") in group_files else "no"]
        for col, value in enumerate(values, start=1):
            if col in (2, 4, 5, 6, 7, 11, 13):
                _text_cell(ws, data_row, col, value)
            else:
                ws.cell(row=data_row, column=col,
                        value=value).font = _CELL_FONT
        data_row += 1


# --- generation -------------------------------------------------------------------

def _audit_context(job_id: str):
    """Read-only audit context for the Needs Review / Source Documents
    sheets. Never written."""
    from apps.web.transfer import extraction as extraction_mod
    from apps.web.transfer import product_lookup as plmod
    enrichment = plmod.load_enrichment(job_id) or {}
    enrichment_by_line = {le.get("line_id"): le
                          for le in enrichment.get("line_enrichments", [])}
    product_warnings = [i for i in enrichment.get("issues", [])
                        if i.get("severity") == "warning"]
    result = extraction_mod.load_result(job_id)
    extraction_docs = ([d.as_dict() for d in result.documents]
                       if result is not None else [])
    return enrichment_by_line, product_warnings, extraction_docs


def build_destination_workbook(job_id: str, group: dict,
                               config: PackingWorkbookConfig,
                               generated_at: str,
                               audit=None) -> Workbook:
    enrichment_by_line, product_warnings, extraction_docs = (
        audit if audit is not None else _audit_context(job_id))
    prepared = pk.load_preparation(job_id) or {}
    workbook = Workbook()
    ws = workbook.active
    ws.title = SHEET_PACKING_LIST
    _build_packing_list_sheet(ws, group, config, generated_at)
    _build_detail_sheet(workbook.create_sheet(SHEET_DETAIL), group,
                        enrichment_by_line)
    _build_carton_mapping_sheet(
        workbook.create_sheet(SHEET_CARTON_MAPPING), group)
    _build_needs_review_sheet(
        workbook.create_sheet(SHEET_NEEDS_REVIEW), group,
        prepared.get("issues") or [], product_warnings)
    _build_source_documents_sheet(
        workbook.create_sheet(SHEET_SOURCE_DOCUMENTS), group,
        extraction_docs)
    return workbook


# --- validation -------------------------------------------------------------------

def validate_workbook(path: Path, group: dict) -> list[dict]:
    """Reopen and verify one generated workbook against its Build 6 group.
    Returns issue dicts (empty = valid)."""
    issues: list[dict] = []

    def issue(code, message, severity=SEV_BLOCKING):
        issues.append({"code": code, "severity": severity,
                       "destination": group.get("destination_code"),
                       "message": message})

    if not path.is_file():
        issue(WORKBOOK_FILE_MISSING, "The workbook file does not exist.")
        return issues
    if path.stat().st_size == 0:
        issue(WORKBOOK_FILE_EMPTY, "The workbook file is empty.")
        return issues
    try:
        with zipfile.ZipFile(path) as archive:
            members = archive.namelist()
            if any(name.endswith("vbaProject.bin") for name in members):
                issue(WORKBOOK_VALIDATION_FAILED,
                      "The workbook contains macros.")
            if archive.testzip() is not None:
                issue(WORKBOOK_OPEN_FAILED, "The XLSX archive is corrupt.")
                return issues
    except zipfile.BadZipFile:
        issue(WORKBOOK_OPEN_FAILED, "The file is not a valid XLSX archive.")
        return issues
    try:
        wb = load_workbook(path)
    except Exception as exc:
        issue(WORKBOOK_OPEN_FAILED,
              f"The workbook could not be reopened ({type(exc).__name__}).")
        return issues
    names = wb.sheetnames
    for required in REQUIRED_SHEETS:
        count = names.count(required)
        if count == 0:
            issue(WORKBOOK_SHEET_MISSING, f"Sheet '{required}' is missing.")
        elif count > 1:
            issue(WORKBOOK_SHEET_DUPLICATE,
                  f"Sheet '{required}' appears {count} times.")
    if getattr(wb, "_external_links", None):
        issue(WORKBOOK_EXTERNAL_LINK_FOUND,
              "The workbook contains external links.")
    if issues:
        return issues

    pl_ws = wb[SHEET_PACKING_LIST]
    header_text = " ".join(str(c.value) for row in
                           pl_ws.iter_rows(min_row=1, max_row=14)
                           for c in row if c.value is not None)
    if group["delivery_invoice_number"] not in header_text \
            or group["destination_code"] not in header_text:
        issue(WORKBOOK_METADATA_MISMATCH,
              "Destination or invoice number missing from the Packing "
              "List header.")

    mapping_ws = wb[SHEET_CARTON_MAPPING]
    mapping_rows = list(mapping_ws.iter_rows(min_row=2,
                                             values_only=True))
    carton_rows = [r for r in mapping_rows if r[2] and r[0] != "Total"
                   and r[0] is not None]
    expected_cartons = [m["generated_carton_number"]
                        for m in group["carton_mappings"]]
    got_cartons = [str(r[2]) for r in carton_rows if str(r[0]) != "Total"]
    if len(got_cartons) != len(expected_cartons):
        issue(WORKBOOK_CARTON_COUNT_MISMATCH,
              f"Carton Mapping has {len(got_cartons)} cartons; the packing "
              f"artifact has {len(expected_cartons)}.")
    elif got_cartons != expected_cartons:
        issue(WORKBOOK_CARTON_COUNT_MISMATCH,
              "Generated carton numbers differ from the Build 6 "
              "assignments.")
    originals = [m.get("original_carton_number")
                 for m in group["carton_mappings"]]
    got_originals = [r[3] if r[3] is None else str(r[3])
                     for r in carton_rows if str(r[0]) != "Total"]
    if len(got_originals) == len(originals) and any(
            (a or "") != (b or "") for a, b in zip(got_originals, originals)):
        issue(WORKBOOK_METADATA_MISMATCH,
              "Original carton numbers were not retained.")

    detail_ws = wb[SHEET_DETAIL]
    detail_rows = list(detail_ws.iter_rows(min_row=2, values_only=True))
    detail_rows = [r for r in detail_rows if r[0] is not None]
    if len(detail_rows) != group["prepared_line_count"]:
        issue(WORKBOOK_LINE_COUNT_MISMATCH,
              f"Detail has {len(detail_rows)} lines; the packing artifact "
              f"has {group['prepared_line_count']}.")
    else:
        total = sum(int(r[19] or 0) for r in detail_rows)
        if total != group["total_units"]:
            issue(WORKBOOK_TOTAL_MISMATCH,
                  f"Detail units total {total}; the packing artifact says "
                  f"{group['total_units']}.")
        expected_eans = [((l.get("product") or {}).get("ean"))
                         for l in group["prepared_lines"]]
        for row, expected in zip(detail_rows, expected_eans):
            got = row[27]
            if expected and (not isinstance(got, str)
                             or got != expected):
                issue(WORKBOOK_IDENTIFIER_FORMAT_ERROR,
                      "An EAN lost its text format or leading zeros in "
                      "Detail.")
                break
        for row in detail_rows:
            if row[3] is not None and not isinstance(row[3], str):
                issue(WORKBOOK_IDENTIFIER_FORMAT_ERROR,
                      "A generated carton number is not stored as text.")
                break
    return issues


# --- output metadata --------------------------------------------------------------

def load_output(job_id: str) -> dict | None:
    try:
        data = json.loads(
            output_meta_path(job_id).read_text(encoding="utf-8"))
    except (JobError, OSError, ValueError):
        return None
    if not isinstance(data, dict) or "job_id" not in data:
        return None
    current = packing_checksum(job_id)
    prepared = pk.load_preparation(job_id)
    data["stale"] = (current is None
                     or data.get("packing_checksum") != current
                     or prepared is None or bool(prepared.get("stale")))
    return data


def _write_output_meta(job_id: str, data: dict) -> None:
    path = output_meta_path(job_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.is_file():
        prior = load_output(job_id)
        if prior is not None and prior.get("stale"):
            stamp = utc_now().replace(":", "").replace("-", "").split(".")[0]
            target = path.with_name(f"result-stale-{stamp}.json")
            counter = 0
            while target.exists():
                counter += 1
                target = path.with_name(
                    f"result-stale-{stamp}-{counter}.json")
            os.replace(path, target)
    tmp = path.with_name(f"{OUTPUT_META_NAME}.tmp-{os.getpid()}")
    tmp.write_text(json.dumps(data, indent=2) + "\n", encoding="utf-8")
    os.replace(tmp, path)


def cleanup_expired_outputs(job_id: str,
                            config: PackingWorkbookConfig) -> int:
    """Minimum safe retention: remove THIS job's workbook/zip files older
    than the configured hours, never while generation is in progress, and
    never touching other jobs or metadata. Full transfer-job retention
    remains the documented Build 1 deferral."""
    import time as _time
    job = jobs.load_transfer_job(job_id)
    if job is None or job.status == JOB_WORKBOOK_GENERATION_IN_PROGRESS:
        return 0
    directory = output_dir(job_id)
    if not directory.is_dir():
        return 0
    removed = 0
    cutoff = _time.time() - config.retention_hours * 3600
    for path in directory.iterdir():
        if path.suffix.lower() not in (".xlsx", ".zip"):
            continue
        try:
            if path.stat().st_mtime < cutoff and not path.is_symlink():
                path.unlink()
                removed += 1
        except OSError:
            continue
    return removed


def cleanup_stale_tmp(job_id: str) -> None:
    directory = output_dir(job_id)
    if directory.is_dir():
        for path in directory.glob("*.tmp-*"):
            try:
                path.unlink()
            except OSError:
                pass


# --- the generation run -----------------------------------------------------------

def generate_workbooks(job_id: str, *,
                       config: PackingWorkbookConfig | None = None,
                       on_progress=None) -> dict:
    """Generate + validate one workbook per destination (and the ZIP when
    multiple), persist output/result.json, and set the job state."""
    config = config or load_workbook_config()
    job, prepared = load_generation_inputs(job_id)
    jobs.update_job_status(job_id, JOB_WORKBOOK_GENERATION_IN_PROGRESS)
    directory = output_dir(job_id)
    directory.mkdir(parents=True, exist_ok=True)
    cleanup_stale_tmp(job_id)
    generated_at = utc_now()
    prior = load_output(job_id)
    prior_files = set()
    if prior is not None:
        prior_files = {w.get("filename")
                       for w in prior.get("destination_workbooks", [])}
        if prior.get("zip") and prior["zip"].get("filename"):
            prior_files.add(prior["zip"]["filename"])

    groups = prepared["destinations"]
    audit = _audit_context(job_id)
    workbook_entries = []
    all_issues: list[dict] = []
    seen_filenames: set[str] = set()
    try:
        for index, group in enumerate(groups, start=1):
            if on_progress is not None:
                try:
                    on_progress(index, len(groups),
                                group["destination_code"])
                except Exception:
                    pass
            filename = workbook_filename(group["destination_code"],
                                         group["delivery_invoice_number"])
            if filename in seen_filenames:
                raise JobError(f"Duplicate workbook filename '{filename}'.")
            seen_filenames.add(filename)
            final_path = (directory / filename)
            if final_path.resolve().parent != directory.resolve():
                raise JobError("Unsafe workbook filename.")
            # openpyxl refuses to reopen non-.xlsx extensions, so the
            # temp name keeps the suffix while still matching *.tmp-*
            tmp_path = directory / f"{filename}.tmp-{os.getpid()}.xlsx"
            workbook = build_destination_workbook(job_id, group, config,
                                                  generated_at, audit=audit)
            workbook.save(tmp_path)
            issues = validate_workbook(tmp_path, group)
            if any(i["severity"] == SEV_BLOCKING for i in issues):
                tmp_path.unlink(missing_ok=True)
                all_issues.extend(issues)
                raise JobError(
                    "Workbook validation failed for destination "
                    f"{group['destination_code']}: "
                    + "; ".join(i["code"] for i in issues))
            os.replace(tmp_path, final_path)
            all_issues.extend(issues)
            workbook_entries.append({
                "destination_code": group["destination_code"],
                "destination_name": group.get("destination_name"),
                "delivery_invoice_number":
                    group["delivery_invoice_number"],
                "filename": filename,
                "relative_path": f"{OUTPUT_DIR}/{filename}",
                "sha256": _sha256(final_path),
                "byte_size": final_path.stat().st_size,
                "sheet_names": list(REQUIRED_SHEETS),
                "carton_count": group["generated_carton_count"],
                "prepared_line_count": group["prepared_line_count"],
                "total_units": group["total_units"],
                "validation_status": ("valid" if not issues
                                      else "valid_with_warnings"),
                "validation_issues": issues,
            })

        zip_entry = None
        if len(workbook_entries) > 1 and config.create_zip_for_multiple:
            zip_entry = _build_zip(job_id, directory, workbook_entries)
    except Exception as exc:
        cleanup_stale_tmp(job_id)
        failed = {
            "schema_version": OUTPUT_SCHEMA_VERSION, "job_id": job_id,
            "packing_checksum": packing_checksum(job_id),
            "created_at": generated_at, "updated_at": utc_now(),
            "status": "failed", "config": config.summary(),
            "destination_workbooks": workbook_entries,
            "zip": None,
            "issues": all_issues + [{
                "code": WORKBOOK_VALIDATION_FAILED,
                "severity": SEV_BLOCKING, "destination": None,
                "message": f"Generation failed ({type(exc).__name__}).",
            }],
            "summary": {},
        }
        _write_output_meta(job_id, failed)
        jobs.update_job_status(job_id, JOB_WORKBOOK_GENERATION_FAILED)
        raise

    # remove superseded files from a previous generation (renamed outputs)
    current_files = {w["filename"] for w in workbook_entries}
    if zip_entry:
        current_files.add(zip_entry["filename"])
    for name in sorted(prior_files - current_files):
        candidate = directory / name
        if candidate.is_file() and candidate.suffix.lower() in (".xlsx",
                                                                ".zip"):
            candidate.unlink()

    warnings = sum(1 for i in all_issues if i["severity"] == SEV_WARNING)
    meta = {
        "schema_version": OUTPUT_SCHEMA_VERSION,
        "job_id": job_id,
        "packing_checksum": packing_checksum(job_id),
        "created_at": (prior or {}).get("created_at") or generated_at,
        "updated_at": utc_now(),
        "status": "complete_with_issues" if warnings else "complete",
        "config": config.summary(),
        "destination_workbooks": workbook_entries,
        "zip": zip_entry,
        "issues": all_issues,
        "summary": {
            "workbooks": len(workbook_entries),
            "total_files": len(workbook_entries) + (1 if zip_entry else 0),
            "total_bytes": (sum(w["byte_size"] for w in workbook_entries)
                            + (zip_entry["byte_size"] if zip_entry else 0)),
            "warnings": warnings,
        },
    }
    _write_output_meta(job_id, meta)
    jobs.update_job_status(
        job_id, JOB_WORKBOOK_GENERATION_WITH_ISSUES if warnings
        else JOB_WORKBOOK_GENERATION_COMPLETE)
    return meta


def _build_zip(job_id: str, directory: Path, workbook_entries) -> dict:
    filename = zip_filename(job_id)
    final_path = directory / filename
    tmp_path = directory / f"{filename}.tmp-{os.getpid()}"
    with zipfile.ZipFile(tmp_path, "w", zipfile.ZIP_DEFLATED) as archive:
        for entry in workbook_entries:        # deterministic member order
            archive.write(directory / entry["filename"], entry["filename"])
    with zipfile.ZipFile(tmp_path) as archive:
        members = archive.namelist()
        expected = [w["filename"] for w in workbook_entries]
        if members != expected or archive.testzip() is not None:
            tmp_path.unlink(missing_ok=True)
            raise JobError("ZIP validation failed.")
    os.replace(tmp_path, final_path)
    return {"filename": filename,
            "relative_path": f"{OUTPUT_DIR}/{filename}",
            "sha256": _sha256(final_path),
            "byte_size": final_path.stat().st_size,
            "member_count": len(workbook_entries)}
